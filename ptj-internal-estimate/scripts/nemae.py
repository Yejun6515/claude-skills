# -*- coding: utf-8 -*-
"""PTJ 내부견적 見積纏め v2 빌더 (REDESIGN_SPEC ①~⑨).

견적부서 見積計算書（まとめ） 파일에 새 '見積纏め_Sales added' 시트(맨 왼쪽)를 추가하고
영업 입력값만 채운다. 계산은 시트 수식(전부 forward·비순환)이 한다 → Excel 반복계산 불필요.

모델(자세히는 REDESIGN_SPEC.md):
  §3 행분해 — 機械(COMM/insurance/bond) · SV(COMM/CIT/Local Tax Agent), 모두 受注価(L)×%.
  비순환 닫힌형: H31 = (D22 + F26 − r機械×H26)/(1 − EBIT − r機械),  r機械=COMM+insurance+bond.
  SV 受注価 H26 = D16(단가)×D11(SV MD) 고정. §5 残業·L24는 D16에 자동연동.
  시작점: まとめ その他 비면 Pure Manf Cost(D22/D26)에 製造原価. CONTI(RC)·FC 채워졌으면 §3에 안 더함.

usage:
  build  <quotation.xlsx> <out.xlsx> --main M --sv S [opts]
  verify <out.xlsx>
"""
import argparse, copy, re, sys, datetime, warnings, os, math, zipfile
import openpyxl
warnings.filterwarnings('ignore')

SKILL_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
TEMPLATE = os.path.join(SKILL_DIR, 'assets', 'nemae_template.xlsx')
EXT = re.compile(r'\[\d+\]')
TPL_SHEET = '見積纏め'
NEW_SHEET = '見積纏め_Sales added'

# 입력셀 맵 (열 레이아웃 템플릿)
C = dict(NAME='C2', DATE='P1', AUTHOR='P2', UNIT='D16', MMUNIT='C16',
         MAIN='C22', SV='C23', SPARE='C24',
         COMM='J5', INS='J6', BOND='J7', CIT='J12', LTA='J13', FC='J14', RC='J15',
         EBIT='L25', OFFER='O25', TOTL='M25', SVL='M23', MECHL='M22')
# SV indivisual Tax source = §1 C16(kJPY/MM). §2-1 J11='=C16', §3 D23='=C16*C11'(절대).

try:
    sys.stdout.reconfigure(encoding='utf-8')
    sys.stderr.reconfigure(encoding='utf-8')
except Exception:
    pass


def copy_sheet(src_ws, tgt_wb, title):
    tgt = tgt_wb.create_sheet(title)
    for k, v in src_ws.column_dimensions.items():
        if v.width is not None:
            tgt.column_dimensions[k].width = v.width
        tgt.column_dimensions[k].hidden = v.hidden
    for k, v in src_ws.row_dimensions.items():
        if v.height is not None:
            tgt.row_dimensions[k].height = v.height
    for row in src_ws.iter_rows():
        for c in row:
            nc = tgt.cell(c.row, c.column, c.value)
            if c.has_style:
                nc.font = copy.copy(c.font)
                nc.fill = copy.copy(c.fill)
                nc.border = copy.copy(c.border)
                nc.alignment = copy.copy(c.alignment)
                nc.number_format = c.number_format
                nc.protection = copy.copy(c.protection)
    for mc in src_ws.merged_cells.ranges:
        tgt.merge_cells(str(mc))
    tgt.sheet_format = copy.copy(src_ws.sheet_format)
    tgt.sheet_properties = copy.copy(src_ws.sheet_properties)
    tgt.freeze_panes = src_ws.freeze_panes
    tgt.sheet_view.showGridLines = src_ws.sheet_view.showGridLines
    return tgt


def strip_external_links(wb, wbv):
    replaced = 0
    for ws in wb.worksheets:
        vsheet = wbv[ws.title] if ws.title in wbv.sheetnames else None
        for row in ws.iter_rows():
            for c in row:
                v = c.value
                if isinstance(v, str) and v.startswith('=') and EXT.search(v):
                    c.value = vsheet[c.coordinate].value if vsheet is not None else None
                    replaced += 1
    n_ext = len(wb._external_links)
    wb._external_links = []
    return replaced, n_ext


def drop_junk_defined_names(wb):
    def junk(val):
        if val is None:
            return False
        v = str(val)
        return ('#REF!' in v) or ('#N/A' in v) or v.strip().startswith('{') or ('[' in v and ']' in v)
    removed = 0
    for nm in list(wb.defined_names.keys()):
        if junk(wb.defined_names[nm].value):
            del wb.defined_names[nm]; removed += 1
    for ws in wb.worksheets:
        try:
            keys = list(ws.defined_names.keys())
        except Exception:
            keys = []
        for nm in keys:
            if junk(ws.defined_names[nm].value):
                del ws.defined_names[nm]; removed += 1
    return removed


def num(ws, addr, default=0.0):
    v = ws[addr].value
    return float(v) if isinstance(v, (int, float)) else default


def read_sv_persons(quotation):
    """現地(Ｓ)Ｖ費 시트에서 (이름, 見積MD, 契約MD) 추출. 못 찾으면 []."""
    try:
        wb = openpyxl.load_workbook(quotation, data_only=True)
    except Exception:
        return []
    def norm(s):
        return s.replace('Ｓ', 'S').replace('Ｖ', 'V') if isinstance(s, str) else s
    sname = next((s for s in wb.sheetnames if '現地SV' in norm(s) or 'SV費' in norm(s)), None)
    if not sname:
        return []
    ws = wb[sname]
    # 헤더로 見積MD/契約MD 열 찾기 (일반형 '見積'+'MD' 또는 工程表형 '移動日含み'/'実労働日')
    col_est = col_con = None
    for row in ws.iter_rows(min_row=1, max_row=6):
        for c in row:
            v = c.value
            if not isinstance(v, str):
                continue
            if col_est is None and (('見積' in v and 'MD' in v) or '移動日含み' in v):
                col_est = c.column
            if col_con is None and (('契約' in v and 'MD' in v) or '実労働日' in v):
                col_con = c.column
    if col_est is None:
        return []
    persons = []
    for r in range(1, ws.max_row + 1):
        nm = ws.cell(r, 1).value
        if isinstance(nm, str) and ('合計' in nm or '計' == nm.strip()):
            break
        est = ws.cell(r, col_est).value
        if isinstance(est, (int, float)) and est > 0:
            con = ws.cell(r, col_con).value if col_con else est
            con = float(con) if isinstance(con, (int, float)) else float(est)
            persons.append((str(nm).strip() if nm else '', float(est), con))
    return persons


def fill_sv_md(ws, persons, svdays):
    """§4 표(rows 36-41): 実働(D)=契約MD, 休日(E)=0, 移動(F)=見積−契約. 計(G)=SUM 자동→G42→D7."""
    total = 0
    if persons:
        for i in range(6):
            r = 29 + i                          # §4 표 rows 29-34
            if i < len(persons):
                nm, est, con = persons[i]
                ws['D%d' % r] = round(con); ws['E%d' % r] = 0; ws['F%d' % r] = round(est - con)
                total += round(est)
            else:
                ws['D%d' % r] = None; ws['E%d' % r] = None; ws['F%d' % r] = None
    elif svdays is not None:
        ws['D29'] = svdays                      # 実働에 計=svdays
        total = svdays
    return total


def solve(main_base, sv_base, spare_base, unit, svmd, comm, ins, bond, cit, lta, ebit,
          sv_indiv_tax=0.0, fc=0.0, daypm=30.5):
    """비순환 닫힌형 — 시트 수식과 동일 결과를 Python으로 산출(리포트·검증용).

    SV indiv Tax = sv_indiv_tax(kJPY/MM) × Man-Months(=svmd/daypm), 절대금액(L×% 아님).
    """
    L_sv = unit * svmd
    svmm = svmd / daypm if daypm else 0
    indiv_tax = sv_indiv_tax * svmm
    r_mech = comm + ins + bond + fc       # 機械 受注価×율: COMM+ins+bond+FC (시트 G22+J22)
    r_sv = comm + cit + lta + fc          # SV 受注価×율: COMM+CIT+LTA+FC (시트 G23+E23+F23+J23)
    J_sv = sv_base + indiv_tax + r_sv * L_sv
    denom = 1 - ebit - r_mech
    L_total = (main_base + J_sv - r_mech * L_sv) / denom if denom else float('nan')
    L_mech = L_total - L_sv
    J_mech = main_base + r_mech * L_mech
    J_total = J_mech + J_sv + spare_base
    return dict(L_sv=L_sv, indiv_tax=indiv_tax, r_mech=r_mech, r_sv=r_sv, J_sv=J_sv,
                L_total=L_total, L_mech=L_mech, J_mech=J_mech, J_total=J_total)


WS_CT = 'application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml'
WS_REL = 'http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet'


def _esc(s):
    return s.replace('&', '&amp;').replace('<', '&lt;').replace('>', '&gt;')


def _ss_list(xml):
    out = []
    for si in re.findall(r'<si>(.*?)</si>', xml, re.S):
        out.append(''.join(re.findall(r'<t[^>]*>(.*?)</t>', si, re.S)))
    return out


def _count(xml, tag):
    m = re.search(r'<%s count="(\d+)"' % tag, xml)
    return int(m.group(1)) if m else 0


def _inner(xml, tag):
    m = re.search(r'<%s count="\d+"[^>]*>(.*?)</%s>' % (tag, tag), xml, re.S)
    return m.group(1) if m else ''


def _append_section(xml, tag, add_inner, add_count):
    """orig styles.xml의 <tag>에 add_inner 추가(count+=). 없으면 <styleSheet> 뒤 생성."""
    if not add_count:
        return xml
    m = re.search(r'(<%s count=")(\d+)("[^>]*>)(.*?)(</%s>)' % (tag, tag), xml, re.S)
    if m:
        rep = m.group(1) + str(int(m.group(2)) + add_count) + m.group(3) + m.group(4) + add_inner + m.group(5)
        return xml[:m.start()] + rep + xml[m.end():]
    block = '<%s count="%d">%s</%s>' % (tag, add_count, add_inner, tag)
    m = re.search(r'<%s count="\d+"\s*/>' % tag, xml)        # self-closing 빈 섹션
    if m:
        return xml[:m.start()] + block + xml[m.end():]
    m = re.search(r'<styleSheet[^>]*>', xml)
    return xml[:m.end()] + block + xml[m.end():]


def inject_sheet(orig_path, sheet_src_path, out_path, sheet_name):
    """원본 xlsx ZIP은 그대로 두고 새 워크시트 1장만 끼워넣는다(맨 왼쪽).
    openpyxl 통째 재저장이 도형/이미지/컨트롤을 떨궈 'Excel 복구' 경고를 내는 문제를 회피한다.
    스타일(font/fill/border/numFmt/cellXfs)은 원본 styles.xml에 offset-append, 문자열은 inline화.
    """
    oz = zipfile.ZipFile(orig_path)
    parts = {n: oz.read(n) for n in oz.namelist()}
    oz.close()
    sz = zipfile.ZipFile(sheet_src_path)
    src_sheet = sz.read('xl/worksheets/sheet1.xml').decode('utf-8')
    src_styles = sz.read('xl/styles.xml').decode('utf-8')
    ss_list = _ss_list(sz.read('xl/sharedStrings.xml').decode('utf-8')) if 'xl/sharedStrings.xml' in sz.namelist() else []
    sz.close()

    styles = parts['xl/styles.xml'].decode('utf-8')
    ofont, ofill, obord, ocxf = (_count(styles, t) for t in ('fonts', 'fills', 'borders', 'cellXfs'))
    # 커스텀 numFmt 재번호 (builtin<164 유지)
    onf = [int(x) for x in re.findall(r'<numFmt numFmtId="(\d+)"', styles)]
    maxnf = max(onf) if onf else 163
    nfmap = {}; new_nf = []
    for sid, code in re.findall(r'<numFmt numFmtId="(\d+)" formatCode="([^"]*)"\s*/>', src_styles):
        maxnf += 1; nfmap[int(sid)] = maxnf
        new_nf.append('<numFmt numFmtId="%d" formatCode="%s"/>' % (maxnf, code))

    def remap_xf(inner):
        inner = re.sub(r'fontId="(\d+)"', lambda m: 'fontId="%d"' % (int(m.group(1)) + ofont), inner)
        inner = re.sub(r'fillId="(\d+)"', lambda m: 'fillId="%d"' % (int(m.group(1)) + ofill), inner)
        inner = re.sub(r'borderId="(\d+)"', lambda m: 'borderId="%d"' % (int(m.group(1)) + obord), inner)
        inner = re.sub(r'numFmtId="(\d+)"', lambda m: 'numFmtId="%d"' % nfmap.get(int(m.group(1)), int(m.group(1))), inner)
        return inner

    styles = _append_section(styles, 'numFmts', ''.join(new_nf), len(new_nf))
    styles = _append_section(styles, 'fonts', _inner(src_styles, 'fonts'), _count(src_styles, 'fonts'))
    styles = _append_section(styles, 'fills', _inner(src_styles, 'fills'), _count(src_styles, 'fills'))
    styles = _append_section(styles, 'borders', _inner(src_styles, 'borders'), _count(src_styles, 'borders'))
    styles = _append_section(styles, 'cellXfs', remap_xf(_inner(src_styles, 'cellXfs')), _count(src_styles, 'cellXfs'))
    parts['xl/styles.xml'] = styles.encode('utf-8')

    # 새 시트: 공유문자열→inline, 스타일 인덱스 offset, rels 필요 요소 제거
    def cellrepl(m):
        attrs, inner = m.group(1), m.group(2)
        if 't="s"' in attrs:
            vi = re.search(r'<v>(\d+)</v>', inner)
            s = ss_list[int(vi.group(1))] if vi and int(vi.group(1)) < len(ss_list) else ''
            return '<c%s><is><t xml:space="preserve">%s</t></is></c>' % (attrs.replace('t="s"', 't="inlineStr"'), _esc(s))
        return m.group(0)
    sheet = re.sub(r'<c([^>]*)>(.*?)</c>', cellrepl, src_sheet, flags=re.S)
    sheet = re.sub(r'\bs="(\d+)"', lambda m: 's="%d"' % (int(m.group(1)) + ocxf), sheet)
    sheet = re.sub(r'\bstyle="(\d+)"', lambda m: 'style="%d"' % (int(m.group(1)) + ocxf), sheet)
    for tag in ('drawing', 'legacyDrawing', 'picture', 'tableParts', 'oleObjects', 'controls', 'extLst'):
        sheet = re.sub(r'<%s\b[^>]*/>' % tag, '', sheet)
        sheet = re.sub(r'<%s\b.*?</%s>' % (tag, tag), '', sheet, flags=re.S)
    sheet = re.sub(r'\sr:id="[^"]*"', '', sheet)
    # ⚠️ 빈 캐시값 <v></v>/<v/> 제거 — openpyxl이 미계산 수식에 빈 <v>를 넣는데,
    #    Excel은 이를 잘못된 값으로 보고 '수식 레코드 제거(복구경고)'를 띄운다. 빼면 Excel이 열 때 계산.
    sheet = re.sub(r'<v\s*/>|<v>\s*</v>', '', sheet)

    # calcChain 제거(시트 추가로 인덱스 어긋남 방지 — Excel 재생성)
    for p in [p for p in parts if p.endswith('calcChain.xml')]:
        del parts[p]
    ct = parts['[Content_Types].xml'].decode('utf-8')
    ct = re.sub(r'<Override PartName="/xl/calcChain\.xml"[^>]*/>', '', ct)
    rels = parts['xl/_rels/workbook.xml.rels'].decode('utf-8')
    rels = re.sub(r'<Relationship[^>]*calcChain\.xml[^>]*/>', '', rels)

    # 등록: workbook.xml(맨앞 sheet), rels, content-types, 새 part
    wbxml = parts['xl/workbook.xml'].decode('utf-8')
    # ⚠️ 새 시트를 맨 앞(index 0)에 넣으면 기존 시트 인덱스가 +1 밀린다.
    #    정의이름 localSheetId·workbookView activeTab(0-based 시트번호)을 +1 보정해야 Excel 복구경고가 안 뜬다.
    wbxml = re.sub(r'localSheetId="(\d+)"', lambda m: 'localSheetId="%d"' % (int(m.group(1)) + 1), wbxml)
    wbxml = re.sub(r'(activeTab=")(\d+)(")', lambda m: '%s%d%s' % (m.group(1), int(m.group(2)) + 1, m.group(3)), wbxml)
    newsid = max(int(x) for x in re.findall(r'sheetId="(\d+)"', wbxml)) + 1
    newrid = 'rId%d' % (max(int(x) for x in re.findall(r'Id="rId(\d+)"', rels)) + 1)
    nums = [int(x) for x in re.findall(r'worksheets/sheet(\d+)\.xml', ' '.join(parts.keys()))]
    partnum = (max(nums) if nums else 0) + 1
    partname = 'xl/worksheets/sheet%d.xml' % partnum
    wbxml = re.sub(r'(<sheets[^>]*>)', r'\1<sheet name="%s" sheetId="%d" r:id="%s"/>'
                   % (_esc(sheet_name), newsid, newrid), wbxml, count=1)
    rels = rels.replace('</Relationships>', '<Relationship Id="%s" Type="%s" Target="worksheets/sheet%d.xml"/></Relationships>'
                        % (newrid, WS_REL, partnum))
    ct = ct.replace('</Types>', '<Override PartName="/%s" ContentType="%s"/></Types>' % (partname, WS_CT))

    parts['xl/workbook.xml'] = wbxml.encode('utf-8')
    parts['xl/_rels/workbook.xml.rels'] = rels.encode('utf-8')
    parts['[Content_Types].xml'] = ct.encode('utf-8')
    parts[partname] = sheet.encode('utf-8')

    with zipfile.ZipFile(out_path, 'w', zipfile.ZIP_DEFLATED) as z:
        for n, data in parts.items():
            z.writestr(n, data)


def cmd_build(a):
    wbv = openpyxl.load_workbook(a.quotation, data_only=True)
    mato = next((s for s in wbv.sheetnames if '見積計算書' in s and 'まとめ' in s
                 and 'リンク' not in s and '詳細' not in s), None)

    # 새 시트는 템플릿(단일시트)에서 만든다 → 채운 뒤 원본 ZIP에 주입(원본 시트 불변)
    twb = openpyxl.load_workbook(TEMPLATE)
    ws = twb[TPL_SHEET]
    ws.title = NEW_SHEET

    ws[C['DATE']] = a.date or datetime.date.today().strftime('%Y%m%d')
    if a.author: ws[C['AUTHOR']] = a.author
    if a.name:   ws[C['NAME']] = a.name
    ws[C['UNIT']] = a.unit_price
    ws[C['MMUNIT']] = a.sv_indiv_tax

    # §3 base는 まとめ 셀 링크(내부참조). 값 하드코딩 금지.
    def resolve(cell_arg, num_arg):
        if cell_arg and mato:
            v = wbv[mato][cell_arg].value if mato in wbv.sheetnames else None
            v = float(v) if isinstance(v, (int, float)) else (num_arg or 0.0)
            return "='%s'!%s" % (mato, cell_arg), v
        return (num_arg or 0.0), (num_arg or 0.0)
    main_cell, main_val = resolve(a.main_cell, a.main)
    sv_cell, sv_val = resolve(a.sv_cell, a.sv)
    spare_cell, spare_val = resolve(a.spare_cell, a.spare)
    ws[C['MAIN']] = main_cell; ws[C['SV']] = sv_cell; ws[C['SPARE']] = spare_cell
    ws[C['COMM']] = a.comm; ws[C['INS']] = a.insurance; ws[C['BOND']] = a.bond
    ws[C['CIT']] = a.cit; ws[C['LTA']] = a.lta; ws[C['FC']] = a.fc; ws[C['RC']] = a.rc
    ws[C['EBIT']] = a.ebit

    persons = read_sv_persons(a.quotation)
    svmd = fill_sv_md(ws, persons, a.svdays)
    if not svmd:
        sys.exit('ERROR: SV MD를 못 구했습니다. --svdays 로 SV 計MD를 지정하세요.')

    r = solve(main_val, sv_val, spare_val, a.unit_price, svmd,
              a.comm, a.insurance, a.bond, a.cit, a.lta, a.ebit, a.sv_indiv_tax, a.fc)
    offer = round(a.offer) if a.offer is not None else int(math.ceil(r['L_total'] / 100.0) * 100)
    ws[C['OFFER']] = offer

    tmp = a.out + '.__tmp__.xlsx'
    twb.save(tmp)
    try:
        inject_sheet(a.quotation, tmp, a.out, NEW_SHEET)
    finally:
        if os.path.exists(tmp):
            os.remove(tmp)

    print('SAVED:', a.out, '(원본 ZIP 보존 + 시트 주입)')
    print('  機械 base=%s(%s)  SV base=%s(%s)  단가=%s/c-md  SV計MD=%s'
          % (round(main_val), main_cell, round(sv_val), sv_cell, a.unit_price, svmd))
    print('  율: COMM=%.3g insur=%.3g bond=%.3g | SVindivTax=%g k/MM CIT=%.3g LocalTaxAgent=%.3g | FC=%.3g RC=%.3g'
          % (a.comm, a.insurance, a.bond, a.sv_indiv_tax, a.cit, a.lta, a.fc, a.rc))
    print('  SV indiv Tax(절대)=%s  EBIT=%.1f%%  SV受注価=%s  Total受注価=%s  Offer=%s'
          % (round(r['indiv_tax']), a.ebit * 100, round(r['L_sv']), round(r['L_total']), offer))
    print('  → 機械 cost=%s/受注価=%s  SV cost=%s/受注価=%s  Total cost=%s'
          % (round(r['J_mech']), round(r['L_mech']), round(r['J_sv']), round(r['L_sv']), round(r['J_total'])))
    print('  SV인원자동=%d  (drawing/이미지/컨트롤 등 원본 파트 그대로 보존)' % len(persons))


def cmd_verify(a):
    z = zipfile.ZipFile(a.file)
    names = z.namelist()
    ext = [n for n in names if 'externalLink' in n]
    draw = [n for n in names if 'drawings/' in n or n.startswith('xl/media/')]
    wb = openpyxl.load_workbook(a.file)
    ws = wb[NEW_SHEET] if NEW_SHEET in wb.sheetnames else None
    ref_nemae = 0
    if ws:
        ref_nemae = sum(1 for row in ws.iter_rows() for c in row
                        if isinstance(c.value, str) and '#REF!' in c.value)
    print('=== verify:', a.file, '===')
    print(' [info] externalLink %d, drawing/media %d 파트 (원본 보존 — 복구경고 방지)' % (len(ext), len(draw)))
    print(' [%s] %s 시트 존재' % ('OK' if ws else 'FAIL', NEW_SHEET))
    print(' [%s] %s 시트 #REF!: %d' % ('OK' if ref_nemae == 0 else 'FAIL', NEW_SHEET, ref_nemae))
    if not ws:
        return
    for k, lbl in [('NAME', '案件名'), ('MAIN', '機械base'), ('SV', 'SVbase'),
                   ('EBIT', 'EBIT'), ('OFFER', 'Offer'), ('UNIT', '単価')]:
        v = ws[C[k]].value
        print('   [%s] %s %s = %r' % ('OK' if v not in (None, '') else '확인', C[k], lbl, v))
    # 순환참조 점검: H31(닫힌형)이 자기참조 안 하는지 + 반복계산 OFF
    tot = ws[C['TOTL']].value
    iterate = getattr(wb.calculation, 'iterate', None)
    selfref = isinstance(tot, str) and (C['TOTL'] in tot.replace('$', ''))
    print('   [%s] %s 닫힌형 비자기참조 (반복계산 iterate=%r)' % ('FAIL' if selfref else 'OK', C['TOTL'], iterate))
    # §1 D7=§4計, §3 SV受注価=단가×MD 링크 확인
    print('   [info] D7=%r (←§4 計)  M23=%r (←단가×MD)  J11=%r (SVindivTax←C16)  C39=%r (←§5 D16연동)'
          % (ws['D7'].value, ws[C['SVL']].value, ws['J11'].value, ws['C39'].value))
    print('   [note] 모든 値은 forward 수식 → Excel Enable Editing 후 표시(반복계산 불필요)')


def main():
    p = argparse.ArgumentParser()
    sub = p.add_subparsers(dest='cmd', required=True)
    b = sub.add_parser('build')
    b.add_argument('quotation'); b.add_argument('out')
    b.add_argument('--name'); b.add_argument('--date'); b.add_argument('--author', default='Kim YJ')
    b.add_argument('--main', type=float, default=None, help='機械 Pure Manf Cost 수치(--main-cell 없을때)')
    b.add_argument('--sv', type=float, default=None, help='SV Pure Manf Cost 수치(--sv-cell 없을때)')
    b.add_argument('--spare', type=float, default=0.0, help='推奨予備品 base 수치')
    b.add_argument('--main-cell', dest='main_cell', help='機械 base를 まとめ 셀 링크로 (예 F36)')
    b.add_argument('--sv-cell', dest='sv_cell', help='SV base를 まとめ 셀 링크로 (예 G36)')
    b.add_argument('--spare-cell', dest='spare_cell', help='予備品 base를 まとめ 셀 링크로')
    b.add_argument('--unit-price', type=float, default=220.0, dest='unit_price', help='SV c-md 단가 D16 (kJPY)')
    b.add_argument('--comm', type=float, default=0.05, help='PTKorea COMM 율 J5 (機械+SV)')
    b.add_argument('--insurance', type=float, default=0.005, help='insurance 율 J6 (機械)')
    b.add_argument('--bond', type=float, default=0.005, help='bond charge 율 J7 (機械)')
    b.add_argument('--sv-indiv-tax', type=float, default=120.0, dest='sv_indiv_tax',
                   help='SV indivisual Tax 단가 C16 (kJPY/MM). §3 indiv.Tax = 이값×Man-Months (절대)')
    b.add_argument('--cit', type=float, default=0.10, help='TAX(CIT) 율 J9 (SV)')
    b.add_argument('--lta', type=float, default=0.02, help='Local Tax Agent Fee 율 J10 (SV)')
    b.add_argument('--fc', type=float, default=0.12, help='FC(Function cost) 율 J14 (기본 12%)')
    b.add_argument('--rc', type=float, default=0.0, help='RC(Risk Conti) 율 J11')
    b.add_argument('--ebit', type=float, default=0.03, help='EBIT G31')
    b.add_argument('--offer', type=float, default=None, help='최종 Offer J31. 미지정시 受注価 라운드업')
    b.add_argument('--svdays', type=int, default=None, help='SV 計MD 수동(現地SV費 자동참조 실패시)')
    b.set_defaults(func=cmd_build)
    v = sub.add_parser('verify'); v.add_argument('file'); v.set_defaults(func=cmd_verify)
    a = p.parse_args()
    a.func(a)


if __name__ == '__main__':
    main()
