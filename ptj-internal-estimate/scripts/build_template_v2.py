# -*- coding: utf-8 -*-
"""見積纏め 템플릿 생성기 (열 레이아웃 복원판).

- §3 = 원래 열 구조: Pure Manf|indv.Tax|CIT|Local Tax Agent|others|RC|Manf.Cost|FC|Total cost|EBIT|TC+EBIT|nego|Offer|Remarks
- 입력셀만 녹색(C6EFCE), 나머지는 테두리만.
- 비순환 닫힌형 M25(총受注価). §4 우측표(MAIN ITEM/Proportion/Submit/price/ea) 없음.
- SV indiv Tax = §1 C16(kJPY/MM)×C11(절대). §5 残業·SV売値는 D16 연동.
실행: python build_template_v2.py
"""
import os, sys
import openpyxl
from openpyxl.styles import Font, Border, Side, Alignment, PatternFill

sys.stdout.reconfigure(encoding='utf-8')
SKILL = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(SKILL, 'assets', 'nemae_template.xlsx')
SHEET = '見積纏め'

FNAME = 'Meiryo UI'                               # 사용자 지정 폰트
thin = Side(style='thin', color='808080')
BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
BASE = Font(name=FNAME, size=11)
BOLD = Font(name=FNAME, size=11, bold=True)
TITLE = Font(name=FNAME, size=12, bold=True)
WRAP = Alignment(wrap_text=True, vertical='center', horizontal='center')
CEN = Alignment(horizontal='center', vertical='center')
IN = PatternFill('solid', fgColor='C6EFCE')      # 입력셀 녹색
PCT = '0.0%'; NUM = '#,##0'; NUM1 = '#,##0.0'

wb = openpyxl.Workbook(); ws = wb.active; ws.title = SHEET
ws.sheet_view.showGridLines = True       # grid(줄무늬) 표시 — 사용자 선호


def S(addr, val, *, font=None, fmt=None, border=False, align=None, inp=False):
    c = ws[addr]; c.value = val
    c.font = font or BASE                          # 모든 셀 Meiryo UI 11 (지정시 그 폰트)
    if fmt: c.number_format = fmt
    if border: c.border = BORDER
    if align: c.alignment = align
    if inp: c.fill = IN; c.border = BORDER          # 입력셀 = 녹색 + 테두리
    return c


# 타이틀
S('B1', '見積纏め  (Sales added)', font=TITLE)
S('B2', '案件名', font=BOLD); S('C2', None, inp=True)
S('P1', None); S('P2', None)

# ── §1. SV md ──────────────────────────────────────────────────────────────
S('B3', '1. S/V md for PTJ use only', font=BOLD)
S('C4', 'Man-Month', font=BOLD, align=CEN); S('D4', 'Man-Calendar day', font=BOLD, align=CEN)
S('C5', 'Spec.'); S('D5', 30.5); S('E5', 'day/month')
S('B6', 'RCM'); S('D6', 'C-md', align=CEN); S('E6', 'W-md', align=CEN)
S('B7', 'SV'); S('C7', '=D7/D5', fmt=NUM1); S('D7', '=G36', fmt=NUM)   # D7 = §4 計
S('B8', 'EA'); S('C8', '=D8/D5', fmt=NUM1); S('D8', 0, fmt=NUM)
S('B11', '合計', font=BOLD); S('C11', '=SUM(C7:C10)', fmt=NUM1); S('D11', '=SUM(D7:D10)', fmt=NUM)
S('B16', '単価', font=BOLD)
S('C16', 120, inp=True, fmt=NUM); S('D16', 220, inp=True, fmt=NUM)
S('E16', 'C16=indiv tax/MM, D16=kJPY/c-md')
S('C17', 'indivisual tax/MM'); S('D17', 'kJPY/c-md'); S('E17', 'kJPY/w-md')

# ── §2-1. Cost/Expenses (cols I-K) ──────────────────────────────────────────
S('I3', '2-1. Cost/Expenses', font=BOLD); S('J3', 'PTJ', font=BOLD, align=CEN); S('K3', 'remarks', font=BOLD)
S('I5', 'PTKorea COMM'); S('J5', 0.05, inp=True, fmt=PCT); S('K5', 'PTKorea (機械+SV)')
S('I6', 'insurance'); S('J6', 0.005, inp=True, fmt=PCT); S('K6', '機械 only')
S('I7', 'bond charge'); S('J7', 0.005, inp=True, fmt=PCT); S('K7', '機械 only')
S('I8', ' total (others)', font=BOLD); S('J8', '=SUM(J5:J7)', fmt=PCT)            # 표시용(입력X)
S('I10', 'TAX', font=BOLD)
S('I11', 'SV indivisual Tax'); S('J11', '=C16', fmt=NUM); S('K11', 'SV  kJPY/MM × Man-Month')
S('I12', 'TAX(CIT)'); S('J12', 0.10, inp=True, fmt=PCT); S('K12', 'SV only')
S('I13', 'Local Tax Agent Fee'); S('J13', 0.02, inp=True, fmt=PCT); S('K13', 'SV only')
S('I14', 'FC (Function cost)'); S('J14', 0.12, inp=True, fmt=PCT); S('K14', '기본 12% (= まとめ R&D+M&S+G&A 있으면 0)')
S('I15', 'RC (Risk Conti)'); S('J15', 0.0, inp=True, fmt=PCT); S('K15', '= まとめ CONTI(WA/EX/PS)')
S('I16', 'EBIT'); S('J16', 'see §3')

# ── §2-2. Exchange Rate (cols M-O) ──────────────────────────────────────────
S('M3', '2-2. Exchange Rate', font=BOLD)
for _lbl, _r in [('USD/JPY', 4), ('Euro/JPY', 5), ('USD/INR', 6), ('Euro/USD', 7)]:
    S('M%d' % _r, _lbl, border=True); S('N%d' % _r, None, inp=True, fmt='0.00')

# ── §3. Cost calculation (열 레이아웃) rows 20-25 ────────────────────────────
# 열: C PureManf | D indvTax | E CIT | F LocalTaxAgent | G others | H RC | I ManfCost | J FC
#     K TotalCost | L EBIT | M TC+EBIT(w/o nego) | N nego | O Offer | P Remarks
S('B20', '3. Cost calculation（本体）', font=BOLD)
hdr = {'C': 'Pure Manf.\nCost', 'D': 'indv. Tax', 'E': 'CIT', 'F': 'Local Tax\nAgent',
       'G': 'others', 'H': 'RC', 'I': 'Manf. Cost', 'J': 'FC', 'K': 'Total cost',
       'L': 'EBIT', 'M': 'TotalCost+EBIT\n(w/o nego)', 'N': 'nego margin',
       'O': '(2)\nOffer price', 'P': 'Remarks'}
for col, t in hdr.items():
    S('%s21' % col, t, font=BOLD, align=WRAP)
allcol = 'CDEFGHIJKLMNOP'

def shell(r, label):                       # 행 테두리 + 라벨
    S('B%d' % r, label, font=BOLD, border=True)
    for col in allcol:
        ws['%s%d' % (col, r)].border = BORDER

# 機械 row 22
shell(22, '機械')
S('C22', 0, inp=True, fmt=NUM)                             # Pure Manf (빌드시 まとめ 링크/값으로 덮음)
S('G22', '=($J$5+$J$6+$J$7)*M22', fmt=NUM)                 # others = (COMM+ins+bond)×受注価
S('I22', '=SUM(C22:H22)', fmt=NUM)                         # Manf.Cost
S('J22', '=$J$14*M22', fmt=NUM)                            # FC = FC%×受注価
S('K22', '=I22+J22', fmt=NUM)                              # Total cost
S('L22', '=(M22-K22)/M22', fmt=PCT)                        # EBIT(결과)
S('M22', '=M25-M23-M24', fmt=NUM)                          # 受注価 = 총 − SV − spare (잔차)
S('N22', '=(O22-M22)/O22', fmt=PCT)                        # nego
S('O22', '=O25-O23-O24', fmt=NUM)                          # Offer

# SV row 23
shell(23, 'SV')
S('C23', 0, inp=True, fmt=NUM)
S('D23', '=C16*C11', fmt=NUM)                              # indiv.Tax(절대)=단가×Man-Months
S('E23', '=$J$12*M23', fmt=NUM)                            # CIT = CIT%×受注価
S('F23', '=$J$13*M23', fmt=NUM)                            # Local Tax Agent = %×受注価
S('G23', '=$J$5*M23', fmt=NUM)                             # others = COMM만(SV)
S('I23', '=SUM(C23:H23)', fmt=NUM)
S('J23', '=$J$14*M23', fmt=NUM)
S('K23', '=I23+J23', fmt=NUM)
S('L23', '=(M23-K23)/M23', fmt=PCT)
S('M23', '=D16*D11', fmt=NUM)                              # SV 受注価 = 단가×MD (고정)
S('N23', '=(O23-M23)/O23', fmt=PCT)
S('O23', '=M23', fmt=NUM)                                  # SV Offer = 受注価(nego 0)

# 推奨予備品 row 24
shell(24, '推奨予備品')
S('C24', 0, inp=True, fmt=NUM)
S('I24', '=SUM(C24:H24)', fmt=NUM); S('K24', '=I24', fmt=NUM)
S('M24', 0, fmt=NUM); S('O24', '=M24', fmt=NUM)

# Total row 25  (L25=EBIT 입력, M25=비순환 닫힌형, O25=Offer 입력)
shell(25, '  Total')
for col in 'CDEFGHIJK':
    S('%s25' % col, '=SUM(%s22:%s24)' % (col, col), fmt=NUM)
S('L25', 0.08, inp=True, fmt=PCT)                          # EBIT 입력
# rM(機械율)=COMM+ins+bond+FC.  M25 = (C22 + K23 + K24 − rM×(M23+M24)) / (1 − EBIT − rM)
S('M25', '=(C22+K23+K24-($J$5+$J$6+$J$7+$J$14)*(M23+M24))/(1-L25-($J$5+$J$6+$J$7+$J$14))', fmt=NUM)
S('N25', '=(O25-M25)/O25', fmt=PCT)
S('O25', None, inp=True, fmt=NUM)                          # 최종 Offer 입력

# ── §4. SV MD count rows 27-36 (우측표 없음) ────────────────────────────────
S('B27', '4. SV MD count（移動含み=見積MD, 実働=契約MD）', font=BOLD)
for col, t in [('C', 'week'), ('D', '実働'), ('E', '休日'), ('F', '移動'), ('G', '計(C-MD)')]:
    S('%s28' % col, t, font=BOLD, align=WRAP, border=True)
for i in range(6):
    r = 29 + i
    S('B%d' % r, 'SV %s' % chr(65 + i), border=True)
    for col in 'DEF': ws['%s%d' % (col, r)].fill = IN     # MD 입력 녹색
    for col in 'CDEF': ws['%s%d' % (col, r)].border = BORDER
    S('G%d' % r, '=SUM(D%d:F%d)' % (r, r), border=True, fmt=NUM)
S('B36', '合計', font=BOLD, border=True)
for col in 'CDEF': ws['%s36' % col].border = BORDER
S('G36', '=SUM(G29:G34)', font=BOLD, border=True, fmt=NUM)

# ── §5. 残業 rows 38-44 — D16 연동 ──────────────────────────────────────────
S('B38', '5. 残業 (overtime)', font=BOLD); S('C38', 'Offer', font=BOLD, align=CEN); S('D38', 'for calc', font=BOLD, align=CEN)
S('B39', '月～金  C-md'); S('C39', '=D16*1000', fmt=NUM); S('D39', '=C39/8', fmt=NUM); S('E39', '/hour')
for r, lbl, mul in [(40, '月～金  17時～22時', 1.25), (41, '月～金  22時以降', 1.5),
                    (42, '土日  8-17時', 1.5), (43, '土日  17時～22時', 1.75), (44, '土日  22時以降', 2)]:
    S('B%d' % r, lbl); S('C%d' % r, '=ROUNDUP(D%d,-3)' % r, fmt=NUM)
    S('D%d' % r, '=$D$39*E%d' % r, fmt=NUM); S('E%d' % r, mul, fmt='0.00')

# 모든 섹션 표 전체 보더(빈 칸 포함) — 사용자 요청 "1번부터 전부 칸 만들어"
# (r1,r2,c1,c2): §1 B4:D16 / §2-1 I5:K16 / §2-2 M4:N7 / §3 B21:P25 / §4 B28:G36 / §5 B38:E44
for r1, r2, c1, c2 in [(4, 16, 2, 4), (5, 16, 9, 11), (4, 7, 13, 14),
                       (21, 25, 2, 16), (28, 36, 2, 7), (38, 44, 2, 5)]:
    for rr in range(r1, r2 + 1):
        for cc in range(c1, c2 + 1):
            ws.cell(rr, cc).border = BORDER

widths = {'A': 2, 'B': 16, 'C': 14.6, 'D': 20.4, 'E': 9, 'F': 10.9, 'G': 11, 'H': 8,
          'I': 17, 'J': 15.4, 'K': 15.1, 'L': 9, 'M': 26.1, 'N': 11, 'O': 12, 'P': 12}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 16.5
ws.row_dimensions[21].height = 31.5               # §3 헤더(줄바꿈)

wb.save(OUT)
print('SAVED template (column layout + green inputs):', OUT)
