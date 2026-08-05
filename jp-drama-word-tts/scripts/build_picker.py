# -*- coding: utf-8 -*-
r"""분석완성.xlsx B열 → TTS 단어 피커 HTML 생성.

사용법: python build_picker.py <화 폴더> <제목>
  예:   python build_picker.py "...\260712_쿠조1" 쿠조1
- 입력: {폴더}\{제목}_분석완성.xlsx (B열: 단어(읽기) - 뜻, ...)
- 단어장: 볼트 15.10. 일본어\_tts단어장.csv (없으면 생성) — 등록 단어는 회색·선택불가
- 출력: {폴더}\단어피커_{제목}.html
"""
import html
import os
import re
import sys

sys.path.insert(0, os.path.expanduser(r"~\.claude\skills\_config"))
from vocab_db import load_done_words, vocab_csv_path  # noqa: E402

EP_DIR = sys.argv[1]
TITLE = sys.argv[2]
XLSX = os.path.join(EP_DIR, f"{TITLE}_분석완성.xlsx")
DB_CSV = vocab_csv_path()
OUT = os.path.join(EP_DIR, f"단어피커_{TITLE}.html")

import openpyxl  # noqa: E402

done_words = load_done_words()

wb = openpyxl.load_workbook(XLSX)
ws = wb.active
entries = {}
order = 0
for r in range(1, ws.max_row + 1):
    cell = ws.cell(r, 2).value
    if not cell:
        continue
    parts = str(cell).split(", ")
    merged = []
    for p in parts:
        if " - " in p or not merged:
            merged.append(p)
        else:  # 뜻 안의 쉼표로 쪼개진 조각 → 앞 항목에 복원
            merged[-1] += ", " + p
    for p in merged:
        if " - " not in p:
            continue
        left, meaning = p.split(" - ", 1)
        left, meaning = left.strip(), meaning.strip()
        m = re.match(r"^(.*?)\((.*?)\)$", left)
        word, reading = (m.group(1), m.group(2)) if m else (left, "")
        key = (word, reading)
        if key not in entries:
            order += 1
            entries[key] = {"meanings": [], "count": 0, "order": order}
        e = entries[key]
        e["count"] += 1
        if meaning not in e["meanings"]:
            e["meanings"].append(meaning)

words, grammar = [], []
for (word, reading), e in sorted(entries.items(), key=lambda kv: kv[1]["order"]):
    item = (word, reading, " / ".join(e["meanings"]), e["count"])
    (grammar if word.startswith(("~", "～")) else words).append(item)


def row_html(word, reading, meaning, count):
    done = word in done_words
    cls = "row done" if done else "row"
    rd = f'<span class="rd">({html.escape(reading)})</span>' if reading else ""
    badge = f'<span class="badge">×{count}</span>' if count > 1 else ""
    tag = '<span class="tag">TTS 있음</span>' if done else ""
    search = html.escape((word + reading + meaning).lower(), quote=True)
    return (
        f'<div class="{cls}" data-w="{html.escape(word, quote=True)}" data-s="{search}">'
        f'<span class="chk"></span><span class="w">{html.escape(word)}{rd}</span>'
        f'<span class="m">{html.escape(meaning)}</span>{badge}{tag}</div>'
    )


word_rows = "\n".join(row_html(*it) for it in words)
grammar_rows = "\n".join(row_html(*it) for it in grammar)

page = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width, initial-scale=1">
<title>{TITLE} — TTS 단어 고르기</title>
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{ font-family: Arial, "Malgun Gothic", sans-serif; color: #000; background: #f7f7f7; }}
  header {{ position: sticky; top: 0; background: #0C2340; color: #fff;
            padding: 14px 20px 12px; border-bottom: 4px solid #E87722; z-index: 10; }}
  header h1 {{ font-size: 18px; font-weight: bold; }}
  header .sub {{ font-size: 12px; color: #97999B; margin-top: 2px; }}
  .bar {{ display: flex; gap: 10px; align-items: center; margin-top: 10px; }}
  .bar input {{ flex: 1; padding: 7px 10px; font-size: 14px; border: none; border-radius: 4px;
               font-family: inherit; }}
  .cnt {{ font-size: 14px; white-space: nowrap; }}
  .cnt b {{ color: #E87722; font-size: 18px; }}
  main {{ max-width: 760px; margin: 0 auto; padding: 14px 12px 120px; }}
  .row {{ display: flex; align-items: center; gap: 10px; background: #fff; border: 1px solid #e2e2e2;
          border-radius: 6px; padding: 9px 12px; margin-bottom: 6px; cursor: pointer; user-select: none; }}
  .row:hover {{ border-color: #E87722; }}
  .chk {{ width: 20px; height: 20px; flex: none; border: 2px solid #97999B; border-radius: 4px;
          position: relative; }}
  .row.sel {{ background: #FDF1E7; border-color: #E87722; }}
  .row.sel .chk {{ background: #E87722; border-color: #E87722; }}
  .row.sel .chk::after {{ content: "✓"; color: #fff; position: absolute; left: 3px; top: -2px;
                          font-size: 15px; font-weight: bold; }}
  .w {{ font-size: 16px; font-weight: bold; color: #0C2340; }}
  .rd {{ font-weight: normal; font-size: 13px; color: #00587C; margin-left: 3px; }}
  .m {{ font-size: 14px; color: #000; flex: 1; }}
  .badge {{ background: #E87722; color: #fff; font-size: 11px; border-radius: 9px;
            padding: 1px 7px; flex: none; }}
  .tag {{ background: #97999B; color: #fff; font-size: 11px; border-radius: 3px;
          padding: 1px 6px; flex: none; }}
  .row.done {{ opacity: .45; pointer-events: none; }}
  details {{ margin-top: 18px; }}
  summary {{ font-size: 14px; font-weight: bold; color: #0C2340; cursor: pointer; padding: 6px 0; }}
  footer {{ position: fixed; bottom: 0; left: 0; right: 0; background: #fff;
            border-top: 2px solid #0C2340; padding: 12px; text-align: center; }}
  footer button {{ background: #E87722; color: #fff; border: none; font-family: inherit;
                   font-size: 16px; font-weight: bold; padding: 10px 26px; border-radius: 6px;
                   cursor: pointer; }}
  footer button:hover {{ background: #d0691e; }}
  #msg {{ font-size: 12px; color: #7A9A01; margin-top: 6px; min-height: 15px; }}
  #out {{ width: 100%; margin-top: 6px; font-family: inherit; font-size: 13px; display: none; }}
</style>
</head>
<body>
<header>
  <h1>{TITLE} — TTS 단어 고르기</h1>
  <div class="sub">모르는 단어를 클릭해서 선택 → 아래 [선택 목록 복사] → Claude 채팅에 붙여넣기</div>
  <div class="bar">
    <input id="q" type="search" placeholder="검색 (단어·읽기·뜻)">
    <span class="cnt">선택 <b id="n">0</b>개</span>
  </div>
</header>
<main>
{word_rows}
<details>
<summary>문법·조사 항목 ({len(grammar)}개) — 필요하면 열어서 선택</summary>
{grammar_rows}
</details>
</main>
<footer>
  <button id="copy">선택 목록 복사</button>
  <div id="msg"></div>
  <textarea id="out" rows="3" readonly></textarea>
</footer>
<script>
var KEY = "picker_{TITLE}";
var rows = Array.prototype.slice.call(document.querySelectorAll(".row"));
var saved = [];
try {{ saved = JSON.parse(localStorage.getItem(KEY) || "[]"); }} catch (e) {{}}
rows.forEach(function (r) {{
  if (saved.indexOf(r.dataset.w) >= 0 && !r.classList.contains("done")) r.classList.add("sel");
  r.addEventListener("click", function () {{
    r.classList.toggle("sel");
    update();
  }});
}});
function selected() {{
  return rows.filter(function (r) {{ return r.classList.contains("sel"); }})
             .map(function (r) {{ return r.dataset.w; }});
}}
function update() {{
  var s = selected();
  document.getElementById("n").textContent = s.length;
  try {{ localStorage.setItem(KEY, JSON.stringify(s)); }} catch (e) {{}}
}}
document.getElementById("q").addEventListener("input", function () {{
  var q = this.value.trim().toLowerCase();
  rows.forEach(function (r) {{
    r.style.display = !q || r.dataset.s.indexOf(q) >= 0 ? "" : "none";
  }});
}});
document.getElementById("copy").addEventListener("click", function () {{
  var text = selected().join(", ");
  var msg = document.getElementById("msg");
  var out = document.getElementById("out");
  out.value = text;
  function fallback() {{
    out.style.display = "block";
    out.select();
    document.execCommand("copy");
    msg.textContent = "복사됐어요. (아래 상자에서 직접 복사도 가능)";
  }}
  if (navigator.clipboard && navigator.clipboard.writeText) {{
    navigator.clipboard.writeText(text).then(function () {{
      msg.textContent = "복사 완료! Claude 채팅에 붙여넣으세요. (" + selected().length + "개)";
    }}, fallback);
  }} else {{ fallback(); }}
}});
update();
</script>
</body>
</html>
"""

with open(OUT, "w", encoding="utf-8") as f:
    f.write(page)

print("어휘:", len(words), "| 문법:", len(grammar), "| 이미 TTS:", len(done_words))
print("단어장:", DB_CSV)
print("OUT:", OUT)
