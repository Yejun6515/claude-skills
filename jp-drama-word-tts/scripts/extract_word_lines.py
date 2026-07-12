# -*- coding: utf-8 -*-
r"""선택된 단어 → 드라마 원문 대사(C열)·번역(D열)·읽기·뜻 매핑 (예문1 재료).

사용법: python extract_word_lines.py <화 폴더> <제목> <단어목록파일>
  - 단어목록파일: 피커에서 복사한 "단어1, 단어2, ..." 텍스트를 저장한 txt
  - 출력: {폴더}\_word_lines.json (단어별 후보 대사 최대 3개, 짧은 순)
"""
import json
import os
import re
import sys

import openpyxl

EP_DIR = sys.argv[1]
TITLE = sys.argv[2]
with open(sys.argv[3], encoding="utf-8") as f:
    WORDS = [w.strip() for w in f.read().replace("\n", ",").split(",") if w.strip()]

wb = openpyxl.load_workbook(os.path.join(EP_DIR, f"{TITLE}_분석완성.xlsx"))
ws = wb.active
rows = []
for r in range(1, ws.max_row + 1):
    rows.append((r, str(ws.cell(r, 2).value or ""), str(ws.cell(r, 3).value or ""),
                 str(ws.cell(r, 4).value or "")))

result = {}
for w in WORDS:
    hits = []
    pat = re.escape(w)
    for r, b, c, d in rows:
        if re.search(pat + r"[(（\s]", b) or (w + " -") in b or w in c:
            m = re.search(re.escape(w) + r"\((.*?)\)\s*-\s*([^,]+)", b)
            reading = m.group(1) if m else ""
            meaning = m.group(2).strip() if m else ""
            if not m:
                m2 = re.search(re.escape(w) + r"\s*-\s*([^,]+)", b)
                meaning = m2.group(1).strip() if m2 else meaning
            hits.append({"row": r, "jp": c.strip(), "ko": d.strip(),
                         "reading": reading, "meaning": meaning})
    hits.sort(key=lambda h: len(h["jp"]))
    result[w] = hits[:3]

missing = [w for w in WORDS if not result[w]]
out = os.path.join(EP_DIR, "_word_lines.json")
with open(out, "w", encoding="utf-8") as f:
    json.dump(result, f, ensure_ascii=False, indent=1)
print("매핑:", len(WORDS) - len(missing), "/", len(WORDS), "| 미발견:", missing)
print("OUT:", out)
