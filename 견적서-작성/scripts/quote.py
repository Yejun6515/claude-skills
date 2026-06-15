# -*- coding: utf-8 -*-
"""
견적서-작성 engine. Mechanical xlsx I/O + verification.
Semantic / must-confirm judgement is done by Claude, not here.

Subcommands:
  new     <machine|spare> <out.xlsx> [--templates DIR]
          Copy the blank master to out.xlsx.

  fill    <out.xlsx> <values.json>
          values.json = {"tokens": {"REF_NO":"0025S242", ...},
                          "cells":  {"N24": 76766000, "V6": "2026-05-15", ...},
                          "sv_payment": "A" | "B"}
          tokens: every {{TOKEN}} occurrence (across all cells) is replaced.
          cells:  direct cell writes (numbers, dates as ISO yyyy-mm-dd).
          sv_payment (machine only): Supervising Services 결제조건 선택.
            "A" (기본) = 50/50 man-day 리포트 기준 (master 기본값, no-op).
            "B" = 100% time sheet 서명 후 1개월 (D67 교체, C68/D68 삭제).

  verify  <out.xlsx> <machine|spare>
          Mechanical checks -> prints a report. Exit 1 if any FAIL.

  dump    <out.xlsx>
          Print all non-empty cells (for Claude to eyeball).

  logo    <out.xlsx>
          Re-inject the Primetals header logo (idempotent). Run on any draft
          whose logo got stripped. `new` and `fill` already call this.
"""
import sys, os, json, shutil, datetime, re, zipfile
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__))
MASTER = {"machine": "machine_master.xlsx", "spare": "spare_master.xlsx"}
LOGO_ASSET = os.path.join(HERE, "logo_primetals.jpeg")

# Primetals logo, top-left header. Fixed-size oneCellAnchor so it looks the
# same on both the machine and the (narrower) spare layout.
_DRAWING_XML = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<xdr:wsDr xmlns:xdr="http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"'
    ' xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">'
    '<xdr:oneCellAnchor><xdr:from><xdr:col>0</xdr:col><xdr:colOff>112835</xdr:colOff>'
    '<xdr:row>0</xdr:row><xdr:rowOff>118697</xdr:rowOff></xdr:from>'
    '<xdr:ext cx="2431073" cy="711444"/>'
    '<xdr:pic><xdr:nvPicPr><xdr:cNvPr id="2" name="Primetals Logo"/>'
    '<xdr:cNvPicPr><a:picLocks noChangeAspect="1" noChangeArrowheads="1"/></xdr:cNvPicPr></xdr:nvPicPr>'
    '<xdr:blipFill><a:blip xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'
    ' r:embed="rId1" cstate="print"/><a:srcRect/><a:stretch><a:fillRect/></a:stretch></xdr:blipFill>'
    '<xdr:spPr bwMode="auto"><a:xfrm><a:off x="0" y="0"/><a:ext cx="2431073" cy="711444"/></a:xfrm>'
    '<a:prstGeom prst="rect"><a:avLst/></a:prstGeom><a:noFill/><a:ln><a:noFill/></a:ln></xdr:spPr>'
    '</xdr:pic><xdr:clientData/></xdr:oneCellAnchor></xdr:wsDr>'
)
_DRAWING_RELS = (
    '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
    '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
    '<Relationship Id="rId1" Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/image"'
    ' Target="../media/image1.jpeg"/></Relationships>'
)
_R_NS = ' xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"'


def ensure_logo(path):
    """Idempotently inject the Primetals logo into an .xlsx via zip surgery.

    openpyxl drops images on load+save, so cmd_fill calls this AFTER wb.save to
    restore the logo. Safe to call repeatedly (no-op if already present).
    """
    if not os.path.exists(LOGO_ASSET):
        print("WARN logo asset missing, skipped:", LOGO_ASSET)
        return
    with zipfile.ZipFile(path) as z:
        names = z.namelist()
        if "xl/media/image1.jpeg" in names and "xl/drawings/drawing1.xml" in names:
            return  # already has it
        items = {n: z.read(n) for n in names}

    ws_name = "xl/worksheets/sheet1.xml"
    rels_name = "xl/worksheets/_rels/sheet1.xml.rels"
    draw_type = "http://schemas.openxmlformats.org/officeDocument/2006/relationships/drawing"

    # --- sheet rels: add a drawing relationship with a free rId ---
    if rels_name in items:
        s = items[rels_name].decode("utf-8")
        used = [int(m) for m in re.findall(r'Id="rId(\d+)"', s)]
        rid = "rId%d" % (max(used) + 1 if used else 1)
        s = s.replace("</Relationships>",
                      '<Relationship Id="%s" Type="%s" Target="../drawings/drawing1.xml"/></Relationships>'
                      % (rid, draw_type))
        items[rels_name] = s.encode("utf-8")
    else:
        rid = "rId1"
        items[rels_name] = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">'
            '<Relationship Id="%s" Type="%s" Target="../drawings/drawing1.xml"/></Relationships>'
            % (rid, draw_type)).encode("utf-8")

    # --- worksheet: ensure r: namespace, append <drawing> before </worksheet> ---
    s = items[ws_name].decode("utf-8")
    head = s[:s.find(">")]
    if "xmlns:r=" not in head:
        idx = s.find("<worksheet") + len("<worksheet")
        s = s[:idx] + _R_NS + s[idx:]
    if "<drawing " not in s:
        s = s.replace("</worksheet>", '<drawing r:id="%s"/></worksheet>' % rid)
    items[ws_name] = s.encode("utf-8")

    # --- content types: jpeg default + drawing override ---
    ct = items["[Content_Types].xml"].decode("utf-8")
    add = ""
    if 'Extension="jpeg"' not in ct:
        add += '<Default Extension="jpeg" ContentType="image/jpeg"/>'
    if "drawings/drawing1.xml" not in ct:
        add += ('<Override PartName="/xl/drawings/drawing1.xml"'
                ' ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>')
    if add:
        ct = ct.replace("</Types>", add + "</Types>")
    items["[Content_Types].xml"] = ct.encode("utf-8")

    # --- new parts ---
    items["xl/drawings/drawing1.xml"] = _DRAWING_XML.encode("utf-8")
    items["xl/drawings/_rels/drawing1.xml.rels"] = _DRAWING_RELS.encode("utf-8")
    items["xl/media/image1.jpeg"] = open(LOGO_ASSET, "rb").read()

    tmp = path + ".tmp"
    with zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as out:
        for n, data in items.items():
            out.writestr(n, data)
    os.replace(tmp, path)


def find_templates(arg=None):
    if arg and os.path.isdir(arg):
        return arg
    # search upward from CWD for a _templates folder holding the masters
    d = os.getcwd()
    for _ in range(6):
        for cand in (os.path.join(d, "_templates"),
                     os.path.join(d, "견적서", "_templates"),
                     d):
            if os.path.exists(os.path.join(cand, "machine_master.xlsx")):
                return cand
        d = os.path.dirname(d)
    raise SystemExit("ERROR: _templates with master files not found. Pass --templates DIR.")


def cmd_new(argv):
    kind, out = argv[0], argv[1]
    tdir = None
    if "--templates" in argv:
        tdir = argv[argv.index("--templates") + 1]
    tdir = find_templates(tdir)
    src = os.path.join(tdir, MASTER[kind])
    shutil.copyfile(src, out)
    ensure_logo(out)
    print(f"NEW  {kind} -> {out}\n     (from {src})")


def _coerce(v):
    if isinstance(v, str) and re.fullmatch(r"\d{4}-\d{2}-\d{2}", v):
        y, m, d = map(int, v.split("-"))
        return datetime.datetime(y, m, d)
    return v


# Supervising Services 결제조건 (machine). Master 기본값은 A(50/50). B는 100%/time-sheet.
SV_PAY_B_D67 = ("100% of the total price: Shall be paid by T/T within 1 month "
                "after the signing of a time sheet.")


def _apply_sv_payment(ws, opt):
    """Switch the Supervising Services payment block (B66 아래 C67/D67, C68/D68).

    A = master 기본(50/50 man-day 리포트) → no-op.
    B = 100% time sheet 서명 후 1개월 → D67 교체, (2)줄(C68/D68) 삭제.
    행은 지우지 않고 값만 비워 아래 셀 참조(D71/D85 등) 보존.
    """
    if not opt or str(opt).upper() == "A":
        return
    if str(opt).upper() != "B":
        print("WARN unknown sv_payment (expected A|B):", opt)
        return
    ws["C67"] = "(1)"
    ws["D67"] = SV_PAY_B_D67
    ws["C68"] = None
    ws["D68"] = None
    print("FILL sv_payment: B (100% / time sheet 서명 후 1개월) — C68/D68 비움")


def cmd_fill(argv):
    out, vjson = argv[0], argv[1]
    with open(vjson, encoding="utf-8") as f:
        vals = json.load(f)
    tokens = vals.get("tokens", {})
    cells = vals.get("cells", {})
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    # token replacement across all cells
    hits = {}
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "{{" in c.value:
                new = c.value
                for k, val in tokens.items():
                    tok = "{{" + k + "}}"
                    if tok in new:
                        new = new.replace(tok, str(val))
                        hits[k] = hits.get(k, 0) + 1
                if new != c.value:
                    c.value = new
    # direct cell writes
    for coord, val in cells.items():
        ws[coord] = _coerce(val)
    # Supervising Services 결제조건 선택 (machine)
    _apply_sv_payment(ws, vals.get("sv_payment"))
    wb.save(out)
    ensure_logo(out)  # openpyxl drops images on save -> restore the logo
    print("FILL tokens applied:", {k: hits.get(k, 0) for k in tokens})
    miss = [k for k in tokens if hits.get(k, 0) == 0]
    if miss:
        print("WARN tokens with no match (typo?):", miss)
    print("FILL cells set:", list(cells))


def _num(v):
    return v if isinstance(v, (int, float)) else None


def cmd_verify(argv):
    out, kind = argv[0], argv[1]
    wb = openpyxl.load_workbook(out)
    ws = wb.active
    fails, warns, oks = [], [], []

    # 1. leftover markers
    left = []
    for row in ws.iter_rows():
        for c in row:
            if isinstance(c.value, str) and "{{" in c.value:
                left.append(f"{c.coordinate}={c.value!r}")
    if left:
        fails.append("미입력 마커 잔존:\n      " + "\n      ".join(left))
    else:
        oks.append("마커 잔존 없음")

    if kind == "machine":
        # equipment subtotal Σ(W24:W27)
        eq = [_num(ws[f"W{r}"].value) for r in range(24, 28)]
        eqsum = sum(x for x in eq if x)
        if eqsum == 0:
            warns.append("장비 금액(W24:W27)이 비어있음")
        else:
            oks.append(f"장비 소계 Σ(W24:W27)={eqsum:,.0f}")
        mds = _num(ws["L34"].value)
        rate = _num(ws["N34"].value)
        if mds and rate:
            oks.append(f"SV {rate:,.0f}×{mds}MD={rate*mds:,.0f}")
        else:
            warns.append("SV MD수(L34) 또는 단가(N34) 비어있음")
        # validity > date  (free text -> just surface)
        warns.append(f"Validity 날짜(D85)={ws['D85'].value!r} 가 견적일({ws['W3'].value!r})보다 미래인지 확인")
        warns.append("결제 % 합(Equipment=100, SV 단독) 및 L/C 조건이 이 고객에 맞는지 확인 (자동통과 금지)")
        # Supervising Services 결제조건: A(50/50) vs B(100%/time sheet) 어느 쪽인지 표면화
        d67 = str(ws["D67"].value or "")
        d68 = str(ws["D68"].value or "")
        if d67.startswith("100%"):
            oks.append("SV 결제 = 옵션 B (100% / time sheet 서명 후 1개월)")
            if d68.strip():
                warns.append(f"SV 옵션 B인데 D68에 (2)줄 잔존={d68!r} — 비워야 함")
        elif d67.startswith("50%"):
            oks.append("SV 결제 = 옵션 A (50/50 man-day 리포트 기준)")
        else:
            warns.append(f"SV 결제조건(D67) 표준문구 아님={d67!r} — 옵션 A/B 확인")

    elif kind == "spare":
        # per-item U=V/S sanity + total
        tot = 0
        for r in (21, 24, 27):
            S, V = _num(ws[f"S{r}"].value), _num(ws[f"V{r}"].value)
            if V:
                tot += V
                if S:
                    oks.append(f"item r{r}: 금액 {V:,.0f} / 수량 {S} -> 단가 {V/S:,.0f}")
                else:
                    warns.append(f"item r{r}: 수량(S{r}) 비어있음")
        if tot == 0:
            warns.append("품목 금액(V21/V24/V27)이 모두 비어있음")
        else:
            oks.append(f"품목 합계 ~{tot:,.0f} (행 추가시 V33 SUM 범위 확인)")
        v6 = ws["V6"].value
        if isinstance(v6, datetime.datetime):
            oks.append(f"견적일 V6={v6:%Y-%m-%d} -> 유효기일 자동 {v6+datetime.timedelta(days=45):%Y-%m-%d}")
        else:
            fails.append(f"견적일 V6 미입력/형식오류={v6!r} (유효기일 깨짐)")
        warns.append("결제조건(I37)·선적개월(I39)·준거법이 이 고객에 맞는지 확인 (자동통과 금지)")

    print("==== VERIFY", kind, "====")
    for x in oks:
        print("  [OK]  ", x)
    for x in warns:
        print("  [확인]", x)
    for x in fails:
        print("  [FAIL]", x)
    if fails:
        print("\n결과: FAIL — 위 [FAIL] 항목을 고칠 것.")
        sys.exit(1)
    print("\n결과: 기계검증 통과. [확인] 항목은 Claude가 사용자와 점검.")


def cmd_dump(argv):
    wb = openpyxl.load_workbook(argv[0], data_only=False)
    ws = wb.active
    for row in ws.iter_rows():
        cells = [f"{c.coordinate}={c.value!r}" for c in row
                 if c.value is not None and str(c.value).strip() != ""]
        if cells:
            print(" | ".join(cells))


def main():
    try:  # Korean paths/labels must not crash on a cp932 console
        sys.stdout.reconfigure(encoding="utf-8")
    except Exception:
        pass
    if len(sys.argv) < 2:
        print(__doc__); sys.exit(1)
    cmd, argv = sys.argv[1], sys.argv[2:]
    {"new": cmd_new, "fill": cmd_fill, "verify": cmd_verify,
     "dump": cmd_dump, "logo": lambda a: ensure_logo(a[0])}[cmd](argv)


if __name__ == "__main__":
    main()
