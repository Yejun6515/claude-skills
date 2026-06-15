# -*- coding: utf-8 -*-
"""
Master quotation templates.

Two tracks:
  • spare_master.xlsx   — GENERATED here from the raw example (blank/tokenize inputs).
  • machine_master.xlsx — HAND-MAINTAINED.  Its 1.Price table layout (numbering
    (1)/(2), N:V merges, table-only borders), print fit, SV 50/50 payment and the
    T&C sections (5 General conditions / 6 Guarantee / 7 Validity / 8 Other
    Conditions) were tuned by hand in Excel.  Do NOT regenerate it from the raw
    example — that would wipe the manual layout.  To fold a hand-edited,
    header-filled draft back into the blank master:
        python build_masters.py <draft.xlsx>

Marker convention:
  {{TOKEN}}  = identity field; a surviving {{...}} in a finished draft is flagged
               by the skill's verify step (catches "이전 건 텍스트 잔존" / problem C).
  None       = numeric input cell, left empty (formulas treat as 0).
Standard term sentences (payment / L-C) keep default text; the skill's check forces
explicit confirmation of those every time (problem D).
"""
import shutil, os, sys, glob, importlib.util, openpyxl

BASE = os.path.dirname(os.path.abspath(__file__))
ROOT = os.path.dirname(BASE)


def add_logo(path):
    """Re-inject the Primetals header logo after an openpyxl save (which drops
    images). Delegates to the skill's quote.ensure_logo so there is one source
    of truth. No-op (with a warning) if the skill script can't be located."""
    cands = glob.glob(os.path.join(
        os.path.expanduser("~"), ".claude", "skills", "*", "scripts", "quote.py"))
    for qp in cands:
        try:
            spec = importlib.util.spec_from_file_location("_quote_logo", qp)
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            if os.path.exists(getattr(mod, "LOGO_ASSET", "")):
                mod.ensure_logo(path)
                return
        except Exception:
            continue
    print("WARN logo not re-injected (skill quote.py/logo asset not found):", path)


def print_fit(ws, print_area):
    """Scale to one page wide so nothing is cut off on the right when printed."""
    from openpyxl.worksheet.properties import PageSetupProperties
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0          # width fixed to 1 page, height flows
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.print_area = print_area


def set_cells(path, edits, blanks):
    wb = openpyxl.load_workbook(path)  # keep formulas
    ws = wb.active
    for coord, val in edits.items():
        ws[coord] = val
    for coord in blanks:
        ws[coord] = None
    wb.save(path)
    return ws.title


# ---------------------------------------------------------------- MACHINE (hand-maintained)
# Identity tokens the skill fills; reblank_machine_master restores them.
MACHINE_TOKENS = {
    "W2":  "Ref. No.:{{REF_NO}}",
    "W3":  "Date: {{DATE}}",
    "B7":  "Messrs: {{CUSTOMER}}",
    "B11": "Subject: {{SUBJECT}}",
    "B14": "We're pleased to offer our proposal for {{SUBJECT}} as follows,",
    "D20": " {{SUBJECT}}",
    "E29": "* Please refer to the Proposal Specification ({{SPEC_NO}}) for the details",
    "E35": "* Please refer to the Proposal Specification ({{SPEC_NO}}) for the details",
    "E50": "Translator shall be prepared by {{CUSTOMER_SHORT}}, if necessary.",
    "E53": "Please see attached our Proposal Specification ({{SPEC_NO}})",
    "D109": "Proposal Specification ({{SPEC_NO}})",
}


def reblank_machine_master(filled_path, dst=os.path.join(BASE, "machine_master.xlsx")):
    """Fold a hand-edited, header-filled machine draft back into the blank master
    by restoring the identity tokens (reverse of the skill's header fill).  The
    price/terms tokens are assumed still un-filled in the draft."""
    shutil.copyfile(filled_path, dst)
    wb = openpyxl.load_workbook(dst)
    ws = wb.active
    for coord, tok in MACHINE_TOKENS.items():
        ws[coord] = tok
    wb.save(dst)
    add_logo(dst)
    print("machine_master.xlsx refreshed from", filled_path)


# ---------------------------------------------------------------- SPARE (generated)
def build_spare():
    sp_src = os.path.join(ROOT, "스페어 견적서", "20260107 quotation for POSCO K1H Spare parts.xlsx")
    sp_dst = os.path.join(BASE, "spare_master.xlsx")
    shutil.copyfile(sp_src, sp_dst)
    sp_edits = {
        "V5": "NO. {{REF_NO}}",
        "A7": "Messrs: {{CUSTOMER}}",
        "C10": "Subject : SPARE PARTS FOR {{PROJECT}} / {{END_USER}} ({{COUNTRY}})",
        # line items (Unit Price U = V/S is a formula, kept) ---
        "D21": "{{ITEM1_DESC}}",
        "D22": "{{ITEM1_DESC2}}  (2nd desc line, clear if unused)",
        "D24": "{{ITEM2_DESC}}",
        "D25": "{{ITEM2_DESC2}}  (2nd desc line, clear if unused)",
        "D27": "{{ITEM3_DESC}}",
        "D28": "{{ITEM3_DESC2}}  (2nd desc line, clear if unused)",
        # terms (MUST-CONFIRM) ---
        "I37": "{{PAYMENT_TERMS}}",
        "I39": "Within {{SHIPMENT_MONTHS}} months after receipt of your purchase order.",
    }
    sp_blanks = ["V6", "S21", "V21", "S24", "V24", "S27", "V27"]
    # I38 'FOB JAPAN (INCOTERMS 2020)' kept as default; K40 validity = V6+45 formula kept.
    title = set_cells(sp_dst, sp_edits, sp_blanks)
    wb = openpyxl.load_workbook(sp_dst)
    print_fit(wb.active, "A1:V72")
    wb.save(sp_dst)
    add_logo(sp_dst)
    print("spare_master.xlsx built  (sheet", repr(title), ")")


if __name__ == "__main__":
    if len(sys.argv) > 1:
        reblank_machine_master(sys.argv[1])     # refold a hand-edited draft into machine master
    else:
        build_spare()                            # default: rebuild spare only
        print("machine_master.xlsx is hand-maintained — left untouched. "
              "Use 'python build_masters.py <draft.xlsx>' to refold a hand-edited draft.")
