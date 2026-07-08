#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""
Generate a meeting-availability grid workbook (.xlsx).

Reproduces the company template:
  A column = date labels (merged per day, e.g. "5/26(Tue)")
  B column = time slots (09:00~10:00 ... 16:00~17:00) with a shaded "Lunch" row
  C column onward = one column per person; body cells left blank for hand-marking

Usage:
    python generate_meeting_grid.py <data.json> <output.xlsx>

data.json shape:
{
  "dates":  ["2026-05-26", "2026-05-27"],          # ISO; weekday computed automatically
  "people": ["Kuroda, Akio", "Abe, Hironori", ...],  # column headers, verbatim
  "time_slots": ["09:00~10:00", "10:00~11:00", "11:00~12:00", "Lunch",
                 "13:00~14:00", "14:00~15:00", "15:00~16:00", "16:00~17:00"],  # optional
  "lunch_label": "Lunch"                             # optional; rows == this get yellow fill
}
"""
import json
import sys
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

DEFAULT_SLOTS = [
    "09:00~10:00", "10:00~11:00", "11:00~12:00", "Lunch",
    "13:00~14:00", "14:00~15:00", "15:00~16:00", "16:00~17:00",
]
WEEKDAY = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]

# --- shared styles -----------------------------------------------------------
THIN = Side(style="thin", color="000000")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
YELLOW = PatternFill(fill_type="solid", fgColor="FFFF00")
LABEL_FONT = Font(name="Malgun Gothic", size=11)          # date / time labels
BODY_FONT = Font(name="Yu Gothic", size=11)               # people headers + cells
CENTER = Alignment(horizontal="center", vertical="center")
LEFT_MID = Alignment(horizontal="left", vertical="center")


def date_label(iso: str) -> str:
    """'2026-05-26' -> '5/26(Tue)'."""
    y, m, d = (int(x) for x in iso.split("-"))
    return f"{m}/{d}({WEEKDAY[date(y, m, d).weekday()]})"


def build(data: dict, out_path: str) -> None:
    dates = data["dates"]
    people = data["people"]
    slots = data.get("time_slots") or DEFAULT_SLOTS
    lunch_label = data.get("lunch_label", "Lunch")

    if not dates:
        raise ValueError("no dates given")
    if not people:
        raise ValueError("no people given")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"
    ws.sheet_view.showGridLines = True

    last_col = 2 + len(people)          # A=1, B=2(Time), people start at C=3
    last_col_letter = get_column_letter(last_col)

    # --- header row (row 2, matching the template's 1-row top margin) --------
    hr = 2
    ws.cell(hr, 1).border = BORDER      # A2 empty
    b2 = ws.cell(hr, 2, "Time")
    b2.font = LABEL_FONT
    b2.border = BORDER
    b2.alignment = CENTER
    for i, name in enumerate(people):
        c = ws.cell(hr, 3 + i, name)
        c.font = BODY_FONT
        c.border = BORDER
        c.alignment = CENTER

    # --- one block per date --------------------------------------------------
    r = hr + 1
    for iso in dates:
        block_start = r
        for slot in slots:
            is_lunch = (slot == lunch_label)
            # B column: time slot label
            bc = ws.cell(r, 2, slot)
            bc.font = LABEL_FONT
            bc.border = BORDER
            bc.alignment = LEFT_MID
            if is_lunch:
                bc.fill = YELLOW
            # A column cell (merged later) — keep border/style on each row
            ac = ws.cell(r, 1)
            ac.border = BORDER
            ac.alignment = CENTER
            # people cells (blank, bordered)
            for i in range(len(people)):
                cell = ws.cell(r, 3 + i)
                cell.font = BODY_FONT
                cell.border = BORDER
                cell.alignment = CENTER
                if is_lunch:
                    cell.fill = YELLOW
            r += 1
        block_end = r - 1
        # merge A column across the block and write the date label
        ws.merge_cells(start_row=block_start, start_column=1,
                       end_row=block_end, end_column=1)
        top = ws.cell(block_start, 1, date_label(iso))
        top.font = LABEL_FONT
        top.alignment = CENTER

    # --- column widths (match template) --------------------------------------
    ws.column_dimensions["A"].width = 12.5
    ws.column_dimensions["B"].width = 20.5
    for i in range(len(people)):
        ws.column_dimensions[get_column_letter(3 + i)].width = 15.75

    wb.save(out_path)
    print(f"OK  {out_path}")
    print(f"    {len(dates)} date block(s) x {len(slots)} slots, {len(people)} people, "
          f"columns A:{last_col_letter}")


def main() -> None:
    if len(sys.argv) != 3:
        print(__doc__)
        sys.exit(1)
    with open(sys.argv[1], "r", encoding="utf-8-sig") as f:
        data = json.load(f)
    build(data, sys.argv[2])


if __name__ == "__main__":
    main()
