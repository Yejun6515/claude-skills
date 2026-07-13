# -*- coding: utf-8 -*-
"""LoA 리스크 매칭 시트 생성 템플릿.
사용법: 이 파일을 scratchpad에 복사 → CONFIG 블록만 계약에 맞게 채우고 실행.
실행: $env:PYTHONIOENCODING='utf-8'; python build_matching.py  (PS5.1 한글 출력 깨짐 방지)

입력 1) consistency xlsx — '근거 (KO)' 시트, 컬럼 레이아웃(1행 헤더):
  A No | B Category | C Risk item | D Classification (draft) | E Source type
  F Rationale | G Source reference | H Conflict(Yes/No) | I Conflict detail
  J Suggested classification | K~V Loop1~4(판정/Δ/comment ×4) | W 특이점합의
입력 2) 해설 매뉴얼 xlsx — 스킬 assets의 Small_LoA_해설매뉴얼.xlsx (E열 한글해석, H열 RC기준)
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ═══════════ CONFIG — 계약별로 여기만 수정 ═══════════
CONS = r'<consistency xlsx 경로>'                       # 계약 폴더의 판정 원본
MANUAL = r'C:\Users\Z006K14G\.claude\skills\loa-risk-matching\assets\Small_LoA_해설매뉴얼.xlsx'
OUT = r'<계약 폴더>\<프로젝트코드>_LoA_리스크매칭_<YYMMDD>.xlsx'
TITLE = '<프로젝트명> (<프로젝트코드>) — Small LoA 리스크 매칭 시트'
N_ITEMS = 74                                            # Small=74 / Standard=100(No가 "1.1" 문자열)

# 기준 문서 3종 + 미입수 문서 — 반드시 실제 계약에 맞게 서술
BASIS = ('기준 문서: ① <고객 일반약관/GTC(버전)> ② <구입사양서 등> ③ <PT draft 제출 조건>. '
         '미입수: <특별조건/TLoA/위임액 등> → 해당 항목은 "보류" 표시, 입수 시 재판정. '
         '판정 원본: <consistency 파일명>. 항목 해설은 LoA 해설매뉴얼(loa-risk-matching 스킬 assets) 참조.')

# 상태 분류 — consistency 판정·status 메모에서 항목번호를 옮겨 적는다
PENDING = {}          # {no: '보류 사유'}  예: {19: 'TLoA 미입수', 22: '위임액 확인 필요'}
JUDGE = set()         # 판단 요(사용자 escalation 판단 대기)
RC_CONFIRMED = {}     # {no: 'RC3'} 등 확정 등급
REGION = set()        # Region/Business 해당 (Small 전용 열)
CLOSED = set()        # N/A 종결(조항 인용으로 확정) — 나머지는 자동으로 'N/A(초안 유지)'
REVISIT = set()       # 상위문서(특별조건 등) 입수 후 재확인할 종결 항목 → 비고
TOP = set()           # ★최우선 특이점
KEY = set()           # ★핵심 특이점
# ════════════════════════════════════════════════════

DB = '0C2340'; OR = 'E87722'; GY = '97999B'; TE = '00587C'; WH = 'FFFFFF'
GN = '7A9A01'; RD = 'CE0037'; LIGHT = 'F2F4F7'

def status_of(no):
    if no in PENDING:
        return (f'🔔 보류 — {PENDING[no]}', RD)
    if no in JUDGE:
        star = ' ★최우선' if no in TOP else (' ★핵심' if no in KEY else '')
        return (f'판단 요{star} — escalation 판단 대기', OR)
    if no in RC_CONFIRMED:
        return (f'{RC_CONFIRMED[no]} 확정', TE)
    if no in REGION:
        return ('Region/Business 해당', TE)
    if no in CLOSED:
        return ('N/A 종결 (조항 확인)', GN)
    return ('N/A (초안 유지)', GN)

def remark_of(no):
    r = []
    if no in REVISIT:
        r.append('상위 계약문서 입수 후 재확인')
    if no in PENDING:
        r.append('문서 입수 시 즉시 판정')
    if no in TOP:
        r.append('Legal 서명·자유기재 근거 항목')
    return ' / '.join(r) or '-'

cwb = openpyxl.load_workbook(CONS, data_only=True)
cws = cwb['근거 (KO)']
cons = {}
for row in cws.iter_rows(min_row=2, values_only=True):
    try:
        no = int(row[0])
    except (TypeError, ValueError):
        continue
    cons[no] = {'cat': row[1], 'draft': row[3], 'src': row[4], 'rat': row[5],
                'ref': row[6], 'conf': row[7], 'confd': row[8], 'sug': row[9]}

mwb = openpyxl.load_workbook(MANUAL, data_only=True)
mws = mwb.active
manual = {}
for r in range(6, mws.max_row + 1):
    v = mws.cell(r, 1).value
    if isinstance(v, int):
        manual[v] = {'ko': mws.cell(r, 5).value, 'rc': mws.cell(r, 8).value}

assert len(cons) == N_ITEMS and len(manual) == N_ITEMS, (len(cons), len(manual))

wb = openpyxl.Workbook(); ws = wb.active; ws.title = '리스크 매칭'
thin = Side(style='thin', color='D0D5DD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)
def kfont(sz=10, bold=False, color='000000'):
    return Font(name='맑은 고딕', size=sz, bold=bold, color=color)

ws.merge_cells('A1:I1'); c = ws['A1']; c.value = TITLE
c.font = kfont(16, True, DB); ws.row_dimensions[1].height = 28
ws.merge_cells('A2:I2'); c = ws['A2']; c.value = BASIS
c.font = kfont(9, False, GY); c.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[2].height = 44
ws.merge_cells('A3:I3'); c = ws['A3']
c.value = ('구분 읽는 법: N/A(초안 유지/종결) = 비해당 확정 · RC 확정 = 해당 등급으로 상신 · '
           '판단 요 = 조항상 해당 소지 확인, 기재(escalation) 여부는 사용자 판단 대기 · 보류 = 근거 문서 미입수. '
           '★최우선/★핵심은 Legal 서명·자유기재 문안의 근거 항목.')
c.font = kfont(9, False, GY); c.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[3].height = 32

HDR = ['No', '질문 (한글 해석)', '이번 계약 판정 (구분)', '근거 문서', '근거 조항', '근거 요지',
       '충돌·특이점 (draft 대비)', 'RC 판정 기준 (매뉴얼)', '비고']
hr = 5
for i, h in enumerate(HDR, 1):
    c = ws.cell(row=hr, column=i, value=h)
    c.font = Font(name='맑은 고딕', size=10, bold=True, color=WH)
    c.fill = PatternFill('solid', fgColor=DB)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws.row_dimensions[hr].height = 22
for i, w in enumerate([5, 40, 26, 15, 22, 50, 44, 28, 26], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = hr + 1; prev_cat = None
for no in sorted(cons):
    d = cons[no]; m = manual[no]
    if d['cat'] != prev_cat:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=str(d['cat']))
        c.font = kfont(11, True, OR); c.fill = PatternFill('solid', fgColor=LIGHT)
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = border
        ws.row_dimensions[row].height = 20
        row += 1; prev_cat = d['cat']
    st, stcol = status_of(no)
    conf = d['confd'] if (d['conf'] and str(d['conf']).strip().lower() == 'yes') else '-'
    vals = [no, m['ko'], st, d['src'] or '-', d['ref'] or '-', d['rat'] or '-', conf, m['rc'], remark_of(no)]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=True, horizontal='center' if i == 1 else 'left')
        if i == 1: c.font = kfont(10, True, DB)
        elif i == 3: c.font = kfont(10, True, stcol)
        elif i in (4, 5): c.font = kfont(9, False, '404040')
        elif i == 7: c.font = kfont(9, False, RD if conf != '-' else GY)
        elif i == 8: c.font = kfont(9, False, TE)
        elif i == 9: c.font = kfont(9, False, GY)
        else: c.font = kfont(10)
    row += 1

row += 1
ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
ws.cell(row=row, column=1, value='요약').font = kfont(12, True, DB); row += 1
counts = {}
for no in cons:
    key = status_of(no)[0].split(' —')[0].split(' (')[0]
    counts[key] = counts.get(key, 0) + 1
for k, v in sorted(counts.items(), key=lambda x: -x[1]):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
    ws.cell(row=row, column=1, value=f'  {k}: {v}건').font = kfont(10)
    row += 1

ws.freeze_panes = 'A6'; ws.sheet_view.zoomScale = 90
ws.auto_filter.ref = f'A{hr}:I{row-1}'
wb.save(OUT)
print('saved:', OUT)
for k, v in counts.items():
    print(f'  {k}: {v}')
