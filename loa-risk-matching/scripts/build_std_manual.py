# -*- coding: utf-8 -*-
"""Standard LoA (Offer Approval) Risk Questionnaire 해설 매뉴얼 xlsx 생성.
소스: HQ-003-05 V5.3 (2026-06-01) 질문 원문 + Explanatory Notes to LoA Risk Questionnaire (rev1)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r'U:\新_海外営業部\Kim Yejun\1_POSCO\3_POSCO Servo valve\K1H\26년계약\260625_LoA준비\workflow\Standard_LoA_RiskQuestionnaire_해설매뉴얼_20260714.xlsx'

DB = '0C2340'; OR = 'E87722'; GY = '97999B'; TE = '00587C'; WH = 'FFFFFF'
LIGHT = 'F2F4F7'

# (ID, 대분류, 소분류, 영어원문, 한글해석, 의미·목적, 예시, RC기준, Small대응)
R = []
def a(*x): R.append(x)

T='Technical Risks'; C='Commercial Risks'; L='Legal Risks'; P='Compliance Risks'; O='Other Risks'

# ── 1. Project scope ─────────────────────────────────────────
a('1.1', T, '1. Project scope',
  'Full Turnkey (Semi Turnkey including Turnkey Balance of Plants (TK BP) components like civils, building structure, cranes, building services) acc. to HQ-061-00.',
  '풀턴키 프로젝트인가? (세미턴키 + 토건·건물구조·크레인·건물설비 등 BoP 포함)',
  '풀턴키는 플랜트 전체(설계~기자재~토건~설치~시운전)를 계약자가 일괄 책임지는, 플랜트 사업에서 가장 리스크가 큰 계약 형태. "키만 돌리면 가동되는 상태"로 인도할 의무를 짐. 금액과 무관하게 최고 경영진 승인 필요.',
  '제철소 신설에서 건물 기초·크레인·유틸리티 공사까지 PT가 일괄 수행하는 계약.',
  '해당 시 RC1 (TSP≥50m EUR 또는 적자 시 MHI 관여, Remark A+B)', '1')
a('1.2', T, '1. Project scope',
  'Semi Turnkey (supply of technological equipment / components including advisory / supervisory services, installation execution) acc. to HQ-061-00.',
  '세미턴키인가? (기술 기자재 공급 + 자문/감리 서비스 + 설치 시공 수행)',
  '설치 시공까지 PT 책임이면 세미턴키. 기자재 공급+감리만 하는 Process Turnkey보다 시공 리스크(현장 인력·안전·공정)가 추가됨. 리뱀프가 설치를 포함하면 1.2와 1.3에 동시 해당.',
  '압연기 설비 공급에 더해 현지 설치공사(Erection)까지 PT가 원청으로 수행하는 계약.',
  'TSP ≤25m → RC3 / >25m → RC2 / ≥100m → RC1 (≥50m은 MHI 관여)', '2')
a('1.3', T, '1. Project scope',
  'Modification / Revamp project.',
  '기존 설비의 개조/리뱀프 프로젝트인가?',
  '기존 설비·인터페이스·인수인계 지점에서 예측 불가능한 장애(도면과 실물 불일치, 노후·마모 등)가 발생할 수 있어 신설보다 리스크가 높다고 보는 항목.',
  '가동 중인 열연공장의 구동·제어 설비 교체 — 기존 배선·기초가 도면과 다를 리스크.',
  'TSP ≤25m → RC3 / >25m → RC2', '3')
a('1.4', T, '1. Project scope',
  'Prototype according to TLoA.',
  'TLoA 기준 프로토타입(상용 실적 없는 최초 적용 기술)이 포함되는가?',
  '검증되지 않은 기술의 최초 적용은 비용 초과·공기 지연으로 이어질 기술 리스크. 프로토타입 상세 리스크는 TLoA에서 CTO 승인을 받아야 하며, LoA에서도 최고 등급으로 상신.',
  '상용 레퍼런스가 없는 신형 냉각 기술을 고객 라인에 처음 적용하는 경우.',
  '해당 시 RC1 + MHI 관여 (Remark A+B)', '4')
a('1.5', T, '1. Project scope',
  'Completeness clause without reference to clear battery limits.',
  '명확한 배터리 리밋(경계) 없이 완전성 조항(completeness clause)이 있는가?',
  '완전성 조항 = 계약서에 명시되지 않았어도 계약 목적 달성에 필요한 공급·서비스는 전부 PT 의무라는 조항. 경계가 없으면 미산정 스코프가 전부 비정합 비용(NCC)이 됨. 최소한 지역적(현장 구역) 또는 기술적(공정·유닛) 배터리 리밋을 정의해 추가 소요를 산정 가능하게 해야 함.',
  '"플랜트 완전 가동에 필요한 일체"를 공급 의무로 규정하면서 경계 정의가 없는 턴키 계약.',
  '배터리 리밋 없으면 RC2 (Remark B)', '-')
a('1.6', T, '1. Project scope',
  'Unclear scope of training program (location, period, duration, number of trainee and trainer, training criteria, etc.).',
  '교육(training) 범위가 불명확한가? (장소·기간·시간·인원·수료 기준 등)',
  '교육 범위가 모호하면 추가 교육 요구·미산정 비용·고객 인력 자격 미달 문제로 이어짐. 장소·기간·횟수·인원·기준을 계약에 명확히 특정해야 함.',
  '"운전원 교육 실시"라고만 쓰여 있고 인원·기간이 없어 고객이 반복 교육을 요구하는 경우.',
  '불명확 시 RC3 (Remark B)', '-')
a('1.7', T, '1. Project scope',
  'Unclear scope of technical advisory / supervisory services (location, period, duration, number of advisor, etc.).',
  '기술 자문(advisory)/감리(supervisory) 서비스 범위가 불명확한가? (장소·기간·MD·인원 등)',
  '자문(올바른 조언만 책임)과 감리(감독 작업의 성공적 완수까지 책임)는 책임 수준이 다름 — 구분을 명확히. man-day 수, 일일 시간, 주당 일수, 야근·교대·휴일 근무와 요율까지 특정해야 분쟁을 예방.',
  '슈퍼바이저 파견 조건이 "필요 기간"으로만 규정돼 연장 비용 부담 주체가 불명확한 경우.',
  '불명확 시 RC3 (Remark B)', '-')
# ── 2. Customer's supplies / Existing plant / Subsurface ─────
a('2.1', T, "2. Customer's supplies / Existing plant / Subsurface risk",
  "Responsibility for customer's existing installations and/or suitability of customer's existing / new plants, equipment and property to be used (e.g. local conditions, access to the site, brown field plant, running plant).",
  '고객 기존 설비, 또는 사용 예정인 고객 설비·자산의 적합성에 대한 책임을 PT가 지는가? (토양/지반 리스크 포함)',
  'PT 프로젝트 수행에 필요한 고객측 공급·기존 설비에 대한 책임은 회피가 원칙. 특히 토양·지반 리스크는 예측 불가한 고액 추가비용으로 이어질 수 있어 한 단계 더 높은 승인 필요.',
  '브라운필드 현장에서 기존 기초의 강도 부족이 발견되면 그 보강까지 PT 부담으로 규정된 계약.',
  '고객 공급·기존설비 책임 → RC3 / 토양·지반 리스크 → RC2', '-')
a('2.2', T, "2. Customer's supplies / Existing plant / Subsurface risk",
  "PT's responsibility for obtaining required permit(s) from any authority which would be in the customer's responsibility.",
  '본래 고객 책임인 관청 인허가(건설허가·가동허가 등)를 PT가 취득할 책임을 지는가?',
  'PT 통상 업무 관련 인허가(수출허가·비자·CE 등)는 원래 PT 몫이라 승인 불요. 그러나 건설허가·조업허가처럼 고객이 신청 주체인 인허가는 소요 시간·비용 예측이 어렵고 법적으로 PT가 취득 불가능할 수도 있음.',
  '플랜트 가동허가 취득을 PT 의무로 규정 — 지연·추가비 리스크를 고객에 전가했더라도 자원 투입 리스크는 잔존.',
  '지연·추가비 리스크가 고객 부담 → RC3 / PT 부담 → RC1 (Remark A)', '-')
a('2.3', T, "2. Customer's supplies / Existing plant / Subsurface risk",
  "Offer is based on customer's as-built documentation and customer does not take the responsibility for its as-built documentation.",
  '견적이 고객 제공 as-built 도면에 기반하는데, 고객이 그 정확성에 대한 책임을 지지 않는가?',
  '리뱀프 견적은 대개 고객 as-built 자료 기반. 자료가 실물과 다르면 설계 변경·비용·공기 파급이 전부 PT 몫이 됨. PT가 현장 대조 검증을 했더라도 편차 책임은 PT에 남으므로 승인 대상.',
  '고객이 "도면은 참고용, 정확성 미보증"을 명시했는데 현장 실사 기회도 제한된 리뱀프 입찰.',
  'PT가 as-built 검증 완료 → RC3 / 충분한 검증 불가 → RC2', '-')
# ── 3. Customer's approval ───────────────────────────────────
a('3.1', T, "3. Customer's approval / Right to modify layout",
  "The customer has a right to approve documents and the contract (i) does not stipulate a maximum period / number of review cycles for customer's approval; and (ii) does not stipulate deemed approval, if the customer fails to approve within such period / number of review cycles.",
  '고객이 문서 승인권을 갖는데 (i) 검토기간·검토횟수 상한이 없고 (ii) 기간 내 미승인 시 간주승인 조항도 없는가?',
  '고객이 검토를 지연시키거나 수정 요구를 반복하면 공정 전체가 끌려다님. 승인권 자체는 거부하지 않되, (1) 명확한 검토기한(예: 14일) 또는 1~2회 검토횟수 제한 + (2) 기한 도과 시 간주승인, 두 가지 안전장치를 요구.',
  '상세설계 승인 조항에 기한·횟수 제한과 간주승인이 모두 없어 고객이 3차, 4차 코멘트를 반복하는 경우.',
  '안전장치 없으면 RC2 (기한·횟수 기재)', '5')
a('3.2', T, "3. Customer's approval / Right to modify layout",
  'The contract allows the customer to change / modify the layout or technical concept after the signature of the contract without adequate cost compensation and extension of time.',
  '계약 서명 후 고객이 적정한 비용 보상·공기 연장 없이 레이아웃/기술 컨셉을 변경할 수 있는가?',
  '"계약은 지켜져야 한다"가 원칙. 서명 후 일방 변경권은 사소한 변경이라도 큰 비용·공기 파급을 낳으므로, 반드시 보상+공기연장이 조건이어야 함.',
  '서명 후 고객이 설비 배치 변경을 지시하면서 "사양 범위 내"라며 추가비를 인정하지 않는 경우.',
  '해당 시 RC1 (Remark A)', '6')
# ── 4. Time schedule ─────────────────────────────────────────
a('4.1', T, '4. Time schedule',
  'The project requires commencement of the project work (start engineering, start procurement of materials etc.) prior to conclusion of the contract without payment security.',
  '계약 체결 전에 지급 담보 없이 선(先)착수(엔지니어링 개시, 자재 조달 등)가 필요한가?',
  '계약이 최종 불발되면 이미 투입한 비용을 회수 못 할 수 있음. 원칙은 선급금 또는 담보 수령 후 착수. 선착수가 불가피하면 의무 부담액 규모에 따라 승인 등급이 갈림. 서명 가이드라인(HQ-004-02)상 별도 승인도 필요할 수 있음.',
  'LOI만 받은 상태에서 장납기 품목(대형 주단조품 등)을 먼저 발주하는 경우.',
  'PT 부담 ≤350k EUR & 장납기 없음 → RC3 / >350k 또는 장납기 발주 → RC2 / 비용 회수권 자체가 없음 → RC1 (Remark A)', '7')
a('4.2', T, '4. Time schedule',
  'Tough time schedule shorter than reference projects despite all possible mitigation measures. (Consult with SC.)',
  '모든 완화 조치를 강구해도 레퍼런스 프로젝트 대비 빠듯한(달성 곤란한) 공정인가?',
  '무리한 일정 수락은 지연·LD로 직결. 유사 실적 프로젝트를 벤치마크로 실현 가능성을 평가하고, 반드시 SC(공급망관리)와 공동 평가.',
  '동급 리뱀프의 실적 공기가 10개월인데 고객이 6개월을 요구하는 경우.',
  '해당 시 RC2 (공정표 제출)', '8')
a('4.3', T, '4. Time schedule',
  "Responsibility of PT for delays which are not caused by PT and/or PT's subcontractors or consortium partners (e.g. caused by customer, third party, Force Majeure).",
  'PT(하도급·컨소시엄 포함) 귀책이 아닌 지연(고객·제3자·불가항력 원인)까지 PT가 책임지는가?',
  '지연 원인은 ①PT 영역 ②고객 영역 ③양쪽 모두의 밖(제3자·FM) 세 가지. "고객 귀책 지연만 면책"이라는 문구는 ③영역 지연까지 PT 부담이 되는 함정. PT는 자기 영역 지연만 책임지는 것이 원칙.',
  '항만 파업으로 자재 입고가 늦어졌는데 면책 사유가 "고객 귀책"으로만 한정돼 PT가 지체상금을 무는 경우.',
  '해당 시 RC1 (Remark A)', '9')
# ── 5. PT work load ──────────────────────────────────────────
a('5.1', T, '5. PT work load (lack of required resources)',
  'Lack of required resources / personnel (in quantity, in quality, in time) which may jeopardize the fulfillment of contractual obligations (including factory load).',
  '계약 이행을 위협할 수 있는 자원·인력 부족(양·질·시기, 공장 부하 포함)이 있는가?',
  '계약 조건이 아니라 PT 내부 역량을 묻는 항목. 입찰 단계부터 자원계획이 필수이며, 부족이 계약 불이행으로 이어질 수 있는 수준이면 경영진에 보고.',
  '동시기 대형 프로젝트가 겹쳐 시운전 엔지니어를 배정 못 할 우려가 있는 경우.',
  '해당 시 RC2 (Remark B)', '10')
# ── 6. Subcontractors ────────────────────────────────────────
a('6.1', T, '6. Availability / Selection of subcontractors',
  'Restriction in choice of / sourcing from subcontractors (restriction beyond pre-agreed supplier list) or insufficient supplier list.',
  '하도급/조달처 선택 제한(사전 합의 벤더리스트 초과 제한) 또는 벤더리스트 부실이 있는가?',
  'PT는 글로벌 자유 소싱(최소한 사전 합의 벤더리스트)이 원칙. 조달처 제한은 비용 초과·품질 저하·핵심 벤더 과부하 지연으로 이어짐. 특정국 거래 금지(보이콧) 조항은 아예 수용 불가.',
  '고객이 자국산 우선 조달을 강제하거나 승인 벤더가 1~2개사뿐인 경우.',
  '해당 시 RC3', '-')
a('6.2', T, '6. Availability / Selection of subcontractors',
  'Insufficient availability of subcontractors for required supplies and/or services (market conditions).',
  '(시장 상황상) 필요한 공급·서비스의 하도급 가용성이 부족한가?',
  '벤더 시장이 타이트하면 6.1과 같은 리스크(지연·가격 급등)가 발생. 입찰 시점의 시장 상황 평가 항목.',
  '대형 주물 공급사 슬롯이 전 세계적으로 소진돼 납기 확보가 불투명한 경우.',
  '해당 시 RC3 (Remark B)', '-')
a('6.3', T, '6. Availability / Selection of subcontractors',
  "Any material concern on major subcontractors' credit rating. Any material concern related to major subcontractors' past performance, delivery record and future work load.",
  '주요 하도급사의 신용등급, 또는 과거 이행실적·납기실적·향후 부하에 중대한 우려가 있는가?',
  '주요 벤더의 파산은 엔지니어링·출하 지연 등 연쇄 파급을 낳고, 단기간에 동등한 대체 벤더를 못 구하면 공정 전체가 무너짐. 그래서 최고 등급.',
  '핵심 감속기 공급사가 법정관리설이 도는 상태에서 대체 소스가 없는 경우.',
  '우려 있으면 RC1 (Remark A+B)', '-')
a('6.4', T, '6. Availability / Selection of subcontractors',
  'SP > 25 m EUR and no Sourcing Concept (e.g. Calcis S138) as aligned with SC.',
  'SP 25m EUR 초과인데 SC와 정렬된 소싱 컨셉(예: CALCIS S138)이 없는가?',
  '소싱 컨셉은 입찰 프로세스(HQ-023-01)의 일부로, 통상 CALCIS S138 리포트 기반(타 지역은 다른 기반도 가능). 대형 프로젝트에서 조달 전략 없이 입찰하는 것 자체가 리스크.',
  'SP 30m 프로젝트를 소싱 컨셉 없이 견적 마감하려는 경우.',
  '해당 시 RC2', '-')
# ── 7. Inspection / Transportation ───────────────────────────
a('7.1', T, '7. Inspection / Transportation',
  'Any material concern related to customer / third party inspection prior to shipment and/or arrival at destination (e.g. customer may deny issuance of inspection certificate at its sole discretion, etc.).',
  '출하 전/도착 시 고객·제3자 검사와 관련한 중대한 우려가 있는가? (예: 검사증명서 발행이 고객 단독 재량)',
  '고객 검사는 조기 결함 발견에 도움도 되지만, 출하 가부의 최종 판단은 PT에 있어야 함. 특히 검사증명서(IC/ICC)가 L/C 네고 서류인 경우, 발행이 고객 단독 재량이면 고객이 (고의로) 발행을 보류해 대금 회수가 막힐 수 있음.',
  '공장 입회검사 합격에도 고객이 증명서 서명을 미뤄 출하·네고가 지연되는 구조.',
  '중대 우려 있으면 RC3', '-')
a('7.2', T, '7. Inspection / Transportation',
  'Any material concern related to transportation and/or customs clearance (e.g. high demurrage, cumbersome customs documents, etc.).',
  '운송·통관 관련 중대한 우려(고액 체선료, 까다로운 통관 서류 등)가 있는가?',
  '통상 운송·통관 요건은 물류 파트너와 함께 산정 가능. 그러나 신규·특수 지역은 사전 정보 확보가 어려워 잔존 우려가 있으면 보고.',
  '내륙 운송 인프라가 열악한 신규 국가 현장으로 초중량물을 운송해야 하는 경우.',
  '중대 우려 있으면 RC3', '-')
# ── 8. Acceptance ────────────────────────────────────────────
a('8.1', T, '8. Acceptance of plant',
  'Obligation to reach the performance parameters without adequate acceptance procedure (e.g. no repeated test runs allowed, testing period too short, etc.).',
  '성능 파라미터 달성 의무는 있는데 적정한 인수(성능시험) 절차가 없는가?',
  '인수는 위험이전·보증개시·잔금·현장철수가 걸린 최대 마일스톤. 절차의 최소요건 4가지 — ①객관적 판정 기준 ②충분한 시험 시간 ③결함 보완 기회 ④보완 후 재시험 허용 — 이 없으면 escalation.',
  '성능시험 1회 실패 시 재시험 없이 곧바로 불합격 처리하는 조건.',
  '최소요건 미충족 시 RC2', '11')
a('8.2', T, '8. Acceptance of plant',
  'No "deemed acceptance clause" which covers at least: (i) delay due to reasons not attributable to PT; and (ii) plant is put into commercial operation by customer.',
  '(i) PT 귀책 아닌 사유로 성능시험을 못 하는 경우와 (ii) 고객이 플랜트를 상업 가동한 경우를 커버하는 간주인수 조항이 없는가?',
  '고객이 자기 생산을 우선해 성능시험을 계속 미루면, 고객은 새 설비로 돈을 벌면서 PT의 인수·대금만 무기한 지연되는 구조. 예정일 후 일정기간 경과 또는 상업운전 개시 시 인수로 간주하는 조항이 방어책.',
  '고객이 특정 강종 생산 주문을 소화하느라 성능시험용 강종 전환을 거부하며 인수를 보류하는 경우.',
  '간주인수 없으면 RC2', '12')
a('8.3', T, '8. Acceptance of plant',
  'No "deemed acceptance clause" for completion of the works which triggers (i) start of warranty period; (ii) transfer of risk; and (iii) payment for acceptance.',
  '간주인수가 있어도 (i) 보증기간 개시 (ii) 위험 이전 (iii) 인수 대금 지급을 트리거하지 않는가?',
  '간주인수는 실제 인수를 대체하는 것이므로 효과도 같아야 함. 세 효과가 모두 따라오지 않으면 반쪽짜리. (위험이전이 Incoterms로 이미 더 일찍 일어나면 (ii)는 논점 아님.)',
  '간주인수가 인정돼도 보증기간은 "실제 인수증명서 발행일"부터 기산한다고 규정된 계약.',
  '해당 시 RC2', '13')
a('8.4', T, '8. Acceptance of plant',
  '"Deemed acceptance" payment is linked to a certificate that is required to be issued or signed by the customer (or a third party appointed by the customer).',
  '간주인수 대금 지급이 고객(또는 고객 지정 제3자)이 발행/서명하는 증명서에 연동되는가?',
  '고객이 인수를 방해하는 상황이라면 증명서 발행에도 협조하지 않는 것이 보통. 간주인수 대금을 L/C에서 회수하려면 고객 관여 없이 네고 가능한 서류 구조여야 함.',
  'L/C 네고 서류에 "고객 서명 인수증명서"가 있어 고객이 서명을 미루면 대금 회수가 막히는 경우.',
  '해당 시 RC2 (담보가 없어 지급이 고객 의사에 달린 경우도 RC2)', '14')
# ── 9. Spare and wear parts ──────────────────────────────────
a('9.1', T, '9. Spare and wear parts',
  "Availability of spare and wear parts required for testing and operation of the plant until acceptance (PAC or FAC) is not ensured (e.g. such parts are not considered at all or are not within PT's scope and PT is not entitled to use customer's spare and wear parts).",
  '인수(PAC/FAC)까지의 시험·운전에 필요한 예비품·마모품 확보가 보장되지 않는가?',
  '시운전~인수 기간에 예비품·마모품이 없으면 설비 정지·성능시험 실패로 직결. PT 공급범위 또는 고객 의무로 확보하고, 고객 재고 사용권을 규정해야 함. (사용분 정산: PT 책임 시운전 중 소모는 PT, 고객 귀책 소모는 고객 부담.)',
  '시운전 중 마모 롤 교체가 필요한데 예비품이 PT 스코프에도 고객 재고에도 없는 경우.',
  '확보 안 되면 RC3', '-')
a('9.2', T, '9. Spare and wear parts',
  "PT's obligation for availability of (i) spare parts > 10 years, or (ii) spare parts for equipment with a life cycle ≤ 3 years > 3 years from acceptance, without right to provide alternative solution.",
  '대체품 제공 권리 없이 (i) 예비품 공급보장 10년 초과 또는 (ii) 수명 3년 이하 장비(IT 등)에 3년 초과 보장인가?',
  '서브서플라이어의 단종·폐업·파산으로 장기 공급보장을 못 지킬 리스크. 기간을 관리 가능한 수준으로 제한하거나, 동등 사양 대체품을 제공할 권리를 확보해야 함.',
  '제어용 IT 하드웨어에 15년 예비품 공급 보장을 요구받는 경우.',
  '해당 시 RC3', '15')
# ── 10. Plant safety ─────────────────────────────────────────
a('10.1', T, '10. Plant safety',
  "PT is responsible for declaring that anything beyond PT's scope of work is safe (e.g. customer's existing / old equipment).",
  'PT 공급범위 밖(고객 기존·노후 설비 등)의 안전까지 PT가 선언/책임지는가?',
  '안전사고는 경제적 책임뿐 아니라 임직원 형사책임까지 갈 수 있는 민감 영역. 통제할 수 없는 범위의 안전책임은 절대 회피 대상.',
  '개조 대상이 아닌 기존 라인 전체에 대한 안전 적합 확인서를 PT 명의로 발행하라는 요구.',
  '해당 시 RC1 (Remark A)', '16')
# ── 11. Local laws ───────────────────────────────────────────
a('11.1', T, '11. Local laws / Standards / General quality criteria',
  "Challenging mandatory laws, regulations, governmental requirements and/or restrictions, norms and technical standards (e.g. DIN, ASME, NAFTA, country specific standards, customer's internal standards imposed in the contract, etc.) to be complied with by PT.",
  '준수 의무가 있는 강행 법규·규격·고객 내부표준이 평가·산정이 어렵거나 기술적으로 도전적인가?',
  '계약에 편입된 표준은 미준수 시 곧 계약위반(시정·배상·해지). 기술적·상업적으로 충분히 평가된 표준만 수용해야 하며, 평가 불가하거나 충족이 어려우면 escalation.',
  '고객 내부표준이 국제 규격보다 엄격한 재질·검사 기준을 요구해 비용 산정이 불확실한 경우.',
  '해당 시 RC3', '17')
# ── 12. TLoA ─────────────────────────────────────────────────
a('12.1', T, '12. Remaining technical risks acc. to TLoA',
  'Highest remaining General Risk acc. to TLoA.',
  'TLoA 잔여 General Risk(자원·일정·현지표준 등) 최고 등급은?',
  '기술 LoA(TLoA)의 General Risk 평가 결과를 상업 LoA에 전기하는 항목. TLoA는 Offer Approval LoA의 필수 문서 — 결과서를 반드시 입수해 함께 보관. 잔여 리스크가 없으면 N/A.',
  'TLoA에서 현지 표준 대응 리스크가 medium으로 남음 → RC2 체크.',
  'No → N/A / Low → RC3 / Medium → RC2 / High → RC1 (Remark A)', '18·19')
a('12.2', T, '12. Remaining technical risks acc. to TLoA',
  'Highest remaining Conceptual Risk acc. to TLoA.',
  'TLoA 잔여 Conceptual Risk(제품·플랜트 영역 컨셉 리스크) 최고 등급은?',
  'TLoA 두 번째 카테고리인 Conceptual Risk(제품/설비 구성 컨셉의 기술 리스크)를 전기. (세 번째 카테고리인 성능 파라미터 리스크는 26.1에서 다룸.)',
  'TLoA에서 신규 설비 구성에 conceptual risk low 판정 → RC3.',
  'No → N/A / Low → RC3 / Medium → RC2 / High → RC1 (Remark A)', '20·21')

# ── 13. Price ────────────────────────────────────────────────
a('13.1', C, '13. Price',
  'Total Sales Price (TSP).',
  '총판매가격(TSP)이 얼마인가?',
  '금액이 클수록 공급범위·인터페이스가 늘어 리스크 발생 확률과 파급이 커지므로, 가격 자체가 승인 등급을 결정. 풀/세미턴키 TSP≥50m EUR은 MHI 관여, TSP≥50bn JPY는 별도 MHI 리스크 리뷰(HQ-003-00).',
  'TSP 30m EUR 프로젝트 → RC2 상신.',
  '≤25m → RC3 / >25m → RC2 / ≥100m → RC1 (Remark A+B)', '22')
a('13.2', C, '13. Price',
  'Price breakdown towards customer does not reflect the calculation.',
  '고객 제출 가격 내역(price breakdown)이 실제 원가 계산과 다르게 구성돼 있는가?',
  '스코프 변경 시 내역 단가 기준으로 증감 정산하는데, 내역이 실제 계산과 다르면 불리한 정산이 발생. 예: 엔지니어링비를 기자재 항목에 얹어 놨는데 기자재가 50% 감액되면, 그대로 수행해야 하는 엔지니어링비까지 깎임.',
  '전략적으로 특정 항목을 낮게, 다른 항목을 높게 쓴 내역서를 제출하는 경우.',
  '해당 시 RC3', '-')
# ── 14. Project income ───────────────────────────────────────
a('14.1', C, '14. Project income',
  'Project income in relation to SP.',
  '프로젝트 수익률(Project income/SP)은 얼마인가?',
  'Project income은 EBIT에서 현금흐름 금융손익과 외국 소득세(CIT/WHT)를 반영한 최종 수익성 지표(HQ-031-03에서 산출, 정부 보조금은 제외). 수익이 낮을수록 상위 승인. 적자면 RC1+MHI, Gross loss ≥1bn JPY는 별도 MHI 리스크 리뷰.',
  'PI 2.5% 프로젝트 → RC2 상신.',
  '≥3% → RC3 / <3% → RC2 / <0% → RC1+MHI (Remark A, *Remark B)', '23·24')
# ── 15. Warranty ─────────────────────────────────────────────
a('15.1', C, '15. Warranty obligations',
  'Base warranty period > 24 months after acceptance.',
  '기본 하자보증기간이 인수 후 24개월을 초과하는가?',
  '24개월이 업계 표준(PT 표준 T&C는 12개월). 기간이 길수록 "인도 시 이미 있던 결함인지, 고객 운영 탓인지" 입증이 어려워짐. 수리 시 보증이 연장·재기산되는 점도 감안. ("guarantee"라는 표현도 LoA 목적상 warranty와 동일 취급.)',
  '고객이 36개월 기본 보증을 요구하는 경우 → RC2.',
  '>24~36개월 → RC2 / >36개월 → RC1 (Remark A, 실제 기간 기재)', '25')
a('15.2', C, '15. Warranty obligations',
  'Explicit extension of or carve out from the base warranty period for steel structures and/or other equipment / items typically having a lifetime > 24 months.',
  '철구조물 등 수명이 긴 품목에 대해 기본 보증기간의 명시적 연장 또는 별도 예외(carve-out)가 있는가?',
  '수명이 긴 스코프(철구조물·코팅 등)에 대한 보증 연장 요구는 흔하지만 여전히 승인 대상 리스크.',
  '철구조물에 5년 보증을 요구받는 경우 → RC3.',
  '연장 ≤5년 → RC3 / >5년 또는 carve-out → RC2 (예외·기간 기재)', '26')
a('15.3', C, '15. Warranty obligations',
  'Base warranty period without (i) latest start and/or expiry date; or (ii) link to deemed acceptance. (At least one of these two precautions must be fulfilled.)',
  '기본 보증기간에 (i) 최종 개시/만료일(latest date)도 (ii) 간주인수 연동도 없는가?',
  '보증 개시점(예: 인수)이 고객 손에 달려 있으면 보증기간이 무한정 뒤로 밀림. "늦어도 계약발효 후 48개월에 만료" 같은 최종일 또는 간주인수 연동, 둘 중 하나는 반드시 확보.',
  '고객이 인수를 미루는 동안 보증 개시도 계속 밀려 총 노출 기간이 늘어나는 경우.',
  '두 장치 모두 없으면 RC2', '27')
a('15.4', C, '15. Warranty obligations',
  'Warranty period for repaired / exchanged parts without latest date.',
  '수리·교체품에 대한 보증에 최종 종료일(cut-off)이 없는가? (체인 보증)',
  '수리→새 보증→또 수리→또 보증으로 이어지는 무한 "체인 보증" 리스크. 해당 부품만 재기산되면 small chain, 플랜트 전체가 재기산되면 big chain. 모든 연장 보증에 cut-off date를 두는 것이 목표.',
  '교체품에 "신품과 동일 조건의 보증"을 무한 반복 적용하는 조항.',
  'Small chain → RC3 / Big chain → RC2', '28')
a('15.5', C, '15. Warranty obligations',
  'Customer decides how to rectify defects at its sole discretion.',
  '하자 구제 방법(수리/교체/배상)을 고객이 단독 재량으로 결정하는가?',
  '구제 방법 선택권은 PT가 갖거나 최소한 상호 합의가 원칙. 고객 단독결정이면 수리로 충분한 하자에 신품 교체 등 과중한 방법을 강요당할 수 있음.',
  '경미한 하자에도 고객이 유닛 전체 신품 교체를 단독 지정할 수 있는 조항.',
  'PT 거부권 없이 고객 단독결정이면 RC2', '29')
a('15.6', C, '15. Warranty obligations',
  'Unusual warranty obligations (e.g. latent / hidden defect clause, implied warranties, design liability, availability guarantee, fit for purpose obligation, etc.).',
  '비통상 보증 의무(잠재하자 조항, 묵시적 보증, 설계책임, 가동률 보장, 목적적합성 보증 등)가 있는가?',
  '산정 불가능한 리스크를 낳는 특수 보증들. 가동률 보장은 판정 산식과 PT 무관 요인(예비품 미보유 정지 등) 배제를 반드시 명시, fit-for-purpose는 "목적"의 구체 특정이 필수. Legal과 공동 평가 권장. (잠재하자 보증은 통상 보증기간 내로 한정되면 no risk.)',
  '보증기간 전체에 걸친 가동률 보장이나 명시 없는 묵시적 성능 보증을 요구받는 경우.',
  '해당 시 RC2', '30')
# ── 16. External consortium ──────────────────────────────────
a('16.1', C, '16. External consortium',
  'Joint and several liability towards the customer.',
  '(외부 컨소시엄) 고객에 대한 연대책임(joint & several)을 지는가?',
  '오픈 컨소시엄에서는 각 멤버가 전체 스코프에 연대책임 — 파트너가 불이행·파산하면 그 몫을 PT가 떠안음. 리스크 노출 파악을 위해 컨소시엄 지분(%)을 코멘트에 기재.',
  '토건 담당 파트너가 파산해 그 잔여 공사와 배상까지 PT가 인수하게 되는 경우.',
  '해당 시 RC3 (Remark B, 지분 기재)', '31')
a('16.2', C, '16. External consortium',
  'Financial strength of external consortium partner(s) acc. to HQ-045-01 (in alignment with FS FIN).',
  '외부 컨소시엄 파트너의 재무 건전성은 어느 수준인가? (FS FIN 협의)',
  '연대책임 구조에서 파트너의 신용은 곧 PT의 리스크. FS FIN SF와 FRMC(금융리스크관리개념)에 따라 평가하고 결과를 코멘트에 기재.',
  '파트너 신용평가 결과 weak → RC2 체크.',
  'Strong → N/A / Moderate → RC3 / Weak → RC2 / Insufficient → RC1 (Remark A+B)', '32')
a('16.3', C, '16. External consortium',
  'Any other serious concerns regarding the capability of consortium partner(s).',
  '컨소시엄 파트너의 (재무 외) 기술·조직 역량에 중대한 우려가 있는가?',
  '노하우·생산능력·레퍼런스 등 수행 역량을 철저히 검증. GBU/Segment와 협의하고, 세미/풀턴키면 TK BP와도 정렬.',
  '파트너가 해당 설비 시공 실적이 전무한 경우.',
  '우려 있으면 RC1 (Remark A+B)', '33')
a('16.4', C, '16. External consortium',
  'PT is liable for damages attributable to the external consortium partner(s) via socialization clause in the consortium agreement.',
  '컨소시엄 내부 계약의 손해 분담(socialization) 조항으로 파트너 귀책 손해(LD 등)를 PT가 분담하는가?',
  '원칙은 각자 자기 스코프 책임. socialization은 귀책 파트너의 LD를 무귀책 파트너도 합의 산식으로 분담하는 조항 — 무귀책자에게 부담 전가.',
  '파트너 지연으로 발생한 LD를 지분 비율대로 공동 부담하기로 한 컨소시엄 계약.',
  '해당 시 RC2', '34')
a('16.5', C, '16. External consortium',
  'Unusual risks under / in the consortium agreement (e.g. liability of external consortium partner(s) towards PT is very limited, unfavourable weight clause, no proper applicable law and/or arbitration clause).',
  '컨소시엄 계약상 기타 비통상 리스크(파트너의 대(對)PT 책임 과소 제한, 불리한 weight clause, 준거법·중재 조항 부적정 등)가 있는가?',
  '파트너 책임한도가 너무 낮으면 고객이 PT에 청구한 손해를 파트너에게 구상 못 함. weight clause(자재 중량 변동 시 정산 조항)가 PT에 불리하게 설계된 경우 등도 포함.',
  '파트너의 대PT 책임한도가 자기 지분의 10%로 제한돼 있는 경우.',
  '해당 시 RC2', '35')
# ── 17. Strategic implications ───────────────────────────────
a('17.1', C, '17. Strategic implications',
  'Equity participation in projects.',
  '프로젝트에 지분 참여(equity participation)를 하는가?',
  'PT그룹사가 플랜트 구매자(고객사)의 지분을 인수하는 것. 설비 공급자를 넘어 투자자가 되는 구조로, 지분투자 가이드라인(HQ-028-00) 적용 대상이자 최고 등급.',
  '고객이 신설 법인 지분 10% 인수를 수주 조건으로 요구하는 경우.',
  '해당 시 RC1 (Remark A+B)', '-')
a('17.2', C, '17. Strategic implications',
  'Build operate transfer (BOT) / Build own operate (BOO) or similar types of projects.',
  'BOT/BOO(건설-운영-이전/건설-소유-운영) 또는 유사 유형 프로젝트인가?',
  '특수목적법인이 자금조달·건설·운영(통상 20~30년)까지 맡아 투자 회수하는 모델. PT가 플랜트 소유자·운영자 리스크까지 지게 되어 통상 사업범위 밖 — HQ-028-00 적용, 최고 등급.',
  '고객이 설비 구매 대신 "PT가 짓고 운영하며 톤당 요금으로 회수"하는 구조를 제안하는 경우.',
  '해당 시 RC1 (Remark A+B)', '-')
a('17.3', C, '17. Strategic implications',
  'Strategic alliances in projects.',
  '프로젝트와 관련한 전략적 제휴(strategic alliance)가 있는가?',
  '특정 프로젝트나 분야(제품·시장·개발)에서 타사와 활동을 조율·협력하기로 하는 약정. 행동의 자유를 제약하고 반독점법(anti-trust) 이슈에 특히 유의해야 하므로 최고 등급.',
  '경쟁사와 특정 시장 공동 대응·기술 공동 개발을 약정하며 입찰하는 경우.',
  '해당 시 RC1 (Remark A+B)', '-')
a('17.4', C, '17. Strategic implications',
  'Non-Compete Obligation.',
  '경업금지 약정(Non-Compete Obligation)이 있는가?',
  'MHI그룹 외부 회사에 "특정 제품·기술·서비스 사업을 다른 회사와 하지 않겠다"고 약속하는 것. 기간 불문 최고 등급이며, PT Head of Legal + MHI 추가 승인이 의무(HQ-008-02).',
  '고객이 "이 기술을 경쟁 제철사에 3년간 공급하지 않는다"는 조항을 요구.',
  '해당 시 RC1 + PT Legal 총괄·MHI 승인 필수 (Remark A)', '71')
# ── 18. Customer's changes ───────────────────────────────────
a('18.1', C, "18. Customer's changes / Variations in PT's scope",
  "The contract requires PT to proceed with change requests without written agreement (price, schedule, performance guarantees, etc.) and/or without acceptable procedures.",
  '가격·공정·성능보증 등에 대한 서면 합의 또는 수용 가능한 절차 없이 변경 지시에 착수해야 하는가?',
  '스코프 변경은 공기·비용·성능·보증에 연쇄 파급. 사전 서면 change order가 원칙이고, 최소한 "변경 실행 시 가격·공정 등의 적정 조정을 명문으로 보장"하는 절차가 있어야 수용 가능.',
  '고객 구두 지시로 먼저 시공하고 정산은 사후 협의하는 관행을 계약이 강제하는 경우.',
  '해당 시 RC2', '36')
# ── 19. Payment ──────────────────────────────────────────────
a('19.1', C, '19. Terms of payment / Payment security',
  "(i) Unclear definition of payment due date. (ii) Entitlement for payment is triggered by activities / items outside PT's control. (iii) No entitlement to claim interest for late payment.",
  '(i) 지급기일 불명확 (ii) 지급 트리거가 PT 통제 밖 (iii) 지연이자 청구권 없음 — 셋 중 해당이 있는가?',
  '각 기성의 지급기일과 필요 서류는 명확해야 함. "고객이 엔드유저에게 받으면 지급"(pay-when-paid)류는 PT가 이행해도 못 받는 구조. 지연이자는 계약 또는 준거법상 확보돼야 함(이슬람권은 이자 금지 → 대체 보상 필요).',
  '지급 조건이 "고객 사내 승인 완료 후"로 규정돼 기일을 특정할 수 없는 경우.',
  '해당 시 RC3', '37')
a('19.2', C, '19. Terms of payment / Payment security',
  'Lack of entitlement to suspend / terminate for late payment.',
  '고객 지급 지연 시 PT의 작업 중지/해지 권리가 없는가?',
  '대금 미수령 상태로 이행을 계속하도록 강요받지 않으려면 지급 지연 시 중지·해지권이 명시돼야 함(특히 담보가 없는 계약에서 필수). 계약에도 준거법에도 근거가 없으면 escalation.',
  '고객이 기성 2회분을 연체 중인데도 공정을 멈출 계약상 근거가 없는 경우.',
  '권리 없으면 RC2', '-')
a('19.3', C, '19. Terms of payment / Payment security',
  '(i) No or insufficient payment security (e.g. lack of Letter of Credit (L/C), parent company guarantee, bank guarantee, PRI, ECA coverage); or (ii) no or insufficient consequences in case issuance of payment security is delayed.',
  '지급담보(L/C·모회사보증·은행보증·PRI·ECA)가 없거나 불충분한가? 담보 발급 지연 시 구제수단(중지·해지권)이 없는가?',
  'PT는 대금 수령 전에 상당한 이행을 먼저 하므로 담보가 원칙. 최종 지급분이 L/C에서 빠져 있는 것도 "불충분". 신용도 확인은 HQ-045-00에 따라 FS FIN SF와 정렬. 신용도가 있으면 RC2, 신용도까지 없으면 최악 시나리오로 RC1.',
  '신용도 양호한 대기업 고객이지만 잔금 10%가 무담보 후불인 경우 → RC2.',
  '고객/모회사 신용도 충분 → RC2 / 신용도 부족 → RC1 (Remark A+B, 담보 유형 기재)', '38·39')
# ── 20. Taxes ────────────────────────────────────────────────
a('20.1', C, '20. Taxes / Duties / Customs',
  'PT bears the risk of tariffs, duties, customs or other penalties for imports into customer country (e.g. DDP acc. Incoterms).',
  '고객국 수입 관세·세금·통관 비용 리스크를 PT가 지는가? (예: DDP 조건)',
  'DDP 등 수입 부담을 공급자에 지우는 Incoterms에 합의하면 예상 못 한 관세·부과금이 전부 PT 몫. 관세 정책이 급변하는 시장(특히 미국 관련 물동)은 한 단계 높은 승인.',
  '고객국 반덤핑 관세 신설로 DDP 조건의 수입 비용이 급증하는 경우.',
  '해당 시 RC3 / 미국 발(發)·향(向) 물동 포함 시 RC2', '-')
# ── 21. Transfer of title and risk ───────────────────────────
a('21.1', C, '21. Transfer of title and risk',
  'Transfer of title to the customer takes place before shipment and either (i) no payment security is in place, or (ii) < 85 % payment is received.',
  '소유권(title) 이전이 선적 전에 일어나는데 (i) 지급담보가 없거나 (ii) 수령 대금이 85% 미만인가?',
  '소유권 이전 시점은 계약에 명시해야(Incoterms는 소유권을 다루지 않음). 선적 전 이전은 담보 또는 85% 이상 수금이 전제. 선적 후 소유권 유보는 국제거래에서 실효성이 약함(준거법상 불인정, 반환 집행 곤란).',
  '기성불 조건으로 제작 진행 중 재공품 소유권이 고객에 넘어가는데 수금율은 50%인 경우.',
  '해당 시 RC2', '-')
a('21.2', C, '21. Transfer of title and risk',
  'Transfer of risk to the customer takes place later than handover.',
  '위험 이전이 핸드오버(순수 공급은 인도조건 시점, 그 외는 인수 시점)보다 늦게 일어나는가?',
  '위험 이전 = 무과실 멸실·파손 책임이 고객으로 넘어가는 시점. 공급자는 최대한 이르게 가져가는 것이 유리. 순수 기자재 공급은 Incoterms 인도 시점, 시공 포함 계약은 인수 시점이 기준선.',
  '현장 반입 후에도 인수 시까지 멸실·파손이 공급자 부담으로 남는 조항.',
  '기준보다 늦으면 RC2', '40')
# ── 22. Insurance ────────────────────────────────────────────
a('22.1', C, '22. Insurance',
  "Insurance cover not based on PT's insurance strategy (e.g. unusual insurance obligations towards the customer acc. to HQ-014-00).",
  'PT 보험 전략(HQ-014-00)에 어긋나는 보험 의무(비통상 부보 요구 등)가 있는가?',
  '고객에 대한 보험 제공 의무는 PT 책임한도와 공급범위를 초과하면 안 됨(보험관리 가이드라인 HQ-014-00). 어긋나는 요구는 FS FIN INS 서면 코멘트를 받아 기재.',
  '고객이 PT 스코프 밖 공사까지 포괄하는 CAR 보험을 PT 명의로 들 것을 요구하는 경우.',
  '해당 시 RC3 (FS FIN INS 코멘트 기재)', '-')
a('22.2', C, '22. Insurance',
  'Contractually required insurance coverage higher than limitation of liability and no clear precedence of limitation of liability cap (as aligned with FS FIN INS).',
  '계약상 요구 보험 부보액이 책임한도(LoL cap)를 초과하는데, 책임한도 우선 원칙이 명확하지 않은가?',
  '보험 부보액이 책임한도보다 크면 "실질 책임한도가 보험액 아니냐"는 해석 분쟁 소지. cap이 보험보다 우선함을 계약에 명시해야 하며, FS FIN INS와 정렬.',
  '책임한도 100%인데 계약이 200% 부보를 요구하고 우선순위 규정이 없는 경우.',
  '해당 시 RC3 (요구 부보액 기재)', '41')
# ── 23. Financial risks ──────────────────────────────────────
a('23.1', C, '23. Financial Risks',
  'Any additional financing risk exceeding the normal extent, including credit obligations or risks because of ECA financing (e.g. retainer beyond the normal extent).',
  '통상 범위를 넘는 추가 금융 리스크(신용 공여, ECA 파이낸싱 관련 리스크 등)가 있는가?',
  'ECA가 통상(5%)보다 높은 리테이너(비커버 리스크 잔존분, 예: 10%)를 조건으로 걸고 은행이 이를 인수하지 않는 경우 등. 통상 범위 내(유보금 등)는 escalation 불요 — FS FIN SF/Local FRM과 협의해 판정.',
  'ECA 커버 조건상 PT가 추가로 떠안는 리스크가 2m EUR인 경우 → RC2.',
  '≤1m EUR → RC3 / >1m EUR → RC2 / >5m EUR → RC1 (Remark A)', '-')
a('23.2', C, '23. Financial Risks',
  "Foreign exchange risk without appropriate countermeasures (e.g. hedging) in customer's / subcontractor's contract(s) acc. to HQ-022-00.",
  '고객/하도급 계약의 외환 리스크에 적절한 대책(헤징 등, HQ-022-00)이 없는가?',
  '외환 리스크는 PT 사업의 일상이지만 HQ Treasury의 대책(헤징 등)으로 커버되는 것이 전제. 가이드라인대로 헤지되지 않은 외화 익스포저는 최고 등급.',
  '수금은 USD, 주요 조달은 EUR인데 환헤지 계획이 없는 경우.',
  '해당 시 RC1 (Remark A)', '-')
a('23.3', C, '23. Financial Risks',
  'No price escalation clause towards the customer or no back to back conditions to the subcontractor(s) in case of long project duration (acceptance > 36 months after contract signature).',
  '장기 프로젝트(서명 후 인수까지 36개월 초과)인데 물가연동(escalation) 조항도, 하도급 back-to-back 고정도 없는가?',
  '장기 프로젝트는 자재·인건비 상승 리스크. 1순위는 고객향 물가연동 조항, 고객이 거부하면 2순위로 하도급 가격을 back-to-back으로 고정해 전가. 둘 다 안 되면 escalation.',
  '42개월 프로젝트에서 고정가 계약 + 벤더 견적 유효기간은 6개월뿐인 경우.',
  '둘 다 없으면 RC2 (Remark B)', '-')
a('23.4', C, '23. Financial Risks',
  'Offer validity > 3 months (e.g. from offer submission date).',
  '견적 유효기간이 3개월을 초과하는가?',
  '유효기간이 긴 확정 견적은 그 사이 임금·자재·에너지 가격 상승 리스크를 PT가 떠안는 것. PT 목표는 유효기간 3개월 이내.',
  '고객 구매 절차상 견적 유효 6개월을 요구받는 경우.',
  '>3개월이면 RC2', '-')
a('23.5', C, '23. Financial Risks',
  "Any deferred payment obligation in customer's favour > 3 months.",
  '고객에게 유리한 3개월 초과 지급 유예(deferred payment)가 있는가?',
  '유예 지급은 PT가 프로젝트를 선(先)금융하는 셈 — 현금흐름 악화와 추가 금융비용. 기준은 PT 인보이스 발행 시점부터 3개월. Standard LoA에서는 최고 등급(RC1)임에 유의.',
  '인수 확인 후 180일 뒤 지급하는 조건.',
  '해당 시 RC1 (Remark A) ※ Small LoA에서는 RC2', '42')
a('23.6', C, '23. Financial Risks',
  'PoC project acc. to HQ-065-00 with negative payment balance > 5 m EUR over an accumulated time period of 3 months acc. to HQ-023-04.',
  'PoC(진행기준 회계) 프로젝트에서 누적 3개월간 지급수지 적자가 5m EUR을 초과하는가?',
  'TSP>10m 프로젝트는 Payment Balance Sheet(HQ-023-04)가 필수. PT 지출이 수금을 초과하는 마이너스 수지가 3개월 연속 기준치를 넘으면 escalation. 예: 6→7→6m EUR 적자 3개월 → RC2.',
  '착수금이 작고 기성 지급이 후반부에 몰려 중반 수지가 -8m EUR로 3개월 지속되는 경우.',
  '>5~15m EUR → RC2 / >15m EUR → RC1 (Remark A)', '-')
a('23.7', C, '23. Financial Risks',
  'PT must provide bank or corporate guarantees (bid / advance payment / performance / warranty / retention bond, etc.) or letter(s) of credit not acceptable to FS FIN SF.',
  'FS FIN SF가 수용할 수 없는 은행·회사 보증(입찰/선수금/이행/하자/유보 본드 등) 또는 L/C를 PT가 제공해야 하는가?',
  '시장 관행보다 가혹한 조건의 보증(무조건 청구 불가 조건, 과도한 금액·기간 등)으로 PT 보증 가이드라인(HQ-016-01/02)에 안 맞는 경우. 유일하게 CEO가 아닌 CFO가 승인하는 RC1 항목.',
  '자동 연장(evergreen) + 무사유 청구 가능한 이행보증을 요구받는 경우.',
  '해당 시 RC1 (승인권자: PT CFO, Remark A)', '-')
a('23.8', C, '23. Financial Risks',
  'TSP > 10 m EUR and no Financial Risk Management Concept acc. to HQ-045-00.',
  'TSP 10m EUR 초과인데 금융리스크관리개념(FRMC, HQ-045-00)이 없는가?',
  'FRMC는 대금 리스크의 식별·평가·정량화와 커버 여부·수단을 정하는 필수 문서. 기준 충족 대형 프로젝트에서 FRMC 없이 상신하는 것 자체가 최고 등급 리스크.',
  'TSP 12m 프로젝트를 FRMC 미작성 상태로 LoA 상정하려는 경우.',
  '해당 시 RC1', '-')
# ── 24. Country risks ────────────────────────────────────────
a('24.1', C, '24. Country risks',
  'Substantial risk because of political or security situation of the country / region where the plant shall be installed and/or work / manufacturing / transport shall be performed.',
  '플랜트 설치국 또는 작업·제조·운송 수행국의 정치·치안 상황에 중대한 리스크가 있는가?',
  '안정적 환경에서의 수행이 견적의 전제. 무장충돌·소요·파업·행정기관의 자의적 조치 등 우려가 있으면 조건에 반영하고 보고.',
  '고위험국 현장 시공이 포함돼 주재원 안전·철수 리스크가 있는 경우.',
  '해당 시 RC1 (Remark A+B)', '43')
a('24.2', C, '24. Country risks',
  'PT is exposed to business risks at the country / region where plant to be installed and/or work / manufacturing / transport shall be performed such as: Transfer price; Anti-dumping tariffs; Any restriction in foreign exchange transaction; VISA / work permit restrictions.',
  '수행국의 사업 리스크 — 이전가격, 반덤핑 관세, 외환거래 제한, 비자/취업허가 제한 — 에 노출되는가?',
  '모든 인지 가능한 비용 요소는 원가와 조건에 반영되는 것이 LoA 상정의 전제. 커버 안 된 사업 리스크 우려가 있으면 전문 부서(Tax, SC/물류, Finance, HR)와 공동 평가 후 보고.',
  '수행국의 외화 송금 규제로 하도급 대금 지급이 지연될 우려가 있는 경우.',
  '해당 시 RC2 (Remark B)', '-')

# ── 25. LD for delay ─────────────────────────────────────────
a('25.1', L, '25. Liquidated damages (LD) for delay',
  'No LD protection for delay. (In case of Full Turnkey aim for LD for start up and not LD for delivery of specific items.)',
  '지연 지체상금(LD) 합의가 없는가?',
  'LD는 지연 손해를 정액화해 리스크에 상한을 만드는 "보호장치". 없으면 고객이 실손 전액(입증 기반)을 청구할 수 있음. 마일스톤 선정도 중요 — 풀턴키는 개별 납품이 아니라 start-up 기준이 적정.',
  'LD 없이 "지연으로 인한 입증 손해 배상"만 규정된 계약.',
  'LD 부재 시 RC2', '44')
a('25.2', L, '25. Liquidated damages (LD) for delay',
  'LD for delay > 10 % of SP.',
  '지연 LD 상한이 SP의 10%를 초과하는가?',
  'LD는 리스크 완화책이지만 상한이 크면 그 자체가 리스크. 통상 요율은 주당 0.5% 수준, 상한 10%가 업계 기준선.',
  '지연 LD 상한 12%를 요구받는 경우 → RC3.',
  '>10~15% → RC3 / >15% → RC2 (요율 기재)', '45')
a('25.3', L, '25. Liquidated damages (LD) for delay',
  'In general: LD cap for delay is reached in < 10 weeks / < 7 weeks.',
  '지연 LD가 상한에 도달하는 기간이 10주 미만인가?',
  '요율이 높아 상한에 빨리 도달하면 이후 고객의 해지권 등 다음 단계 리스크가 조기에 열림. 상한 도달까지의 기간을 최대화하는 것이 PT 목표. 마일스톤별 LD 각각에 적용.',
  '주당 1% × 상한 8% → 8주 만에 도달 → RC3.',
  '<10주 → RC3 / <7주 → RC2 (도달 기간 기재)', '46')
a('25.4', L, '25. Liquidated damages (LD) for delay',
  'Only applicable in case of shutdown: LD cap for delay of start-up is reached in < 30 days / < 10 days.',
  '(셧다운 수반 리뱀프에만 적용) 재기동 지연 LD 상한 도달이 30일 미만인가?',
  '가동 중 설비를 세우고 하는 리뱀프는 고객이 조업 재개에 사활을 걸므로 단기·고율 LD를 요구하는 것이 통례 — 그래서 완화된 별도 기준 적용.',
  '열연 셧다운 공사에서 LD가 20일 만에 상한 도달하는 조건 → RC3.',
  '<30일 → RC3 / <10일 → RC2 (도달 일수 기재)', '47')
a('25.5', L, '25. Liquidated damages (LD) for delay',
  'LD for delay are not sole and exclusive (financial) remedy.',
  '지연 LD가 유일·배타적 금전 구제(sole remedy)가 아닌가?',
  'sole remedy가 아니면 LD는 상한이 아니라 "최저 책임"이 됨 — LD를 물고도 그 위에 실손 배상이 얹힘. 준거법에 따라 당연히 보장되지 않으므로 명시 조항이 필요.',
  'GTC에 "지체상금과 별도로 손해배상을 청구할 수 있다"고 규정된 경우.',
  'sole remedy 명시 없으면 RC2', '48')
a('25.6', L, '25. Liquidated damages (LD) for delay',
  '"Time is of the essence" clause / "Fixgeschäft" giving the customer an immediate termination right in case of delay without agreement of LD for delay.',
  '단기 지연도 중대 위반으로 보아 즉시 해지권을 주는 "기한 엄수 본질(time is of the essence)" 조항이 있는가?',
  '통상은 LD 상한 도달 전까지 고객이 해지 못 함. 이 조항이 있으면 하루 지연도 즉시 해지 사유가 될 수 있어, LD 우선 적용이 명확히 규정되지 않는 한 고위험.',
  '영미법계 계약서에 "time is of the essence"가 들어 있는데 지연 LD 합의가 없는 경우.',
  '해당 시 RC1 (Remark A)', '49')
# ── 26. LD for non-performance ───────────────────────────────
a('26.1', L, '26. Liquidated damages (LD) for non-performance',
  'No LD model which covers low, medium or high risk performance parameters acc. to TLoA.',
  'TLoA상 저·중·고 위험 성능 파라미터를 커버하는 성능 LD 모델이 없는가?',
  '성능 LD가 없으면 미달 시 절대적 make-good 의무(최소 수준까지 무한 보완) + 고객의 플랜트 거부권/실손배상으로 이어짐. LD 모델이 있거나 TLoA상 성능 리스크가 "no risk"면 N/A. LD 모델이 없으면 26.2·26.3은 N/A 처리하고 코멘트 기재.',
  '처리능력·품질 보증치는 있는데 미달 시 LD 테이블이 없는 계약 — TLoA 성능리스크 medium이면 RC2.',
  '(LD 모델 부재 시) TLoA Low → RC3 / Medium → RC2 / High → RC1 (Remark A+B)', '50')
a('26.2', L, '26. Liquidated damages (LD) for non-performance',
  'LD for shortfall of low, medium or high risk performance parameters acc. to TLoA are not sole and exclusive (financial) remedy.',
  '성능 미달 LD가 유일·배타적 금전 구제가 아닌가? (LD를 내고도 성능 달성 의무가 남는 등)',
  '25.5의 성능판 논리. LD가 sole remedy가 아니면 최저 책임으로 변질. 성능 LD가 sole remedy이거나, TLoA 성능 리스크가 없거나, LD 모델 자체가 없으면(26.1에 반영) N/A.',
  '성능 LD를 지급해도 "보증치 달성 시까지 개조 의무"가 별도로 남는 계약.',
  'TLoA Low → RC3 / Medium → RC2 / High → RC1 (Remark A)', '51')
a('26.3', L, '26. Liquidated damages (LD) for non-performance',
  'LD for shortfall of performance parameters > 10 % of the SP.',
  '성능 미달 LD 상한이 SP의 10%를 초과하는가?',
  '성능 LD도 합리적 범위 내 상한 설정이 업계 관행 — 25.2의 성능판.',
  '성능 LD 상한 12% 요구 → RC3.',
  '>10~15% → RC3 / >15% → RC2 (요율 기재)', '52')
# ── 27. Liability ────────────────────────────────────────────
a('27.1', L, '27. Liability',
  'Limitation of liability / indemnity obligations > 100 % of TSP.',
  '총 책임한도(LoL)가 TSP의 100%를 초과하는가? (한도 조항이 아예 없어 무제한인 경우 포함)',
  '책임한도는 프로젝트가 완전히 실패했을 때 회사 존립을 지키는 최후의 방어선. 100% cap이 업계 표준 관행. 계약에 cap 조항이 없으면 책임은 무제한 = 당연 해당. (정당한 거부에 따른 기수령 대금 반환은 손해배상이 아니므로 cap 계산에서 제외.)',
  '고객 표준약관에 책임한도 조항이 아예 없는 경우 — 무제한 책임 구조로 해당.',
  '해당 시 RC1 (Remark A+B, 실제 한도 % 기재)', '53')
a('27.2', L, '27. Liability',
  'Only applicable in case of external consortium (open consortium or PT non-leading partner in silent consortium): Limitation of liability / indemnity obligations > 100 % of SP.',
  '(외부 컨소시엄일 때만) 책임한도가 PT SP의 100%를 초과하는가?',
  '고객 계약의 cap은 전체 계약가 기준이므로, PT 몫(SP)으로 환산하면 100%를 넘게 됨. 계산례: TSP 110m, PT SP 70m, cap 100% → PT 노출 = 110×100%÷70 = 157%.',
  '오픈 컨소시엄에서 cap 100% of TSP인데 PT 지분이 60%인 경우 → PT 기준 167% → RC2.',
  '>100% → RC3 / >150% → RC2 / >170% → RC1 (실제 % 기재)', '54')
a('27.3', L, '27. Liability',
  'Explicit exclusion / carve out in the contract of IP infringement; and/or breach of confidentiality; and/or insurance proceeds; from the overall limitation of liability.',
  'IP 침해·비밀유지 위반·보험금 수령분이 총 책임한도에서 명시적으로 제외(carve-out)되는가?',
  '이 3가지는 상대적으로 발생 확률이 낮거나(IP·비밀) 보험이 커버해(보험금) RC3 수준으로 수용 가능. 단 결과손해 배제(27.5)는 유지돼야 하고, 보험금 제외는 FS FIN INS 사전 협의 필수.',
  '고객이 "IP 침해 책임은 한도 제외"를 요구하는 경우 → RC3 + 27.5 유지 확인.',
  '해당 시 RC3 (제외 항목·보험 커버 기재)', '55')
a('27.4', L, '27. Liability',
  'Any other exclusions / carve outs (except gross negligence / wilful misconduct, personal injury, product liability or matters that cannot be limited under applicable law) from the overall limitation of liability not mentioned under 27.3 leading to unlimited liability.',
  '27.3 외의 기타 제외 사유(법상 배제 불가 항목 제외)로 인해 사실상 무제한 책임이 되는가?',
  '중과실·고의·인적 손해·제조물책임 등은 어차피 법상 제한 불가라 제외해도 승인 불요. 정당 거부 시 기수령 대금 반환 제외도 승인 불요. 그 밖의 제외는 cap을 무력화하므로 최고 등급.',
  '"모든 하자보수 비용은 한도 제외"처럼 광범위한 carve-out을 요구받는 경우.',
  '해당 시 RC1 (Remark A+B, 제외 항목 기재)', '56')
a('27.5', L, '27. Liability',
  'No or insufficient exclusion of liability for: (i) indirect and consequential damages; and (ii) loss of revenue / profit; and (iii) loss of use / production.',
  '(i) 간접·결과손해 (ii) 수익/이익 상실 (iii) 사용/생산 손실에 대한 책임 배제가 없거나 불충분한가?',
  '지연·성능미달·기존설비 손상은 고객의 생산 손실→매출 손실로 이어지는데, 이 규모는 프로젝트 금액과 비교가 안 됨(제철소 조업정지 하루 손실이 계약가를 넘을 수 있음). 배제가 업계 표준이자 PT 기본 요건.',
  'GTC에 조업차질 손해 배상이 명시돼 있고 결과손해 배제 조항이 없는 경우.',
  '배제 없거나 불충분하면 RC2 (Remark B, 누락 배제 기재)', '57')
a('27.6', L, '27. Liability',
  'Aggregate cap (maximum) for LD for delay and shortfall of performance parameters > 10 % of SP.',
  '지연 LD + 성능 LD의 합산 상한(aggregate cap)이 SP의 10%를 초과하는가?',
  '개별 sub-cap(25.2·26.3)과 별도로, 두 LD를 합친 총액에도 개별 합계보다 낮은 상한을 두는 것이 목표.',
  '지연 10% + 성능 10%인데 합산 상한 15%로 합의한 경우 → RC3.',
  '>10~20% → RC3 / >20~30% → RC2 / >30% → RC1 (Remark A, 실제 % 기재)', '58')
# ── 28. Know-how / IP ────────────────────────────────────────
a('28.1', L, '28. Know-how protection / Confidentiality / IP infringement',
  'The contract requires unconditional (i) handover or (ii) grant of right of use of proprietary information / sensible know-how (e.g. source code, drawings etc. of important nature) as defined in HQ-038-00 to the customer.',
  '소스코드·핵심 도면 등 독점정보/민감 노하우(HQ-038-00)의 무조건적 인도 또는 사용권 부여를 계약이 요구하는가?',
  '독점 노하우는 PT 장기 경쟁력의 핵심 자산. 인도를 피할 수 없다면 고객의 사용 목적을 운전·정비·오버홀·복구·수리로 한정하는 조건부여야 함. 무조건 이전은 최고 등급.',
  '제어 소스코드 원본 인도를 무조건 요구받는 경우.',
  '조건 한정 불가 시 RC1 (Remark A+B)', '59')
a('28.2', L, '28. Know-how protection / Confidentiality / IP infringement',
  'Insufficient confidentiality obligations of customer regarding scope and/or duration (minimum 10 years from acceptance of the plant).',
  '고객의 비밀유지 의무가 범위 또는 기간(인수 후 최소 10년) 면에서 불충분한가?',
  'PT가 제공하는 모든 정보가 보호 대상이어야 하고 기간은 10년이 표준 타깃. PT 표준계약의 모델 조항 대비 의무를 크게 줄이는 수정은 승인 대상.',
  '비밀유지 기간을 3년으로 줄이자는 고객 요구.',
  '해당 시 RC2 (실제 기간 기재)', '60')
a('28.3', L, '28. Know-how protection / Confidentiality / IP infringement',
  'Excessive grant of rights to customer (e.g. right to use confidential information beyond operation, maintenance, overhaul, restoration and repair, or to pass it to third parties); Transfer of ownership / title of intellectual property to the customer.',
  '비밀정보의 과도한 사용권(목적 외 사용·개조·복제·제3자 제공) 또는 IP 소유권 이전이 있는가?',
  '고객의 사용권은 공급 설비의 운전·정비·오버홀·복구·수리로 한정이 원칙. 개량·복제 목적 사용, 제3자 제공, 소유권 이전은 회피 대상.',
  '고객이 도면을 자기 계열사·타 벤더에 제공할 권리를 요구하는 경우.',
  '해당 시 RC1 (Remark A+B)', '61')
a('28.4', L, '28. Know-how protection / Confidentiality / IP infringement',
  'In case of violation of third party rights by PT, PT is not entitled to control the defense and the corrective actions (e.g. pay royalty, change design, settlement) and customer is entitled to decide.',
  'PT의 제3자 권리(IPR) 침해 시 방어·시정조치(로열티 지급/설계변경/화해)의 주도권이 PT가 아닌 고객에게 있는가?',
  '침해 대응은 상황별 최적 수단(로열티·설계변경·화해)을 고를 수 있어야 비용을 통제함. 고객이 결정권을 가지면 PT에 더 부담스러운 방법을 고를 수 있음.',
  'IP 분쟁 발생 시 고객이 "설계 전면 변경"을 단독 지정할 수 있는 조항.',
  '고객 결정권이면 RC2', '62')
# ── 29. Suspension / Termination ─────────────────────────────
a('29.1', L, '29. Suspension / Termination of contract',
  'The contract allows customer to unilaterally suspend the contract for convenience (i) without full reimbursement of PT cost; and/or (ii) without defined maximum suspension period.',
  '고객이 사유 없이 계약을 중지(suspension for convenience)할 수 있는데 (i) PT 비용 전액 보전이 없거나 (ii) 최대 중지 기간 제한이 없는가?',
  '임의 중지권은 반드시 비용·경비 전액 보전이 조건. 최대 기간 제한도 필수 — 없으면 고객이 프로젝트를 무기한 표류시키면서 PT를 기존 조건(가격·벤더)에 묶어둘 수 있음.',
  '고객이 투자 재검토를 이유로 6개월째 중지시키면서 대기 비용을 인정하지 않는 경우.',
  '해당 시 RC2 (최대 중지 기간 기재)', '-')
a('29.2', L, '29. Suspension / Termination of contract',
  'The contract allows customer to unilaterally terminate the contract for convenience without full reimbursement of PT cost.',
  '고객이 사유 없이(termination for convenience) 해지할 수 있으면서, PT 기발생 비용·경비의 전액 보전 의무가 없는가?',
  '임의 해지권 자체는 흔하지만 반드시 "해지로 인한 PT의 모든 비용·경비 보전"이 명시 조건이어야 함. 보전 없는 임의해지는 최고 등급.',
  '고객 사정(투자 보류)으로 해지하면서 기성분만 정산하고 해지 비용·이익은 불인정하는 조항.',
  '해당 시 RC1 (Remark A)', '63')
# ── 30. Dispute resolution ───────────────────────────────────
a('30.1', L, '30. Applicable dispute resolution / Law',
  'Lack of dispute resolution clause (for exclusive and final settlement of disputes) acceptable to Legal.',
  'Legal이 수용 가능한 (배타적·최종적) 분쟁해결 조항이 없는가?',
  '합의 없으면 분쟁은 관할 법원행 — 법치가 불안한 국가의 법원이면 치명적. ICC 중재 등 국제적 신뢰성 있는 중재를 지향하고, 수용 가능 여부는 반드시 PT Legal과 정렬.',
  '고객국 지방법원 전속관할 조항만 있는 계약.',
  'Legal 수용 불가 시 RC2 (Remark B: ICC 중재 지향)', '64')
a('30.2', L, '30. Applicable dispute resolution / Law',
  'Lack of applicable law acceptable to Legal.',
  'Legal이 수용 가능한 준거법 합의가 없는가?',
  '국제 프로젝트는 여러 법역이 얽히므로 명시적 준거법 합의가 필요. 중립법(스위스법, 영국법 등)을 지향하고 고객국 법은 회피. Legal과 정렬 필수.',
  '준거법 조항이 아예 없거나 고객국 법으로 강제되는 경우.',
  'Legal 수용 불가 시 RC2 (중립법 지향)', '65')
# ── 31. Changes in law ───────────────────────────────────────
a('31.1', L, '31. Changes in law or regulations',
  "Responsibility of PT for consequences of changes in law, regulations, standards in customer's country after signature of contract (which may effect on time, cost, performance, payment and other terms and conditions).",
  '계약 서명 후 고객국의 법령·규격 변경에 따른 결과(공기·비용·성능·지급 영향)를 PT가 책임지는가?',
  '서명 시점까지의 법규는 견적에 반영 가능하지만, 서명 후 변경은 예측·산정 불가. 최소한 고객국의 법령 변경 리스크는 고객 책임으로 규정해야 함.',
  '서명 후 강화된 환경 규제로 추가 설비가 필요해졌는데 비용 분담 조항이 없는 경우.',
  '고객 책임 규정 없으면 RC2', '-')
# ── 32. Export Control / FM ──────────────────────────────────
a('32.1', L, '32. Export Control / Force Majeure',
  'Any concerns out of the EC checks acc. to HQ-052-00 (High Risk Country, military / nuclear related, sanctioned party, other Red Flag) and (i) no GECO approval received; or (ii) inability / concerns to comply with its preconditions.',
  '수출통제 체크(고위험국 수출, 군사·원자력 관련, 제재리스트 당사자, 기타 Red Flag)에서 우려가 있는데 GECO 승인이 없거나 승인 조건 준수가 어려운가?',
  '수출통제 위반은 수출 특권·라이선스 상실 등 회사 전체 리스크. 우려 항목이 하나라도 있으면 RECO를 통해 GECO 승인을 받아야 하며, 승인 없이는 최고 등급.',
  '엔드유저가 제재리스트 계열사로 확인됐는데 GECO 검토가 완료되지 않은 경우.',
  '해당 시 RC1 (Remark A)', '66')
a('32.2', L, '32. Export Control / Force Majeure',
  'No proper Export Control Clause excluding PT for breach of contract in case of boycott / export bans / sanctions / revoking of export license. No prohibition of re-export. No notification obligation in case of change of control. No termination right of PT in case fulfillment is affected by sanctions.',
  '수출통제 면책 조항, 재수출 금지, 고객 지배구조 변경 통지의무, 제재 시 PT 해지권 — 이런 보호 장치가 계약에 없는가?',
  '수출 규제는 PT 통제 밖 사유이므로 그로 인한 불이행은 면책돼야 함. 고객의 재수출은 새로운 수출이라 금지 명문화 필요. 각 보호장치 부재 시 escalation. 모델 조항은 PT 표준계약에 있음.',
  '고객 GTC에 수출통제 관련 조항이 전무한 경우.',
  '보호장치 부재 시 RC3 (Remark B)', '67')
a('32.3', L, '32. Export Control / Force Majeure',
  'For PT EU entities when customer is outside of EU or partner countries: No proper No-Russia / No-Belarus Clause combined with a termination right due to violation.',
  '(PT EU 법인이 EU/파트너국 밖 고객과 계약 시) No-Russia/No-Belarus 조항 + 위반 시 해지권이 없는가?',
  'EU 제재법(Council Regulation (EU) 833/2014)상 의무 조항. 적용 여부·예외는 EC 부서와 협의(생략은 EC 서면 승인 필요). 한국·일본·미국 등은 파트너국이라 통상 미적용.',
  'EU 법인이 파트너국 외 제3국 고객과 계약하는데 해당 조항이 없는 경우.',
  '해당 시 RC1 (Remark A)', '68')
a('32.4', L, '32. Export Control / Force Majeure',
  "Force Majeure defined too narrow. No time / schedule relief in case of FM. No termination right of PT if FM lasts longer than 6 months. Customer's right to terminate if FM lasts shorter than 3 months, or without obligation to pay for performed work.",
  'FM 정의가 협소한가? FM 시 공기 유예가 없는가? FM 6개월 초과 시 PT 해지권이 없는가? 고객이 3개월 미만 FM으로도 해지 가능한가? FM 해지 시 기성 대금 지급 의무가 없는가?',
  'FM은 양측 통제 밖 사유의 리스크 배분 장치. PT 표준 모델 조항 대비 보호가 부족한 수정(열거식 협소 정의, 공기 유예 없음 등)은 승인 대상.',
  'FM을 "천재지변"으로만 좁게 정의하고 효과도 일부 면제에 그치는 GTC.',
  '보호 부족 시 RC2 (Remark B)', '69')
# ── 33. Miscellaneous ────────────────────────────────────────
a('33.1', L, '33. Miscellaneous',
  "Other unusual clauses / obligations with major risks, e.g.: change of control in PT gives customer right to terminate; assignment of rights by customer without PT's consent; unreasonable notice periods; customer's right to reject PT's site personnel at any time without reason; customer's explicit right to set-off.",
  '기타 중대 리스크가 있는 비통상 조항 — 예: PT 지배구조 변경 시 고객 해지권, PT 동의 없는 고객의 권리 양도, 무리한 통지기간, 현장인력 무사유 거부권, 고객의 명시적 상계권.',
  '체크리스트가 못 담는 계약 특유의 리스크를 잡는 항목. 담당자(BPM 등)의 개별 판단으로 비통상 조항을 발굴해 기재.',
  '고객이 채권을 제3자에게 자유 양도할 수 있고, PT 대금과 타 채권을 임의 상계할 수 있는 조항.',
  '해당 시 RC2', '70')

# ── 34. Compliance ───────────────────────────────────────────
a('34.1', P, '34. Compliance risks',
  'Result of the Compliance Assessment acc. to HQ-003-06.',
  'HQ-003-06 컴플라이언스 평가 결과(신호등)는?',
  'Compliance Risk Assessment는 Standard LoA의 필수 문서. 뇌물·비즈니스 파트너 계약 미승인 등 리스크를 신호등으로 평가 — Green이면 비해당, Yellow면 RC2, Red면 LoA 상정 자체가 불가(HQ-003-00).',
  '판매 에이전트 커미션 구조에 불투명성이 있어 평가 결과 Yellow인 경우.',
  'Green → N/A / Yellow → RC2 (Red는 상정 불가)', '72')

# ── 35. Other ────────────────────────────────────────────────
a('35.1', O, '35. Other Risks (not mentioned above)',
  '( to be inserted ) — catch-all for any other risk not addressed above, together with a proper risk class evaluation.',
  '위 항목들에 없는 기타 리스크 자유기재 — 리스크 내용과 함께 RC 등급을 스스로 평가해 기재.',
  '"캐치올" 항목. 어떤 체크리스트도 모든 리스크를 못 담으므로, BPM 또는 관여 부서 판단으로 경영 승인이 필요한 미분류 리스크를 여기에 서술하고 RC3/RC2/RC1을 선택.',
  '고객 표준약관 기반 계약(비 PT 표준 T&C), 상위 계약문서 미입수 등 구조적 리스크를 여기에 기재.',
  '기재 시 RC3 / RC2 / RC1 선택 (Remark A+B)', '74')

assert len(R) == 100, len(R)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '해설 매뉴얼'

thin = Side(style='thin', color='D0D5DD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def kfont(sz=10, bold=False, color='000000'):
    return Font(name='맑은 고딕', size=sz, bold=bold, color=color)

ws.merge_cells('A1:I1')
c = ws['A1']; c.value = 'Standard LoA Risk Questionnaire 해설 매뉴얼 (HQ-003-05 V5.3 / 100항목)'
c.font = kfont(16, True, DB); c.alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:I2')
c = ws['A2']
c.value = ('목적: LoA(Offer Approval)를 처음 작성하는 담당자가 각 리스크 항목이 "무엇을, 왜 묻는지"를 이해하도록 돕는 해설서. '
           '근거: LoA Risk Questionnaire (HQ-003-05) V5.3 (2026-06-01) 질문 원문 + Explanatory Notes to LoA Risk Questionnaire (rev1). '
           'Explanatory Notes의 장·절 번호는 이 시트의 No.와 동일 체계. 자매 문서: Small_LoA_RiskQuestionnaire_해설매뉴얼 (HQ-003-09, TSP≤10m EUR용).')
c.font = kfont(9, False, GY); c.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[2].height = 40

ws.merge_cells('A3:I3')
c = ws['A3']
c.value = ('읽는 법: RC(Risk Class)는 escalation 등급 — N/A(비해당/Frame Approval 승인 완료) < RC3 < RC2 < RC1(CEO 승인, 최고 등급; 예외적으로 23.7만 CFO). '
           'Standard 폼에는 Small LoA의 "Region/Business" 열이 없음. 예시는 이해를 돕기 위한 일반 사례(가상)임. '
           '원칙: 리스크는 고객 원안이 아니라 "PT가 승인받고자 하는 조건(PT 수정 반영본)" 기준으로 체크하고, N/A 이외 등급은 반드시 코멘트란에 승인 요청 내용을 기재한다. '
           'Remark A = RC1 프로젝트면 최고 RC1 항목 전부에 코멘트 기재 / Remark B = MHI 관여 필요 여부 확인.')
c.font = kfont(9, False, GY); c.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[3].height = 46

HDR = ['No', '대분류', '소분류 (섹션)', '질문 원문 (English)', '한글 해석', '이 질문의 의미·목적', '예시 (일반 사례)', 'RC 판정 기준', 'Small LoA 대응']
hr = 5
for i, h in enumerate(HDR, 1):
    c = ws.cell(row=hr, column=i, value=h)
    c.font = Font(name='맑은 고딕', size=10, bold=True, color=WH)
    c.fill = PatternFill('solid', fgColor=DB)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws.row_dimensions[hr].height = 22

widths = [6, 12, 24, 52, 42, 55, 44, 30, 9]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = hr + 1
prev_cat = None
for (no, cat, sub, en, ko, mean, ex, rc, small) in R:
    if cat != prev_cat:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=cat)
        c.font = kfont(11, True, OR)
        c.fill = PatternFill('solid', fgColor=LIGHT)
        c.alignment = Alignment(vertical='center')
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = border
        ws.row_dimensions[row].height = 20
        row += 1
        prev_cat = cat
    vals = [no, cat.split()[0], sub, en, ko, mean, ex, rc, small]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=True,
                                horizontal='center' if i in (1, 9) else 'left')
        if i == 1:
            c.font = kfont(10, True, DB)
        elif i == 5:
            c.font = kfont(10, True)
        elif i == 8:
            c.font = kfont(9, False, 'CE0037' if 'RC1' in str(v) else TE)
        elif i == 9:
            c.font = kfont(9, False, GY)
        elif i == 4:
            c.font = Font(name='Arial', size=9, color='404040')
        elif i == 7:
            c.font = kfont(9, False, '404040')
        else:
            c.font = kfont(10)
    row += 1

ws.freeze_panes = 'A6'
ws.sheet_view.zoomScale = 90
ws.auto_filter.ref = f'A{hr}:I{row-1}'

wb.save(OUT)
print('saved:', OUT, 'items:', len(R))
