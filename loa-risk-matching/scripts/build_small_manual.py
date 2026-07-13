# -*- coding: utf-8 -*-
"""Small LoA Risk Questionnaire 해설 매뉴얼 xlsx 생성.
소스: HQ-003-09 V5.4 (LoA for Small Projects xlsm) + Explanatory Notes to LoA Risk Questionnaire (rev1)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = r'U:\新_海外営業部\Kim Yejun\1_POSCO\3_POSCO Servo valve\K1H\26년계약\260625_LoA준비\workflow\Small_LoA_RiskQuestionnaire_해설매뉴얼_20260714.xlsx'

DB = '0C2340'; OR = 'E87722'; GY = '97999B'; TE = '00587C'; WH = 'FFFFFF'
LIGHT = 'F2F4F7'

# (no, 대분류, 소분류, 영어원문, 한글해석, 의미·목적, 예시, RC기준, 해설§)
R = []
def a(*x): R.append(x)

T='Technical Risks'; C='Commercial Risks'; L='Legal Risks'; P='Compliance Risks'; O='Other Risks'

a(1, T, 'Project Scope',
  'Full Turnkey (Semi Turnkey including Turnkey Balance of Plants (TK BP) components like civils, building structure, cranes, building services) acc. to HQ-061-00.',
  '풀턴키 프로젝트인가? (세미턴키 + 토건·건물구조·크레인·건물설비 등 BoP 포함)',
  '풀턴키는 플랜트 전체(설계~기자재~토건~설치~시운전)를 계약자가 일괄 책임지는, 플랜트 사업에서 가장 리스크가 큰 계약 형태. "키만 돌리면 가동되는 상태"로 인도할 의무를 짐. 금액과 무관하게 최고 경영진 승인이 필요.',
  '제철소 설비 신설에서 건물 기초·크레인·유틸리티 공사까지 PT가 일괄 수행하는 계약.',
  '해당 시 RC1 (Remark A+B: TSP≥50m EUR 또는 적자 시 MHI 관여)', '1.1')
a(2, T, 'Project Scope',
  'Semi Turnkey (supply of technological equipment / components including advisory / supervisory services, installation execution) acc. to HQ-061-00.',
  '세미턴키인가? (기술 기자재 공급 + 자문/감리 서비스 + 설치 시공 수행)',
  '설치 시공까지 PT 책임이면 세미턴키. 기자재 공급+감리만 하는 Process Turnkey보다 시공 리스크(현장 인력·안전·공정)가 추가되므로 승인 대상. 리뱀프 프로젝트가 설치를 포함하면 2번과 3번에 동시 해당.',
  '이번 K1H처럼 서보밸브 시스템 공급 + E&C 서비스 + 설치(Erection & Installation)를 PT가 수행하는 계약.',
  '해당 시 RC3 (TSP≤25m EUR 기준; Small LoA는 TSP≤10m이므로 항상 RC3) + Remark B(MHI 관여 여부 확인)', '1.2')
a(3, T, 'Project Scope',
  'Modification / Revamp project.',
  '기존 설비의 개조/리뱀프 프로젝트인가?',
  '기존 설비·인터페이스·인수인계 지점에서 예측 불가능한 장애(도면과 실물 불일치, 노후·마모 등)가 발생할 수 있어 신설보다 리스크가 높다고 보는 항목.',
  '가동 중인 열연공장의 제어시스템 교체 — 기존 배선·기초가 as-built 도면과 다를 리스크.',
  '해당 시 RC3', '1.3')
a(4, T, 'Project Scope',
  'Prototype according to TLoA.',
  'TLoA 기준 프로토타입(상용 실적 없는 최초 적용 기술)이 포함되는가?',
  '검증되지 않은 기술의 최초 적용은 비용 초과·공기 지연으로 이어질 기술 리스크. 프로토타입 상세 리스크는 TLoA에서 CTO 승인을 받아야 하며, LoA에서도 최고 등급으로 상신.',
  '상용 레퍼런스가 없는 신형 서보 구동 방식을 이번 프로젝트에 처음 적용하는 경우.',
  '해당 시 RC1 + MHI 관여 (Remark A+B)', '1.4')
a(5, T, "Customer's approval / Right to modify layout",
  "The customer has a right to approve documents and the contract (i) does not stipulate a maximum period / number of review cycles for customer's approval; and (ii) does not stipulate deemed approval, if the customer fails to approve within such period / number of review cycles.",
  '고객이 문서 승인권을 갖는데 (i) 검토기간·검토횟수 상한이 없고 (ii) 기간 내 미승인 시 간주승인 조항도 없는가?',
  '고객이 검토를 지연시키거나 수정 요구를 반복하면 공정 전체가 끌려다님. 승인권 자체는 거부하지 않되, (1) 명확한 검토기한(예: 14일) 또는 1~2회 검토횟수 제한 + (2) 기한 도과 시 간주승인, 두 가지 안전장치를 요구하는 항목.',
  '이번 건: POSCO 일반약관 제7조 — 검토횟수 무제한 + 간주승인 없음 + 승인 후에도 매도인 무면책 → 해당 소지.',
  '안전장치 없으면 RC2', '3.1')
a(6, T, "Customer's approval / Right to modify layout",
  'The contract allows the customer to change / modify the layout or technical concept after the signature of the contract without adequate cost compensation and extension of time.',
  '계약 서명 후 고객이 적정한 비용 보상·공기 연장 없이 레이아웃/기술 컨셉을 변경할 수 있는가?',
  '"계약은 지켜져야 한다"가 원칙. 서명 후 일방 변경권은 사소한 변경이라도 큰 비용·공기 파급을 낳으므로, 반드시 보상+공기연장이 조건이어야 함. 조건 없는 일방 변경권은 고위험.',
  '이번 건: 일반약관 제9조⑧(설계서 불합치 시 증액 없는 대체·개조 명령)·제15조③(단순 물량변동 조정 불가) → 부분 해당 소지.',
  '해당 시 RC1 (Remark A)', '3.2')
a(7, T, 'Time schedule',
  'The project requires commencement of the project work (start engineering, start procurement of materials etc.) prior to conclusion of the contract without payment security.',
  '계약 체결 전에 지급 담보 없이 선(先)착수(엔지니어링 개시, 자재 조달 등)가 필요한가?',
  '계약이 최종 불발되면 이미 투입한 비용을 회수 못 할 수 있음. 원칙은 선급금 또는 담보 수령 후 착수. 선착수가 불가피하면 의무 부담액 규모에 따라 승인 등급이 갈림.',
  'LOI만 받은 상태에서 장납기 품목(대형 주단조품 등)을 먼저 발주하는 경우.',
  'PT 부담 ≤350k EUR & 장납기 발주 없음 → RC3 / >350k EUR 또는 장납기 발주 → RC2 / 비용 회수권 자체가 없음 → RC1', '4.1')
a(8, T, 'Time schedule',
  'Tough time schedule shorter than reference projects despite all possible mitigation measures. (Consult with SC.)',
  '모든 완화 조치를 강구해도 레퍼런스 프로젝트 대비 빠듯한(달성 곤란한) 공정인가?',
  '무리한 일정 수락은 지연·LD로 직결. 유사 실적 프로젝트를 벤치마크로 실현 가능성을 평가하고, 반드시 SC(공급망관리)와 공동 평가해야 하는 항목.',
  '동급 리뱀프의 실적 공기가 10개월인데 고객이 6개월을 요구하는 경우.',
  '해당 시 RC2 (해당 시 공정표 제출)', '4.2')
a(9, T, 'Time schedule',
  "Responsibility of PT for delays which are not caused by PT and/or PT's subcontractors or consortium partners (e.g. caused by customer, third party, Force Majeure).",
  'PT(하도급·컨소시엄 포함) 귀책이 아닌 지연(고객·제3자·불가항력 원인)까지 PT가 책임지는가?',
  '지연 원인은 ①PT 영역 ②고객 영역 ③양쪽 모두의 밖(제3자·FM) 세 가지. "고객 귀책 지연만 면책"이라는 문구는 ③영역 지연까지 PT 부담이 되는 함정. PT는 자기 영역 지연만 책임지는 것이 원칙.',
  '항만 파업으로 자재 입고가 늦어졌는데 면책 사유가 "고객 귀책"으로만 한정돼 있어 PT가 지체상금을 무는 경우.',
  '해당 시 RC1', '4.3')
a(10, T, 'PT work load (resources)',
  'Lack of required resources / personnel (in quantity, in quality, in time) which may jeopardize the fulfillment of contractual obligations (including factory load).',
  '계약 이행을 위협할 수 있는 자원·인력 부족(양·질·시기, 공장 부하 포함)이 있는가?',
  '계약 조건이 아니라 PT 내부 역량을 묻는 항목. 입찰 단계부터 자원계획이 필수이며, 부족이 계약 불이행으로 이어질 수 있는 수준이면 경영진에 보고.',
  '같은 시기에 대형 프로젝트가 겹쳐 시운전 엔지니어를 배정 못 할 우려가 있는 경우.',
  '해당 시 RC2 (Remark B)', '5.1')
a(11, T, 'Acceptance of plant',
  'Obligation to reach the performance parameters without adequate acceptance procedure (e.g. no repeated test runs allowed, testing period too short, etc.).',
  '성능 파라미터 달성 의무는 있는데 적정한 인수(성능시험) 절차가 없는가?',
  '인수는 위험이전·보증개시·잔금·현장철수가 걸린 최대 마일스톤. 절차의 최소요건 4가지 — ①객관적 판정 기준 ②충분한 시험 시간 ③결함 보완 기회 ④보완 후 재시험 허용 — 이 없으면 escalation.',
  '성능시험 1회 실패 시 재시험 없이 곧바로 불합격 처리하는 조건.',
  '최소요건 미충족 시 RC2', '8.1')
a(12, T, 'Acceptance of plant',
  'No "deemed acceptance clause" which covers at least: (i) delay due to reasons not attributable to PT; and (ii) plant is put into commercial operation by customer.',
  '(i) PT 귀책 아닌 사유로 성능시험을 못 하는 경우와 (ii) 고객이 플랜트를 상업 가동한 경우를 커버하는 간주인수 조항이 없는가?',
  '고객이 자기 생산을 우선해 성능시험을 계속 미루면, 고객은 새 설비로 돈을 벌면서 PT의 인수·대금만 무기한 지연되는 구조가 됨. 예정일 후 일정기간 경과 또는 상업운전 개시 시 인수로 간주하는 조항이 방어책.',
  '이번 건: POSCO 일반약관 전문에 간주인수 조항 전무(제17조②) → 해당.',
  '간주인수 없으면 RC2', '8.2')
a(13, T, 'Acceptance of plant',
  'No "deemed acceptance clause" for completion of the works which triggers (i) start of warranty period; (ii) transfer of risk; and (iii) payment for acceptance.',
  '간주인수가 있어도 (i) 보증기간 개시 (ii) 위험 이전 (iii) 인수 대금 지급을 트리거하지 않는가?',
  '간주인수는 실제 인수를 대체하는 것이므로 효과도 같아야 함. 세 가지 효과가 모두 따라오지 않으면 반쪽짜리. (위험이전이 Incoterms로 이미 더 일찍 일어나는 계약이면 (ii)는 논점 아님.)',
  '이번 건: 인수 지연 시 보증·위험·지급 기산이 전부 밀리는 구조(제17조⑥, 제31조③④) → 해당 소지.',
  '해당 시 RC2', '8.3')
a(14, T, 'Acceptance of plant',
  '"Deemed acceptance" payment is linked to a certificate that is required to be issued or signed by the customer (or a third party appointed by the customer).',
  '간주인수 대금 지급이 고객(또는 고객 지정 제3자)이 발행/서명하는 증명서에 연동되는가?',
  '고객이 인수를 방해하는 상황이라면 증명서 발행에도 협조하지 않는 것이 보통. 간주인수 대금을 L/C에서 회수하려면 고객 관여 없이 네고 가능한 서류 구조여야 함.',
  'L/C 네고 서류에 "고객 서명 인수증명서"가 들어 있어 고객이 서명을 미루면 대금 회수가 막히는 경우.',
  '해당 시 RC2 (지급담보가 아예 없어 지급이 고객 의사에 달린 경우도 RC2)', '8.4')
a(15, T, 'Spare parts',
  "PT's obligation for availability of (i) spare parts > 10 years, or (ii) spare parts for equipment with a life cycle ≤ 3 years > 3 years from acceptance, without right to provide alternative solution.",
  '대체품 제공 권리 없이 (i) 예비품 공급보장이 10년 초과이거나 (ii) 수명 3년 이하 장비(IT 등)에 3년 초과 보장인가?',
  '서브서플라이어의 단종·폐업·파산으로 장기 공급보장을 못 지킬 리스크. 기간을 관리 가능한 수준으로 제한하거나, 동등 사양 대체품을 제공할 권리를 확보해야 함.',
  '제어용 IT 하드웨어에 15년 예비품 공급 보장을 요구받는 경우.',
  '해당 시 RC3', '9.2')
a(16, T, 'Plant safety',
  "PT is responsible for declaring that anything beyond PT's scope of work is safe (e.g. customer's existing / old equipment).",
  'PT 공급범위 밖(고객 기존·노후 설비 등)의 안전까지 PT가 선언/책임지는가?',
  '안전사고는 경제적 책임뿐 아니라 임직원 형사책임까지 갈 수 있는 민감 영역. 통제할 수 없는 범위의 안전책임은 절대 회피 대상.',
  '개조 대상이 아닌 기존 라인 전체에 대한 안전 적합 확인서를 PT 명의로 발행하라는 요구.',
  '해당 시 RC1 (Remark A)', '10.1')
a(17, T, 'Local laws / Standards',
  "Challenging mandatory laws, regulations, governmental requirements and/or restrictions, norms and technical standards (e.g. DIN, ASME, country specific standards, customer's internal standards imposed in the contract) to be complied with by PT.",
  '준수 의무가 있는 강행 법규·규격·고객 내부표준이 평가·산정이 어렵거나 기술적으로 도전적인가?',
  '계약에 편입된 표준은 미준수 시 곧 계약위반(시정·배상·해지). 기술적·상업적으로 충분히 평가된 표준만 수용해야 하며, 평가 불가하거나 충족이 어려우면 escalation.',
  '이번 건: 구입사양서 §4.1 — KS + POSCO 내부표준(SZ, SZ 우선) 부과 + 일반약관 제39조 매수인 제반규정 준수 의무 → 해당 가능.',
  '해당 시 RC3', '11.1')
a(18, T, 'Remaining technical risks acc. TLoA',
  'Highest remaining General Risk acc. to TLoA - if SP is ≤ 1 m EUR.',
  'TLoA 잔여 General Risk(자원·일정·현지표준 등) 최고 등급 — SP 1m EUR 이하인 경우.',
  '기술 LoA(TLoA, HQ-059-00)의 General Risk 평가 결과를 상업 LoA에 전기하는 항목. Small LoA에서는 SP≤1m이면 완화되어 no/low는 N/A 처리.',
  'TLoA에서 일정 리스크가 medium으로 남음 → RC3 체크.',
  'Medium → RC3 / High → RC2 (No·Low는 N/A)', '12.1')
a(19, T, 'Remaining technical risks acc. TLoA',
  'Highest remaining General Risk acc. to TLoA - if SP is > 1 m EUR.',
  'TLoA 잔여 General Risk 최고 등급 — SP 1m EUR 초과인 경우.',
  '18번과 동일하나 SP>1m이면 완화 없이 low부터 escalation. TLoA 결과 없이는 판정 불가 — TLoA 결과서를 반드시 입수해 함께 보관해야 함.',
  '이번 건: SP 1811k EUR로 이 항목 적용 — TLoA 미입수로 판정 보류 중.',
  'Low → Region/Business / Medium → RC2 / High → RC1 (Remark A)', '12.1')
a(20, T, 'Remaining technical risks acc. TLoA',
  'Highest remaining Conceptual Risk acc. to TLoA - if SP is ≤ 1 m EUR.',
  'TLoA 잔여 Conceptual Risk(제품·플랜트 영역 컨셉 리스크) 최고 등급 — SP 1m EUR 이하.',
  'TLoA 두 번째 카테고리인 Conceptual Risk(제품/설비 구성 컨셉의 기술 리스크)를 전기하는 항목. 등급 체계는 18번과 동일.',
  'TLoA에서 신규 설비 구성에 conceptual risk medium 판정 → RC3.',
  'Medium → RC3 / High → RC2 (No·Low는 N/A)', '12.2')
a(21, T, 'Remaining technical risks acc. TLoA',
  'Highest remaining Conceptual Risk acc. to TLoA - if SP is > 1 m EUR.',
  'TLoA 잔여 Conceptual Risk 최고 등급 — SP 1m EUR 초과.',
  '19번과 동일 구조. (TLoA 세 번째 카테고리인 성능 파라미터 리스크는 50번에서 다룸.)',
  '이번 건: SP>1m으로 이 항목 적용 — TLoA 미입수로 판정 보류 중.',
  'Low → Region/Business / Medium → RC2 / High → RC1 (Remark A)', '12.2')

a(22, C, 'Price',
  'Total Sales Price (TSP). Overview of delegated approval authorities: LoA Intranet.',
  '총판매가격(TSP)이 얼마인가? (승인 권한 위임액과 비교)',
  '금액이 클수록 공급범위·인터페이스가 늘어 리스크 발생 확률과 파급이 커지므로, 가격 자체가 승인 등급을 결정. Small LoA는 지역별 위임액(LoA Intranet 게시) 기준으로 등급이 갈림.',
  '이번 건: TSP 1811k EUR — 위임액(LoA Intranet 매트릭스) 미확인으로 Region/Business vs RC3 판정 보류.',
  'TSP ≤1m → N/A / ≤위임액 → Region/Business / >위임액 → RC3', '13.1')
a(23, C, 'Project income',
  'Project income in relation to SP - if SP ≤ 1 m EUR.',
  '프로젝트 수익률(Project income/SP) — SP 1m EUR 이하인 경우.',
  'Project income은 EBIT에서 현금흐름 금융손익과 외국 소득세(CIT/WHT)를 반영한 최종 수익성 지표(HQ-031-03 Sales Calculation Summary에서 산출). 수익이 낮을수록 상위 승인.',
  'SP 0.8m, PI 4% → RC3.',
  '≥5% → N/A / <5% → RC3 (MS는 위임 기준)', '14.1')
a(24, C, 'Project income',
  'Project income in relation to SP - if SP > 1 m EUR.',
  '프로젝트 수익률 — SP 1m EUR 초과인 경우.',
  '23번과 동일 지표, SP>1m 구간의 등급표. 적자(<0%)면 RC1에 MHI까지 관여.',
  '이번 건: PI 8.6% → "≥5%" 구간, Region/Business 체크(기존 판정 유지).',
  '≥5% → Region/Business / ≥3% → RC3 / <3% → RC2 / <0% → RC1+MHI (Remark A+B)', '14.1')
a(25, C, 'Warranty obligations',
  'Base warranty period > 24 months after acceptance.',
  '기본 하자보증기간이 인수 후 24개월을 초과하는가?',
  '24개월이 업계 표준(PT 표준 T&C는 12개월). 기간이 길수록 "인도 시 이미 있던 결함인지, 고객 운영 탓인지" 입증이 어려워짐. 수리 시 보증이 연장·재기산되는 점도 감안.',
  '고객이 36개월 기본 보증을 요구하는 경우.',
  '>24~36개월 → RC2 / >36개월 → RC1 (실제 기간 기재)', '15.1')
a(26, C, 'Warranty obligations',
  'Explicit extension of or carve out from the base warranty period for steel structures and/or other equipment / items typically having a lifetime > 24 months.',
  '철구조물 등 수명이 긴 품목에 대해 기본 보증기간의 명시적 연장 또는 별도 예외(carve-out)가 있는가?',
  '수명이 긴 스코프(철구조물·코팅 등)에 대한 보증 연장 요구는 흔하지만 여전히 승인 대상 리스크.',
  '이번 건: 일반약관 제21조⑥ — 고의·과실 하자는 보증 만료 후에도 법정기간 내 배상청구 가능 → 사실상 연장/예외로 해당 소지.',
  '연장 ≤5년 → RC3 / >5년 또는 carve-out → RC2 (예외 내용·기간 기재)', '15.2')
a(27, C, 'Warranty obligations',
  'Base warranty period without (i) latest start and/or expiry date; or (ii) link to deemed acceptance. (At least one of these two precautions must be fulfilled.)',
  '기본 보증기간에 (i) 최종 개시/만료일(latest date)도 (ii) 간주인수 연동도 없는가?',
  '보증 개시점(예: 인수)이 고객 손에 달려 있으면 보증기간이 무한정 뒤로 밀림. "늦어도 계약발효 후 48개월에 만료" 같은 최종일 또는 간주인수 연동, 둘 중 하나는 반드시 확보.',
  '고객이 인수를 미루는 동안 보증 개시도 계속 밀려 총 노출 기간이 늘어나는 경우.',
  '두 장치 모두 없으면 RC2', '15.3')
a(28, C, 'Warranty obligations',
  'Warranty period for repaired / exchanged parts without latest date.',
  '수리·교체품에 대한 보증에 최종 종료일(cut-off)이 없는가? (체인 보증)',
  '수리→새 보증→또 수리→또 보증으로 이어지는 무한 "체인 보증" 리스크. 해당 부품만 재기산되면 small chain, 플랜트 전체가 재기산되면 big chain. 모든 연장 보증에 cut-off date를 두는 것이 목표.',
  '이번 건: 일반약관 제21조⑤ — 교체 시 "최초 납품과 같은 조건의 하자보증" → 재기산(rolling) 해당 소지.',
  'Small chain → RC3 / Big chain → RC2', '15.4')
a(29, C, 'Warranty obligations',
  'Customer decides how to rectify defects at its sole discretion.',
  '하자 구제 방법(수리/교체/배상)을 고객이 단독 재량으로 결정하는가?',
  '구제 방법 선택권은 PT가 갖거나 최소한 상호 합의가 원칙. 고객 단독결정이면 수리로 충분한 하자에 신품 교체 등 과중한 방법을 강요당할 수 있음.',
  '이번 건: 일반약관 제21조④ "매수인의 선택에 따라" 대물충당·보수·손해배상 → 해당(핵심 특이점).',
  'PT 거부권 없이 고객 단독결정이면 RC2', '15.5')
a(30, C, 'Warranty obligations',
  'Unusual warranty obligations (e.g. latent / hidden defect clause, implied warranties, design liability, availability guarantee, fit for purpose obligation, etc.).',
  '비통상 보증 의무(잠재하자 조항, 묵시적 보증, 설계책임, 가동률 보장, 목적적합성 보증 등)가 있는가?',
  '산정 불가능한 리스크를 낳는 특수 보증들. 가동률 보장은 판정 산식과 PT 무관 요인(예비품 미보유 정지 등) 배제를 반드시 명시해야 하고, fit-for-purpose는 "목적"의 구체 특정이 필수. Legal과 공동 평가 권장.',
  '이번 건: 구입사양서 §3.5.1 — 시운전 1개월 가동률 99.7% 이상 보장 → availability guarantee 해당 가능.',
  '해당 시 RC2', '15.6')
a(31, C, 'External consortium',
  'Joint and several liability towards the customer.',
  '(외부 컨소시엄) 고객에 대한 연대책임(joint & several)을 지는가?',
  '오픈 컨소시엄에서는 각 멤버가 전체 스코프에 연대책임 — 파트너가 불이행·파산하면 그 몫을 PT가 떠안음. 리스크 노출 파악을 위해 컨소시엄 지분(%)을 코멘트에 기재.',
  '토건 담당 파트너가 파산해 그 잔여 공사와 배상까지 PT가 인수하게 되는 경우.',
  '해당 시 RC3 (Remark B, 지분 기재)', '16.1')
a(32, C, 'External consortium',
  'Financial strength of external consortium partner(s) acc. to HQ-045-01 (in alignment with FS FIN).',
  '외부 컨소시엄 파트너의 재무 건전성은 어느 수준인가? (FS FIN 협의)',
  '연대책임 구조에서 파트너의 신용은 곧 PT의 리스크. FS FIN SF와 FRMC(금융리스크관리개념)에 따라 평가하고 결과를 코멘트에 기재.',
  '파트너 신용평가 결과 moderate → RC3 체크.',
  'Strong → N/A / Moderate → RC3 / Weak → RC2 / Insufficient → RC1 (Remark A+B)', '16.2')
a(33, C, 'External consortium',
  'Any other serious concerns regarding the capability of consortium partner(s).',
  '컨소시엄 파트너의 (재무 외) 기술·조직 역량에 중대한 우려가 있는가?',
  '노하우·생산능력·레퍼런스 등 수행 역량을 철저히 검증. GBU/Segment와 협의하고, 세미/풀턴키면 TK BP와도 정렬.',
  '파트너가 해당 설비 시공 실적이 전무한 경우.',
  '우려 있으면 RC1 (Remark A)', '16.3')
a(34, C, 'External consortium',
  'PT is liable for damages attributable to the external consortium partner(s) via socialization clause in the consortium agreement.',
  '컨소시엄 내부 계약의 손해 분담(socialization) 조항으로 파트너 귀책 손해(LD 등)를 PT가 분담하는가?',
  '원칙은 각자 자기 스코프 책임. socialization은 귀책 파트너의 LD를 무귀책 파트너도 합의 산식으로 분담하는 조항 — 무귀책자에게 부담 전가.',
  '파트너 지연으로 발생한 LD를 지분 비율대로 공동 부담하기로 한 컨소시엄 계약.',
  '해당 시 RC2', '16.4')
a(35, C, 'External consortium',
  'Unusual risks under / in the consortium agreement (e.g. liability of external consortium partner(s) towards PT is very limited, unfavourable weight clause, no proper applicable law and/or arbitration clause).',
  '컨소시엄 계약상 기타 비통상 리스크(파트너의 대(對)PT 책임 과소 제한, 불리한 weight clause, 준거법·중재 조항 부적정 등)가 있는가?',
  '파트너 책임한도가 너무 낮으면 고객이 PT에 청구한 손해를 파트너에게 구상 못 함. weight clause(자재 중량 변동 시 정산 조항)가 PT에 불리하게 설계된 경우 등도 포함.',
  '파트너의 대PT 책임한도가 자기 지분의 10%로 제한돼 있는 경우.',
  '해당 시 RC2', '16.5')
a(36, C, "Customer's changes / Variations",
  "The contract requires PT to proceed with change requests without written agreement (price, schedule, performance guarantees, etc.) and/or without acceptable procedures.",
  '가격·공정·성능보증 등에 대한 서면 합의 또는 수용 가능한 절차 없이 변경 지시에 착수해야 하는가?',
  '스코프 변경은 공기·비용·성능·보증에 연쇄 파급. 사전 서면 change order가 원칙이고, 최소한 "변경 실행 시 가격·공정 등의 적정 조정을 명문으로 보장"하는 절차가 있어야 수용 가능.',
  '고객 구두 지시로 먼저 시공하고 정산은 사후 협의하는 관행을 계약이 강제하는 경우.',
  '해당 시 RC2', '18.1')
a(37, C, 'Terms of payment / Payment security',
  '(i) Unclear definition of payment due date. (ii) Entitlement for payment is triggered by activities / items outside PT\'s control. (iii) No entitlement to claim interest for late payment.',
  '(i) 지급기일 불명확 (ii) 지급 트리거가 PT 통제 밖 (iii) 지연이자 청구권 없음 — 셋 중 해당이 있는가?',
  '각 기성의 지급기일과 네고 서류는 명확해야 함. "고객이 엔드유저에게 받으면 지급"(pay-when-paid)류는 PT가 이행해도 못 받는 구조. 지연이자는 계약 또는 준거법상 확보돼야 함(이슬람권은 대체 보상 필요).',
  '이번 건: 일반약관 제31조 — 검수 연동이지만 지급기한(60일)은 명시 → 비해당 판정 유지.',
  '해당 시 RC3', '19.1')
a(38, C, 'Terms of payment / Payment security',
  'If SP ≤ 1 m EUR: No or insufficient payment security and customer / parent company are not considered as creditworthy.',
  '(SP 1m EUR 이하) 지급담보가 없거나 불충분하고, 고객·모회사 신용도도 부족한가?',
  'Small LoA 완화 규정 — SP≤1m이면 고객(또는 모회사) 신용도가 충분하면 담보 없이도 N/A. 신용도까지 부족할 때만 escalation. 신용도는 HQ-045-00에 따라 FS FIN SF와 확인.',
  '신용 정보가 전혀 없는 신규 소형 고객에 무담보 후불 조건으로 공급.',
  '해당 시 RC2 (Remark B, 담보 유형/신용도 기재)', '19.3')
a(39, C, 'Terms of payment / Payment security',
  'If SP > 1 m EUR: (i) No or insufficient payment security (L/C, parent company guarantee, bank guarantee, PRI, ECA); and/or (ii) no or insufficient consequences in case issuance of payment security is delayed.',
  '(SP 1m EUR 초과) 지급담보(L/C·모회사보증·은행보증·PRI·ECA)가 없거나 불충분한가? 담보 발급 지연 시 구제수단(중지·해지권)이 없는가?',
  'PT는 대금 수령 전에 상당한 이행을 먼저 하므로 담보가 원칙. 신용도가 충분한 고객이면 미담보 리스크 금액에 따라 RC3/RC2, 신용도까지 부족하면 RC1. 최종 지급분이 L/C에서 빠져 있는 것도 "불충분"에 해당.',
  '이번 건: SP 1811k, POSCO 신용 기반 무담보 거래 — 미담보 노출액 기준으로 판정.',
  '(신용도 충분 시) 노출 ≤3m EUR → RC3 / >3m EUR → RC2 / 신용도 부족 → RC1 (Remark A+B)', '19.3')
a(40, C, 'Transfer of risk',
  'Transfer of risk to the customer takes place later than handover.',
  '위험 이전이 핸드오버(순수 공급은 인도조건 시점, 그 외는 인수 시점)보다 늦게 일어나는가?',
  '위험 이전 = 무과실 멸실·파손 책임이 고객으로 넘어가는 시점. 공급자는 최대한 이르게 가져가는 것이 유리. 순수 기자재 공급은 Incoterms 인도 시점, 시공 포함 계약은 인수 시점이 기준선.',
  '이번 건: 일반약관 제17조⑥ — 현장 반입 후에도 검사합격+인수까지 멸실·파손이 매도인 부담 → 해당 소지.',
  '기준보다 늦으면 RC2', '21.2')
a(41, C, 'Insurance',
  'Contractually required insurance coverage higher than limitation of liability and no clear precedence of limitation of liability cap (as aligned with FS FIN INS).',
  '계약상 요구 보험 부보액이 책임한도(LoL cap)를 초과하는데, 책임한도 우선 원칙이 명확하지 않은가?',
  '보험 부보액이 책임한도보다 크면 "실질 책임한도가 보험액 아니냐"는 해석 분쟁 소지. cap이 보험보다 우선함을 계약에 명시해야 하며, FS FIN INS와 정렬.',
  '책임한도 100%인데 계약이 200% 부보를 요구하고 우선순위 규정이 없는 경우.',
  '해당 시 RC3 (요구 부보액 기재)', '22.2')
a(42, C, 'Financial Risks',
  "Any deferred payment obligation in customer's favour > 3 months.",
  '고객에게 유리한 3개월 초과 지급 유예(deferred payment)가 있는가?',
  '유예 지급은 PT가 프로젝트를 선(先)금융하는 셈 — 현금흐름 악화와 추가 금융비용. 기준은 PT 인보이스 발행 시점부터 3개월.',
  '인수 확인 후 180일 뒤 지급하는 조건.',
  '해당 시 RC2 (Small LoA; Standard LoA에서는 RC1)', '23.5')
a(43, C, 'Country risks',
  'Substantial risk because of political or security situation of the country / region where the plant shall be installed and/or work / manufacturing / transport shall be performed.',
  '플랜트 설치국 또는 작업·제조·운송 수행국의 정치·치안 상황에 중대한 리스크가 있는가?',
  '안정적 환경에서의 수행이 견적의 전제. 무장충돌·소요·파업·행정기관의 자의적 조치 등 우려가 있으면 조건에 반영하고 보고.',
  '고위험국 현장 시공이 포함돼 주재원 안전·철수 리스크가 있는 경우.',
  '해당 시 RC1 (Remark A+B)', '24.1')

a(44, L, 'LD for delay',
  'No LD protection for delay.',
  '지연 지체상금(LD) 합의가 없는가?',
  'LD는 지연 손해를 정액화해 리스크에 상한을 만드는 "보호장치". 없으면 고객이 실손 전액(입증 기반)을 청구할 수 있음. 마일스톤 선정도 중요 — 풀턴키는 개별 납품이 아니라 start-up 기준이 적정.',
  '이번 건: LD 체계는 있음(제32조, 0.1%/일·상한 10/15%) — 단 제23조①로 LD 외 추가 손배가 열려 있어 보호가 불완전(48번 논점).',
  'LD 부재 시 RC2', '25.1')
a(45, L, 'LD for delay',
  'LD for delay > 10 % of SP.',
  '지연 LD 상한이 SP의 10%를 초과하는가?',
  'LD는 리스크 완화책이지만 상한이 크면 그 자체가 리스크. 통상 요율은 주당 0.5% 수준, 상한 10%가 업계 기준선.',
  '이번 건: 지연 LD 상한 10%(draft) → 비해당(종결).',
  '>10~15% → RC3 / >15% → RC2', '25.2')
a(46, L, 'LD for delay',
  'In general: LD cap for delay is reached in less than 10 / 7 weeks.',
  '지연 LD가 상한에 도달하는 기간이 10주 미만인가?',
  '요율이 높아 상한에 빨리 도달하면 이후 고객의 해지권 등 다음 단계 리스크가 조기에 열림. 상한 도달까지의 기간을 최대화하는 것이 PT 목표.',
  '이번 건: 0.1%/일 × 100일 = 10% → 약 14.3주로 여유 → 비해당(종결).',
  '<10주 → RC3 / <7주 → RC2 (도달 기간 기재)', '25.3')
a(47, L, 'LD for delay',
  'Only applicable in case of shutdown: LD cap for delay of start-up is reached in less than 30 / 10 days.',
  '(셧다운 수반 리뱀프에만 적용) 재기동 지연 LD 상한 도달이 30일 미만인가?',
  '가동 중 설비를 세우고 하는 리뱀프는 고객이 조업 재개에 사활을 걸므로 단기·고율 LD를 요구하는 것이 통례 — 그래서 완화된 별도 기준 적용.',
  '열연 셧다운 공사에서 LD가 20일 만에 상한 도달하는 조건 → RC3.',
  '<30일 → RC3 / <10일 → RC2 (도달 일수 기재)', '25.4')
a(48, L, 'LD for delay',
  'LD for delay are not sole and exclusive (financial) remedy.',
  '지연 LD가 유일·배타적 금전 구제(sole remedy)가 아닌가?',
  'sole remedy가 아니면 LD는 상한이 아니라 "최저 책임"이 됨 — LD를 물고도 그 위에 실손 배상이 얹힘. 준거법에 따라 당연히 보장되지 않으므로 명시 조항이 필요.',
  '이번 건: 일반약관 제23조① — 지체상금과 "별도로" 조업차질 손해배상 명문 → 해당(핵심 특이점).',
  'sole remedy 명시 없으면 RC2', '25.5')
a(49, L, 'LD for delay',
  '"Time is of the essence" clause / "Fixgeschäft" giving the customer an immediate termination right in case of delay without agreement of LD for delay.',
  '단기 지연도 중대 위반으로 보아 즉시 해지권을 주는 "기한 엄수 본질(time is of the essence)" 조항이 있는가?',
  '통상은 LD 상한 도달 전까지 고객이 해지 못 함. 이 조항이 있으면 하루 지연도 즉시 해지 사유가 될 수 있어, LD 우선 적용이 명확히 규정되지 않는 한 고위험.',
  '이번 건: 일반약관 제36조②1 — "납기 지연이 확실시"만으로 최고·유예 없이 해제 가능 → 유사 효과로 해당 소지.',
  '해당 시 RC1 (Remark A)', '25.6')
a(50, L, 'LD for non-performance',
  'No LD model which covers low, medium or high risk performance parameters acc. to TLoA.',
  'TLoA상 저·중·고 위험 성능 파라미터를 커버하는 성능 LD 모델이 없는가?',
  '성능 LD가 없으면 미달 시 절대적 make-good 의무(최소 수준까지 무한 보완) + 고객의 플랜트 거부권/실손배상으로 이어짐. LD 모델이 있거나 TLoA상 성능 리스크가 "no risk"면 N/A. LD 모델이 없으면 51·52번은 N/A 처리하고 코멘트 기재.',
  '이번 건: TLoA(HQ-059-00) 미입수로 판정 보류 중 — TLoA 결과 확보가 선결.',
  '(LD 모델 부재 시) TLoA Low → RC3 / Medium → RC2 / High → RC1 (Remark A+B)', '26.1')
a(51, L, 'LD for non-performance',
  'LD for shortfall of low, medium or high risk performance parameters acc. to TLoA are not sole and exclusive (financial) remedy.',
  '성능 미달 LD가 유일·배타적 금전 구제가 아닌가? (LD를 내고도 성능 달성 의무가 남는 등)',
  '48번의 성능판 논리. LD가 sole remedy가 아니면 최저 책임으로 변질. 성능 LD가 sole remedy이거나, TLoA 성능 리스크가 없거나, LD 모델 자체가 없으면(50번에 반영) N/A.',
  '이번 건: 일반약관 제20조② — PAT/FAT 불합격 시 제23조 배상책임 → 비배타적, 해당(핵심 특이점).',
  'TLoA Low → RC3 / Medium → RC2 / High → RC1 (Remark A)', '26.2')
a(52, L, 'LD for non-performance',
  'LD for shortfall of performance parameters > 10 % of the SP.',
  '성능 미달 LD 상한이 SP의 10%를 초과하는가?',
  '성능 LD도 합리적 범위 내 상한 설정이 업계 관행 — 45번의 성능판.',
  '성능 LD 상한 12% 요구 → RC3.',
  '>10~15% → RC3 / >15% → RC2 (요율 기재)', '26.3')
a(53, L, 'Liability',
  'Limitation of liability / indemnity obligations > 100 % of TSP.',
  '총 책임한도(LoL)가 TSP의 100%를 초과하는가? (한도 조항이 아예 없어 무제한인 경우 포함)',
  '책임한도는 프로젝트가 완전히 실패했을 때 회사 존립을 지키는 최후의 방어선. 100% cap이 업계 표준 관행. 계약에 cap 조항이 없으면 책임은 무제한 = 당연 해당. (정당한 거부에 따른 기수령 대금 반환은 손해배상이 아니므로 cap 계산에서 제외.)',
  '이번 건: POSCO 일반약관 전문에 책임한도 조항 부재 → 무제한 책임 구조로 해당 — 이번 계약 최우선 특이점.',
  '해당 시 RC1 (Remark A+B, 실제 한도 % 기재)', '27.1')
a(54, L, 'Liability',
  'Only applicable in case of external consortium: Limitation of liability / indemnity obligations > 100 % of SP.',
  '(외부 컨소시엄일 때만) 책임한도가 PT SP의 100%를 초과하는가?',
  '고객 계약의 cap은 전체 계약가 기준이므로, PT 몫(SP)으로 환산하면 100%를 넘게 됨. 계산례: TSP 110m, PT SP 70m, cap 100% → PT 노출 = 110×100%÷70 = 157%.',
  '오픈 컨소시엄에서 cap 100% of TSP인데 PT 지분이 60%인 경우 → PT 기준 167% → RC2.',
  '>100% → RC3 / >150% → RC2 / >170% → RC1 (실제 % 기재)', '27.2')
a(55, L, 'Liability',
  'Explicit exclusion / carve out in the contract of IP infringement; and/or breach of confidentiality; and/or insurance proceeds; from the overall limitation of liability.',
  'IP 침해·비밀유지 위반·보험금 수령분이 총 책임한도에서 명시적으로 제외(carve-out)되는가?',
  '이 3가지는 상대적으로 발생 확률이 낮거나(IP·비밀) 보험이 커버해(보험금) RC3 수준으로 수용 가능. 단 결과손해 배제(57번)는 유지돼야 하고, 보험금 제외는 FS FIN INS 사전 협의 필수.',
  '이번 건: cap 자체가 없어 "한도의 예외" 논의가 무의미 → 53번에 연동(비해당 유지).',
  '해당 시 RC3 (제외 항목·보험 커버 기재)', '27.3')
a(56, L, 'Liability',
  'Any other exclusions / carve outs (except gross negligence / wilful misconduct, personal injury, product liability or matters that cannot be limited under applicable law) from the overall limitation of liability leading to unlimited liability.',
  '55번 외의 기타 제외 사유(법상 배제 불가 항목 제외)로 인해 사실상 무제한 책임이 되는가?',
  '중과실·고의·인적 손해·제조물책임 등은 어차피 법상 제한 불가라 제외해도 승인 불요. 정당 거부 시 기수령 대금 반환 제외도 승인 불요. 그 밖의 제외는 cap을 무력화하므로 최고 등급.',
  '이번 건: cap 부재로 기본적으로 무제한 구조 → 53번에 연동(비해당 유지). 비밀정보 전액배상(제34조② 등)은 참고 기재.',
  '해당 시 RC1 (Remark A+B, 제외 항목 기재)', '27.4')
a(57, L, 'Liability',
  'No or insufficient exclusion of liability for: (i) indirect and consequential damages; and (ii) loss of revenue / profit; and (iii) loss of use / production.',
  '(i) 간접·결과손해 (ii) 수익/이익 상실 (iii) 사용/생산 손실에 대한 책임 배제가 없거나 불충분한가?',
  '지연·성능미달·기존설비 손상은 고객의 생산 손실→매출 손실로 이어지는데, 이 규모는 프로젝트 금액과 비교가 안 됨(제철소 조업정지 하루 손실이 계약가를 넘을 수 있음). 배제가 업계 표준이자 PT 기본 요건.',
  '이번 건: 일반약관 제23조① — "조업상 차질" 손해 배상을 명문으로 포함 + 배제 조항 부재 → 해당 — 이번 계약 최우선 특이점.',
  '배제 없거나 불충분하면 RC2 (Remark B, 누락 배제 기재)', '27.5')
a(58, L, 'Liability',
  'Aggregate cap (maximum) for LD for delay and shortfall of performance parameters > 10 % of SP.',
  '지연 LD + 성능 LD의 합산 상한(aggregate cap)이 SP의 10%를 초과하는가?',
  '개별 sub-cap(45·52번)과 별도로, 두 LD를 합친 총액에도 개별 합계보다 낮은 상한을 두는 것이 목표.',
  '이번 건: 지연 10% + PAC 지연 10%, 합산 15% 상한(제32조①3호 = draft 일치) → RC3 확정(어제 판정).',
  '>10~20% → RC3 / >20~30% → RC2 / >30% → RC1 (Remark A, 실제 % 기재)', '27.6')
a(59, L, 'Know-how / Confidentiality / IP',
  'The contract requires unconditional (i) handover or (ii) grant of right of use of proprietary information / sensible know-how (e.g. source code, drawings etc. of important nature) as defined in HQ-038-00 to the customer.',
  '소스코드·핵심 도면 등 독점정보/민감 노하우(HQ-038-00)의 무조건적 인도 또는 사용권 부여를 계약이 요구하는가?',
  '독점 노하우는 PT 장기 경쟁력의 핵심 자산. 인도를 피할 수 없다면 고객의 사용 목적을 운전·정비·오버홀·복구·수리로 한정하는 조건부여야 함. 무조건 이전은 최고 등급.',
  '이번 건: 일반약관 제38조② — 계약종료 후 3년 내 매도인 개량 IP도 매수인 귀속 → 해당 소지(핵심 특이점).',
  '조건 한정 불가 시 RC1 (Remark A+B)', '28.1')
a(60, L, 'Know-how / Confidentiality / IP',
  'Insufficient confidentiality obligations of customer regarding scope and/or duration (minimum 10 years from acceptance of the plant).',
  '고객의 비밀유지 의무가 범위 또는 기간(인수 후 최소 10년) 면에서 불충분한가?',
  'PT가 제공하는 모든 정보가 보호 대상이어야 하고 기간은 10년이 표준 타깃. PT 표준계약의 모델 조항 대비 의무를 크게 줄이는 수정은 승인 대상.',
  '비밀유지 기간을 3년으로 줄이자는 고객 요구.',
  '해당 시 RC2 (실제 기간 기재)', '28.2')
a(61, L, 'Know-how / Confidentiality / IP',
  'Excessive grant of rights to customer (e.g. right to use confidential information beyond operation, maintenance, overhaul, restoration and repair, or to pass it to third parties); Transfer of ownership / title of intellectual property to the customer.',
  '비밀정보의 과도한 사용권(목적 외 사용·개조·복제·제3자 제공) 또는 IP 소유권 이전이 있는가?',
  '고객의 사용권은 공급 설비의 운전·정비·오버홀·복구·수리로 한정이 원칙. 개량·복제 목적 사용, 제3자 제공, 소유권 이전은 회피 대상.',
  '이번 건: 제38조①② 권리귀속 구조 → 59번과 연동, 해당 소지.',
  '해당 시 RC1 (Remark A+B)', '28.3')
a(62, L, 'Know-how / Confidentiality / IP',
  'In case of violation of third party rights by PT, PT is not entitled to control the defense and the corrective actions (e.g. pay royalty, change design, settlement) and customer is entitled to decide.',
  'PT의 제3자 권리(IPR) 침해 시 방어·시정조치(로열티 지급/설계변경/화해)의 주도권이 PT가 아닌 고객에게 있는가?',
  '침해 대응은 상황별 최적 수단(로열티·설계변경·화해)을 고를 수 있어야 비용을 통제함. 고객이 결정권을 가지면 PT에 더 부담스러운 방법을 고를 수 있음.',
  '이번 건: 일반약관 제37조 — PT 자비 전면 부담 + 방어 주도권·고객 협력 조항 없음 → 해당 소지.',
  '고객 결정권이면 RC2', '28.4')
a(63, L, 'Termination of contract',
  'The contract allows customer to unilaterally terminate the contract for convenience without full reimbursement of PT cost.',
  '고객이 사유 없이(termination for convenience) 해지할 수 있으면서, PT 기발생 비용·경비의 전액 보전 의무가 없는가?',
  '임의 해지권 자체는 흔하지만 반드시 "해지로 인한 PT의 모든 비용·경비 보전"이 명시 조건이어야 함. 보전 없는 임의해지는 최고 등급.',
  '고객 사정(투자 보류)으로 해지하면서 기성분만 정산하고 해지 비용·이익은 불인정하는 조항.',
  '해당 시 RC1 (Remark A)', '29.2')
a(64, L, 'Dispute resolution / Law',
  'Lack of dispute resolution clause (for exclusive and final settlement of disputes) acceptable to Legal.',
  'Legal이 수용 가능한 (배타적·최종적) 분쟁해결 조항이 없는가?',
  '합의 없으면 분쟁은 관할 법원행 — 법치가 불안한 국가의 법원이면 치명적. ICC 중재 등 국제적 신뢰성 있는 중재를 지향하고, 수용 가능 여부는 반드시 PT Legal과 정렬.',
  '고객국 지방법원 전속관할 조항만 있는 계약.',
  'Legal 수용 불가 시 RC2 (Remark B: ICC 중재 지향)', '30.1')
a(65, L, 'Dispute resolution / Law',
  'Lack of applicable law acceptable to Legal.',
  'Legal이 수용 가능한 준거법 합의가 없는가?',
  '국제 프로젝트는 여러 법역이 얽히므로 명시적 준거법 합의가 필요. 중립법(스위스법, 영국법 등)을 지향하고 고객국 법은 회피. Legal과 정렬 필수.',
  '준거법 조항이 아예 없거나 고객국 법으로 강제되는 경우.',
  'Legal 수용 불가 시 RC2 (중립법 지향)', '30.2')
a(66, L, 'Export Control / Force Majeure',
  'Any concerns out of the EC checks acc. to HQ-052-00 (High Risk Country, military / nuclear related, sanctioned party, other Red Flag) and (i) no GECO approval received; or (ii) inability / concerns to comply with its preconditions.',
  '수출통제 체크(고위험국 수출, 군사·원자력 관련, 제재리스트 당사자, 기타 Red Flag)에서 우려가 있는데 GECO 승인이 없거나 승인 조건 준수가 어려운가?',
  '수출통제 위반은 수출 특권·라이선스 상실 등 회사 전체 리스크. 우려 항목이 하나라도 있으면 RECO를 통해 GECO 승인을 받아야 하며, 승인 없이는 최고 등급.',
  '엔드유저가 제재리스트 계열사로 확인됐는데 GECO 검토가 완료되지 않은 경우.',
  '해당 시 RC1 (Remark A)', '32.1')
a(67, L, 'Export Control / Force Majeure',
  'No proper Export Control Clause excluding PT for breach of contract in case of boycott / export bans / sanctions / revoking of export license. No prohibition of re-export. No notification obligation in case of change of control. No termination right of PT in case fulfillment is affected by sanctions.',
  '수출통제 면책 조항, 재수출 금지, 고객 지배구조 변경 통지의무, 제재 시 PT 해지권 — 이런 보호 장치가 계약에 없는가?',
  '수출 규제는 PT 통제 밖 사유이므로 그로 인한 불이행은 면책돼야 함. 고객의 재수출은 새로운 수출이라 금지 명문화 필요. 각 보호장치 부재 시 escalation.',
  '이번 건: POSCO 일반약관에 수출통제 관련 조항 전무 → 특별조건 확인 후 재판정 대상.',
  '보호장치 부재 시 RC3 (Remark B)', '32.2')
a(68, L, 'Export Control / Force Majeure',
  'For PT EU entities when customer is outside of EU or partner countries: No proper No-Russia / No-Belarus Clause combined with a termination right due to violation.',
  '(PT EU 법인이 EU/파트너국 밖 고객과 계약 시) No-Russia/No-Belarus 조항 + 위반 시 해지권이 없는가?',
  'EU 제재법(Council Regulation (EU) 833/2014)상 의무 조항. 적용 여부·예외는 EC 부서와 협의. 한국은 EU 파트너국 목록에 있어 한국 고객 직계약은 통상 미적용.',
  '이번 건: 고객이 한국(파트너국) → 통상 N/A. EC 확인 후 종결 가능.',
  '해당 시 RC1 (Remark A)', '32.3')
a(69, L, 'Export Control / Force Majeure',
  'Force Majeure defined too narrow. No time / schedule relief in case of FM. No termination right of PT if FM lasts longer than 6 months. Customer\'s right to terminate if FM lasts shorter than 3 months, or without obligation to pay for performed work.',
  'FM 정의가 협소한가? FM 시 공기 유예가 없는가? FM 6개월 초과 시 PT 해지권이 없는가? 고객이 3개월 미만 FM으로도 해지 가능한가? FM 해지 시 기성 대금 지급 의무가 없는가?',
  'FM은 양측 통제 밖 사유의 리스크 배분 장치. PT 표준 모델 조항 대비 보호가 부족한 수정(열거식 협소 정의, 공기 유예 없음 등)은 승인 대상.',
  '이번 건: 일반약관 제32조②2 — FM 효과가 지체상금 면제에만 한정되고 일반 면책·해지권 구조 없음 → 해당 소지.',
  '보호 부족 시 RC2 (Remark B)', '32.4')
a(70, L, 'Miscellaneous',
  "Other unusual clauses / obligations with major risks, e.g.: change of control in PT gives customer right to terminate; assignment of rights by customer without PT's consent; unreasonable notice periods; customer's right to reject PT's site personnel at any time without reason; customer's explicit right to set-off.",
  '기타 중대 리스크가 있는 비통상 조항 — 예: PT 지배구조 변경 시 고객 해지권, PT 동의 없는 고객의 권리 양도, 무리한 통지기간, 현장인력 무사유 거부권, 고객의 명시적 상계권.',
  '체크리스트가 못 담는 계약 특유의 리스크를 잡는 항목. 담당자(BPM 등)의 개별 판단으로 비통상 조항을 발굴해 기재.',
  '이번 건: 제43조(상계+연6% 지연손해금), 제36조③(해제 시 기지급 대금 반환), 제36조의2·3(입찰제한·담합 손배), 제35조의2(재무상태 통지) 등 다수 → 선별 기재 대상.',
  '내용에 따라 RC 판정(사안별 평가)', '33.1')
a(71, L, 'Miscellaneous',
  'Non-Compete Obligation.',
  '경업금지 약정(Non-Compete Obligation)이 있는가?',
  'MHI그룹 외부 회사에 "특정 제품·기술·서비스 사업을 다른 회사와 하지 않겠다"고 약속하는 것. 기간 불문 최고 등급이며, PT Head of Legal + MHI 추가 승인이 의무(HQ-008-02).',
  '고객이 "이 기술을 경쟁 제철사에 3년간 공급하지 않는다"는 조항을 요구.',
  '해당 시 RC1 + PT Legal 총괄·MHI 승인 필수 (Remark A)', '17.4')

a(72, P, 'Compliance',
  'Only applicable in case of involvement of a Sales Partner or Trading House acting as a Sales Partner: Result of the Compliance Assessment acc. to HQ-003-06.',
  '(판매 파트너/상사가 판매 파트너로 개입하는 경우만) HQ-003-06 컴플라이언스 평가 결과는?',
  '뇌물·계약 미승인 등 컴플라이언스 리스크를 신호등으로 평가. Green이면 비해당, Yellow면 RC2, Red면 LoA 상정 자체가 불가(HQ-003-00).',
  '대만 향 판매에 현지 에이전트가 개입해 Compliance Assessment 결과가 Yellow인 경우.',
  'Green/미개입 → N/A / Yellow → RC2 (Red는 상정 불가)', '34.1')
a(73, P, 'Compliance',
  'Is there a Business Partner other than Sales Partner / Trading House involved in the project? If "yes", please indicate the name and the CDD ID.',
  '판매 파트너/상사 외의 비즈니스 파트너가 프로젝트에 개입하는가? 있으면 이름과 CDD ID를 기재.',
  '컴플라이언스 실사(CDD, Compliance Due Diligence) 추적을 위한 정보 기재 항목. 유효한 CDD가 없으면 Local Compliance Officer에게 문의. (Explanatory Notes에 별도 해설 없음.)',
  '현지 설치 하도급이 아닌 컨설턴트·중개인이 개입하는 경우 그 CDD ID를 기재.',
  '정보 기재 항목 (RC 판정 없음)', '-')

a(74, O, 'Other Risks',
  '( to be inserted ) — catch-all for any other risk not addressed above, together with a proper risk class evaluation.',
  '위 항목들에 없는 기타 리스크 자유기재 — 리스크 내용과 함께 RC 등급을 스스로 평가해 기재.',
  '"캐치올" 항목. 어떤 체크리스트도 모든 리스크를 못 담으므로, BPM 또는 관여 부서 판단으로 경영 승인이 필요한 미분류 리스크를 여기에 서술하고 RC3/RC2/RC1을 선택.',
  '이번 건 기재 권고: ① POSCO 일반약관 기반 계약(PT 표준 T&C 아님 → Legal 서명 필요) ② 특별조건·특별약관 미입수(우선순위 상위 문서) ③ cap 부재+결과손해 포함(53·57번)이 최대 리스크.',
  '기재 시 RC3 / RC2 / RC1 선택 (Remark A)', '35.1')

assert len(R) == 74, len(R)

wb = openpyxl.Workbook()
ws = wb.active
ws.title = '해설 매뉴얼'

thin = Side(style='thin', color='D0D5DD')
border = Border(left=thin, right=thin, top=thin, bottom=thin)

def kfont(sz=10, bold=False, color='000000'):
    return Font(name='맑은 고딕', size=sz, bold=bold, color=color)

# 제목 블록
ws.merge_cells('A1:I1')
c = ws['A1']; c.value = 'Small LoA Risk Questionnaire 해설 매뉴얼 (HQ-003-09 V5.4 / 74항목)'
c.font = kfont(16, True, DB); c.alignment = Alignment(vertical='center')
ws.row_dimensions[1].height = 28

ws.merge_cells('A2:I2')
c = ws['A2']
c.value = ('목적: LoA를 처음 작성하는 담당자가 각 리스크 항목이 "무엇을, 왜 묻는지"를 이해하도록 돕는 해설서. '
           '근거: LoA for Small Projects (HQ-003-09) V5.4 질문 원문 + Explanatory Notes to LoA Risk Questionnaire (rev1, HQ-003-09 V5.2 기준). '
           '"해설 §"는 Explanatory Notes의 장·절 번호(Standard LoA HQ-003-05 항목번호와 동일 체계).')
c.font = kfont(9, False, GY); c.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[2].height = 40

ws.merge_cells('A3:I3')
c = ws['A3']
c.value = ('읽는 법: RC(Risk Class)는 escalation 등급 — N/A(비해당/Frame Approval 승인 완료) < Region/Business < RC3 < RC2 < RC1(CEO, 최고 등급). '
           '"예시" 열의 "이번 건"은 POSCO K1H Servo valve(YR4N/C01, POSCO 일반약관 240827 기반) 검토 사례. '
           '원칙: 리스크는 고객 원안이 아니라 "PT가 승인받고자 하는 조건(PT 수정 반영본)" 기준으로 체크하고, N/A 이외 등급은 반드시 코멘트란에 승인 요청 내용을 기재한다.')
c.font = kfont(9, False, GY); c.alignment = Alignment(vertical='top', wrap_text=True)
ws.row_dimensions[3].height = 40

HDR = ['No', '대분류', '소분류', '질문 원문 (English)', '한글 해석', '이 질문의 의미·목적', '예시', 'RC 판정 기준', '해설 §']
hr = 5
for i, h in enumerate(HDR, 1):
    c = ws.cell(row=hr, column=i, value=h)
    c.font = Font(name='맑은 고딕', size=10, bold=True, color=WH)
    c.fill = PatternFill('solid', fgColor=DB)
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    c.border = border
ws.row_dimensions[hr].height = 22

widths = [5, 12, 22, 52, 42, 55, 46, 30, 7]
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w

row = hr + 1
prev_cat = None
for (no, cat, sub, en, ko, mean, ex, rc, ref) in R:
    if cat != prev_cat:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=9)
        c = ws.cell(row=row, column=1, value=cat)
        c.font = kfont(11, True, OR)
        c.fill = PatternFill('solid', fgColor=LIGHT)
        c.alignment = Alignment(vertical='center')
        for col in range(1, 10):
            ws.cell(row=row, column=col).border = border
        ws.row_dimensions[row].height = 20
        row += 1
        prev_cat = cat
    vals = [no, cat.split()[0], sub, en, ko, mean, ex, rc, ref]
    for i, v in enumerate(vals, 1):
        c = ws.cell(row=row, column=i, value=v)
        c.border = border
        c.alignment = Alignment(vertical='top', wrap_text=True,
                                horizontal='center' if i in (1, 9) else 'left')
        if i == 1:
            c.font = kfont(10, True, DB)
        elif i == 5:
            c.font = kfont(10, True)
        elif i == 8:
            c.font = kfont(9, False, 'CE0037' if 'RC1' in str(v) else TE)
        elif i == 9:
            c.font = kfont(9, False, GY)
        elif i == 4:
            c.font = Font(name='Arial', size=9, color='404040')
        elif i == 7:
            c.font = kfont(9, False, '404040')
        else:
            c.font = kfont(10)
    row += 1

ws.freeze_panes = 'A6'
ws.sheet_view.zoomScale = 90
ws.auto_filter.ref = f'A{hr}:I{row-1}'

wb.save(OUT)
print('saved:', OUT, 'rows:', row - 1)
