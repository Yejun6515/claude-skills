# -*- coding: utf-8 -*-
"""Build the blank Business-trip master (header + formatting only, no day data).
Run once to (re)generate _templates/business_trip_master.xlsx.
The master is copied and filled by generate_schedule.py."""
import openpyxl, os
from openpyxl.styles import Font, Alignment, Border, Side

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUT = os.path.join(HERE, "_templates", "business_trip_master.xlsx")

COLS = {"B": 25.2, "C": 18.4, "D": 67.8, "E": 23.2, "F": 33.0}
medium = Side(style="medium")
HEAD_FONT = Font(name="Yu Gothic", size=11, bold=True)
A_C = Alignment(horizontal="center", vertical="center")

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Schedule"
ws.sheet_view.showGridLines = False
for col, w in COLS.items():
    ws.column_dimensions[col].width = w
ws.row_dimensions[1].height = 19.5
ws.row_dimensions[2].height = 19.5
for col, label in zip("BCDEF", ["Date", "Time", "Agenda", "Venue", "Notes"]):
    c = ws[f"{col}2"]
    c.value = label
    c.font = HEAD_FONT
    c.alignment = A_C
    c.border = Border(top=medium, bottom=medium)
# keep the companion sheets present (left blank, like the original template)
wb.create_sheet("Hotel")
wb.create_sheet("Visitors")

os.makedirs(os.path.dirname(OUT), exist_ok=True)
wb.save(OUT)
print("master built:", OUT)
