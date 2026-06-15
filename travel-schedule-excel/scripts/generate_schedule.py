# -*- coding: utf-8 -*-
"""Fill the blank Business-trip master from a JSON schedule and save an xlsx.

Usage:
    python generate_schedule.py <data.json> <output.xlsx>

data.json shape:
{
  "days": [
    {"iso": "2026-06-15", "venue": "Seoul", "notes": "",
     "am": "Move to Korea", "pm": "PTKR Monthly Meeting"},
    {"iso": "2026-06-19", "venue": "", "notes": "", "fullday": "Return to Japan"}
  ]
}

Rules (match the Business_trip template):
- One block per listed day = 2 rows (AM / PM). Days you omit are simply skipped.
- Date label is computed from "iso" -> e.g. "Jun/15(Mon)".
- Agenda (D) is merged across the 2 rows when the day has a single entry
  ("fullday", or only one of am/pm); split into AM/PM rows when both are given.
- Date / Venue / Notes are always merged across the 2 rows.
- Evening events normally go in Notes (e.g. "Dinner").
- All text must be ENGLISH.
"""
import openpyxl, os, sys, json, shutil
from datetime import date
from openpyxl.styles import Font, Alignment, Border, Side

HERE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
MASTER = os.path.join(HERE, "_templates", "business_trip_master.xlsx")

BODY_FONT = Font(name="Arial", size=11)
HEAD_FONT = Font(name="Yu Gothic", size=11, bold=True)
medium, hair = Side(style="medium"), Side(style="hair")
A_C  = Alignment(horizontal="center", vertical="center")
A_CW = Alignment(horizontal="center", vertical="center", wrap_text=True)
A_LW = Alignment(vertical="center", wrap_text=True)
A_L  = Alignment(vertical="center")


def date_label(iso):
    d = date.fromisoformat(iso)
    return f"{d:%b}/{d.day}({d:%a})"


def build(data_path, out_path):
    with open(data_path, encoding="utf-8") as fh:
        days = json.load(fh)["days"]
    shutil.copyfile(MASTER, out_path)
    wb = openpyxl.load_workbook(out_path)
    ws = wb["Schedule"]
    last = 2 + 2 * len(days)
    for i, d in enumerate(days):
        r = 3 + 2 * i
        ws.merge_cells(f"B{r}:B{r+1}"); ws[f"B{r}"].value = date_label(d["iso"]); ws[f"B{r}"].alignment = A_C
        ws[f"C{r}"].value, ws[f"C{r+1}"].value = "AM", "PM"
        ws[f"C{r}"].alignment = ws[f"C{r+1}"].alignment = A_C
        has_am, has_pm = bool(d.get("am")), bool(d.get("pm"))
        if d.get("fullday") or (has_am ^ has_pm):
            ws.merge_cells(f"D{r}:D{r+1}")
            ws[f"D{r}"].value = d.get("fullday") or d.get("am") or d.get("pm")
            ws[f"D{r}"].alignment = A_LW
        else:
            ws[f"D{r}"].value = d.get("am", ""); ws[f"D{r+1}"].value = d.get("pm", "")
            ws[f"D{r}"].alignment = ws[f"D{r+1}"].alignment = A_LW
        ws.merge_cells(f"E{r}:E{r+1}"); ws[f"E{r}"].value = d.get("venue", ""); ws[f"E{r}"].alignment = A_L
        ws.merge_cells(f"F{r}:F{r+1}"); ws[f"F{r}"].value = d.get("notes", ""); ws[f"F{r}"].alignment = A_CW
    # fonts + horizontal-only grid (hair between rows, medium around header and table bottom)
    for row in range(2, last + 1):
        for col in "BCDEF":
            c = ws[f"{col}{row}"]
            c.font = HEAD_FONT if row == 2 else BODY_FONT
            c.border = Border(top=medium if row in (2, 3) else hair,
                              bottom=medium if row in (2, last) else hair)
    wb.save(out_path)
    print("saved:", out_path)


if __name__ == "__main__":
    if len(sys.argv) != 3:
        sys.exit("usage: python generate_schedule.py <data.json> <output.xlsx>")
    build(sys.argv[1], sys.argv[2])
